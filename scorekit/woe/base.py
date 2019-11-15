# -*- coding: utf-8 -*-

from ..data import Data, DataSamples
from ..processor import MissingProcessor, GiniChecker, BusinessLogicChecker, WOEOrderChecker
from .._utils import color_background
import pandas as pd
#import math as m
import numpy as np
import matplotlib
import matplotlib.colors as colors
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import seaborn as sns
from sklearn.model_selection import cross_val_score, StratifiedKFold, train_test_split, GridSearchCV, PredefinedSplit
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import roc_auc_score, roc_curve, auc
#rom scipy.stats import chi2, chisquare, ks_2samp, ttest_ind
#import statsmodels.formula.api as sm
import warnings
from abc import ABCMeta, abstractmethod
#from sklearn.feature_selection import GenericUnivariateSelect, f_classif
from patsy import dmatrices
from statsmodels.stats.outliers_influence import variance_inflation_factor
import re
import ast
import os
import xlsxwriter
import openpyxl
from PIL import Image
import datetime
from dateutil.relativedelta import *
import gc
#import weakref
import copy

#warnings.simplefilter('ignore')

plt.rc('font', family='Verdana')
plt.style.use('seaborn-darkgrid')

gc.enable()


class WOE:
    '''
    WOE transformation of all/specified features of a specific DataSamples object (using train sample).
    Stores WOE parameters for each feature (FeatureWOEs) of the dataset it was fitted on.
    Can be used for trasformation of other Data objects.

    Parameters
    -----------
    datasamples: DataSamples object to calculate WOE
    features: list of features to calculate WOE for. If None then all the data.features will be processed.
    add_categorical: whether to transform features to categorical automatically or not
    interval_min_unique: if number of unique values of a feature is less than this parameter, the feature is considered categorical
    categorical: user-defined list of categorical features
    '''
    def __init__(self, datasamples, features=None, add_categorical=False, interval_min_unique=30, categorical = None):

        if categorical is None:
            categorical=[]

        self.data = datasamples.train
        self.datasamples = datasamples
        # feature_woes - {feature: FeatureWOE}
        self.feature_woes = {}
        self.categorical = categorical

        if features is None:
            cycle_features=list(datasamples.train.features)
        else:
            cycle_features=list(features)

        if add_categorical:
            for feature in cycle_features:
                unique_amount=len(datasamples.train.dataframe[feature].unique())
                field_type=datasamples.train.dataframe[feature].dtype
                if field_type==object or unique_amount<interval_min_unique:
                    categorical.append(feature)


        for feature in cycle_features:
            if feature in datasamples.train.features:
                self.feature_woes[feature] = FeatureWOE(self.datasamples, feature, categorical = (feature in categorical))
            else:
                print('No', feature, 'in datasamples.train.features. Skipping..')

        self.excluded_feature_woes = {}
        # for report
        self.stats = pd.DataFrame()
        self.auto_fit_parameters = {}




    def fit(self, features=None, alpha = 10, new_groups = True, alpha_range = None, n_folds = 5, max_depth = None, max_leaf_nodes=None,
            min_samples_leaf=None, simple=False, woe_adjust=0.5, missing_process='separate', scoring='neg_log_loss', plot_flag=True,
            out=None, sep=';'):
        '''
        Fits FeatureWOEs for each feature.

        Parameters
        -----------
        features: list of features to be fitted (if omitted, then all features in self.data will be fitted)
        alpha: regularization coefficient
        alpha_range: a range for optimal alpha calculation
        n_folds: number of fonds for WOE calculation
        max_depth: list of max depth values for Decision Tree GridSearch in WOE calculation
        max_leaf_nodes: list of leaves number values for Decision Tree GridSearch in WOE calculation
        min_samples_leaf: a value of minimal observations number for a leaf in Decision Tree GridSearch in WOE calculation
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        missing_process: way of handling missing values
            'separate' - keep missing values in a separate bin
            'worst' - assign missing values to the worst bin
            'nearest' - assign missing values to the nearest bin by WoE
            'worst_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the worst bin, otherwise keep separate
            'best_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the best bin, otherwise keep separate
            'nearest_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the nearest bin, otherwise keep separate
        scoring: a measure for cross-validation used used for optimal WOE splitting
        plot_flag: should WoE plots be printed or not
        out: a path for csv/xlsx output file (default YYYYMMDD_HHMMSS_groups_backup.xlsx)
        sep: the separator to be used in case of csv export
        '''

        if out is None:
            out=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")+'_groups_backup.xlsx'

        if features is None:
            cycle_features=list(self.feature_woes)
        else:
            cycle_features=list(features)

        times=[]

        for feature in cycle_features:
            if feature not in self.feature_woes and feature in self.excluded_feature_woes:
                self.feature_woes[feature] = self.excluded_feature_woes[feature]
                del self.excluded_feature_woes[feature]
            elif feature not in self.feature_woes:
                self.feature_woes[feature] = FeatureWOE(self.data, feature)

            if self.feature_woes[feature].data.dataframe[feature].isnull().sum() > 0:
                print ('Alarm! Missings in', feature, '!')
                #return None
            print (feature, 'processing started!')
            current_time=datetime.datetime.now()
            print('-- Starting at:                  '+str(current_time))
            times.append(current_time)
            if len(times)>1:
                time_per_feature=np.mean([times[i+1]-times[i] for i in range(len(times)-1)])
                print('-- Estimated time of completion: '+str(datetime.datetime.now()+time_per_feature*(len(cycle_features)-cycle_features.index(feature)))+' (average per feature = '+str(time_per_feature)+')')

            self.feature_woes[feature].fit(alpha = alpha, alpha_range = alpha_range, new_groups = new_groups, n_folds = n_folds, max_depth = max_depth,
                                             max_leaf_nodes=max_leaf_nodes, min_samples_leaf=min_samples_leaf, simple=simple, woe_adjust=woe_adjust,
                                             missing_process=missing_process, scoring=scoring, plot_flag=plot_flag)
            self.export_groups(out=out, sep=sep, features=cycle_features[:cycle_features.index(feature)+1])
            gc.collect()
            print ('---------------------------------------\n')



    def auto_fit(self, features=None, groups_range=range(6,1,-1), verbose=True, scoring = 'neg_log_loss', cv = 7,
                 alpha = 10, alpha_scoring = 'neg_log_loss', alpha_best_criterion = 'min', n_folds = 5, alpha_range = None,
                 alpha_classifier = LogisticRegression(random_state = 42), simple = None, woe_adjust = 0.5, plot_flag = True,
                 max_depth = None, min_samples_leaf = None, missing_process = 'separate', missing_min_part=0.05,
                 WOEM_on=True, WOEM_woe_threshold=0.05, WOEM_with_missing=False,
                 SM_on=True, SM_target_threshold=5, SM_size_threshold=100,
                 G_on=True, G_gini_threshold=5, G_gini_decrease_threshold=0.2, G_gini_increase_restrict=True, G_with_test=False,
                 BL_on=True, BL_conditions_dict=None,
                 WOEO_on=True, WOEO_dr_threshold=0.01, WOEO_correct_threshold=0.85, WOEO_miss_is_incorrect=True,
                 out=None, sep=';', to_continue=None, out_check=None):
        '''
        Auto-fits FeatureWOEs for each feature in WOE.feature_woes/input feature list.
        Values for groups amount are taken from groups_range in the specified order, feature is being fitted,
        then gini, business logic and WoE order checks are ran. If all checks are passed, then current binning stays,
        otherwise the feature is fitted with the next value of groups amount and checks are carried out again.
        If checks are failed for all values from groups_range, then feature is excluded from self.feature_woes.

        Parameters
        -----------
        features: list of features to auto-fit
        groups_range: list of integers that will be used as max_leaf_nodes value in FeatureWOE.fit method in the specified order
        verbose: if comments and graphs from checks should be printed
        out: a path for csv/xlsx output file (default YYYYMMDD_HHMMSS_auto_groups_backup.xlsx)
        sep: the separator to be used in case of csv export
        to_continue: file with backup groups to import and continue after the last feature from this file
        out_check: a path for xlsx output file with check results (default YYYYMMDD_HHMMSS_features_check.xlsx)

        Fit options:
        -----------
        alpha: regularization paramter, can be set by user or calculated automatically
        scoring: a measure for cross-validation used used for optimal WOE splitting
        cv: the numbder of folds for cross-validation used ufor optimal WOE splitting
        alpha_scoring: a measure for cross-validation used used for optimal alpha search
        alpha_best_criterion: if 'min', the lower alpha_scoring - the better, if 'max', the higher the measure - the better
        n_folds: number of folds for WOE calculation
        alpha_range: a range of values accesible for the optimal alpha choice
        alpha_classifier: a calssifier used for the optimal alpha search
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        plot_flag: should bins be plotted or not
        max_depth: the maximum of the DecisionTree depth used for optimal WOE splitting
        min_samples_leaf: the minimum of the DecisionTree leaf size used for optimal WOE splitting (one value)
        missing_process: way of handling missing values
            'separate' - keep missing values in a separate bin
            'worst' - assign missing values to the worst bin
            'nearest' - assign missing values to the nearest bin by WoE
            'worst_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the worst bin, otherwise keep separate
            'best_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the best bin, otherwise keep separate
            'nearest_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the nearest bin, otherwise keep separate
        missing_min_part: threshold for missing values part in case of missing_process=='worst_or_separate'

        Size merge options (SM):
        -----------------------
        SM_on: flag to turn on/off WoE merge process
        SM_target_threshold: min number of targets for group to not be considered small
        SM_size_threshold: min number of observations for group to not be considered small

        WoE merge options (WOEM):
        -----------------------
        WOEM_on: flag to turn on/off WoE merge process
        WOEM_woe_threshold: if woe difference between groups (neighboring groups for interval) is less then this threshold, then they are to be merged
        WOEM_with_missing: should woe difference with missing group also be checked

        Gini check options (G):
        ----------------------
        G_on: flag to turn on/off Gini check
        G_gini_threshold: gini on train and validate/95% bootstrap should be greater then this
        G_gini_decrease_threshold: gini decrease from train to validate/95% bootstrap deviation from mean to mean should be greater then this
        G_gini_increase_restrict: if gini increase should also be restricted
        G_with_test: should features be also checked on test sample (and be excluded in case of failure)

        Business logic check options (BL):
        ---------------------------------
        BL_on: flag to turn on/off business logic check
        BL_conditions_dict: adress for excel-file with business logic conditions (columns 'variable' and 'condition' are mandatory)

        WoE order check options (WOEO):
        ------------------------------
        WOEO_on: flag to turn on/off WoE order check
        WOEO_dr_threshold: if WoE order is not correct, then default rate difference between swaped bins is checked
        WOEO_correct_threshold: what part of checks on bootstrap should be correct for feature to pass the check
        WOEO_woe_adjust: woe adjustment factor (for Default_Rate_i formula)
        WOEO_miss_is_incorrect: is there is no data for a bin on bootstrap sample, should it be treated as error or not
        '''
        self.auto_fit_parameters = {'WOEM_on' : WOEM_on,
                                    'WOEM_woe_threshold' : WOEM_woe_threshold,
                                    'WOEM_with_missing' : WOEM_with_missing,
                                    'SM_on' : SM_on,
                                    'SM_target_threshold' : SM_target_threshold,
                                    'SM_size_threshold' : SM_size_threshold,
                                    'G_on' : G_on,
                                    'G_gini_threshold' : G_gini_threshold,
                                    'G_gini_decrease_threshold': G_gini_decrease_threshold,
                                    'G_gini_increase_restrict' : G_gini_increase_restrict,
                                    'G_with_test' : G_with_test,
                                    'BL_on' : BL_on,
                                    'BL_conditions_dict' : BL_conditions_dict,
                                    'WOEO_on' : WOEO_on,
                                    'WOEO_dr_threshold' : WOEO_dr_threshold,
                                    'WOEO_correct_threshold' : WOEO_correct_threshold,
                                    'WOEO_miss_is_incorrect' : WOEO_miss_is_incorrect
                                    }


        time_string=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        if out is None:
            out=time_string+'_auto_groups_backup.xlsx'

        if out_check is None:
            out_check=time_string+'_features_check.xlsx'

        all_log=pd.DataFrame(columns=['feature', 'iteration','result', 'reason'])
        all_gini=pd.DataFrame(columns=['feature', 'iteration', 'Train', 'Validate', 'Test']+['Bootstrap'+str(i) for i in range(len(self.datasamples.bootstrap))])
        all_bl=pd.DataFrame(columns=['feature', 'iteration', 'categorical', 'condition', 'fact', 'condition_result', 'overall_result'])
        all_woe=pd.DataFrame(columns=['feature', 'iteration', 'group', 'Train', 'Validate', 'Test']+['Bootstrap'+str(i) for i in range(len(self.datasamples.bootstrap))])
        all_er=pd.DataFrame(columns=['feature', 'iteration', 'group', 'Train', 'Validate', 'Test']+['Bootstrap'+str(i) for i in range(len(self.datasamples.bootstrap))])

        if features is None:
            cycle_features=list(self.feature_woes)
        else:
            cycle_features=list(features)

        if BL_conditions_dict is not None:
            conditions=pd.read_excel(BL_conditions_dict).set_index('variable')
            conditions['conditions']=conditions['conditions'].apply(lambda x: '' if (pd.isnull(x)) else x)

        times=[]

        if to_continue is not None and os.path.isfile(to_continue):
            print('Continuing after the last succesfully exported feature from '+str(to_continue)+':')
            if isinstance(to_continue, str):
                if to_continue[-4:]=='.csv':
                    imported_groups=pd.read_csv(to_continue)
                elif to_continue[-5:]=='.xlsx' or to_continue[-4:]=='.xls':
                    imported_groups=pd.read_excel(to_continue)
                else:
                    print('Unknown format of import file. Abort.')
                    return None
            self.exclude(cycle_features[:cycle_features.index(imported_groups.feature.tolist()[-1])+1])
            self.import_groups(to_continue, replace=True, exclude_rest=False, fit_flag=False, alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds = n_folds)
            cycle_features=cycle_features[cycle_features.index(imported_groups.feature.tolist()[-1])+1:]

        for feature in cycle_features:
            print ('---------------------------------------', feature, 'processing started! ---------------------------------------')
            current_time=datetime.datetime.now()
            print('-- Starting at:                  '+str(current_time))
            times.append(current_time)
            if len(times)>1:
                checked=len(times)-1
                time_per_feature=np.mean([times[i+1]-times[i] for i in range(checked)])
                print('-- Estimated time of completion: '+str(datetime.datetime.now()+time_per_feature*(len(cycle_features)-checked))+' (average per feature = '+str(time_per_feature)+')')
                print('-- Current progress: checked', checked, 'of', len(cycle_features),'(',len([x for x in cycle_features[:checked] if x in self.feature_woes]),'successful)')

            if BL_conditions_dict is not None:
                try:
                    current_condition=conditions.loc[feature]['conditions']
                except Exception:
                    current_condition=''
            else:
                current_condition=''

            if feature not in self.feature_woes and feature in self.excluded_feature_woes:
                self.feature_woes[feature] = self.excluded_feature_woes[feature]
                del self.excluded_feature_woes[feature]
            elif feature not in self.feature_woes:
                self.feature_woes[feature] = FeatureWOE(self.datasamples, feature)

            auto_fit_success, feature_log, feature_gini, feature_bl, feature_woe, feature_er = self.feature_woes[feature].auto_fit(self.datasamples,
                                                       alpha = alpha, groups_range=groups_range, scoring = scoring, cv = cv, alpha_scoring = alpha_scoring,
                                                       alpha_best_criterion = alpha_best_criterion, alpha_range = alpha_range, alpha_classifier = alpha_classifier,
                                                       n_folds = n_folds, simple = simple, woe_adjust = woe_adjust, plot_flag = plot_flag,
                                                       max_depth = max_depth, min_samples_leaf = min_samples_leaf, missing_process = missing_process, missing_min_part=missing_min_part,
                                                       SM_on=SM_on, SM_target_threshold=SM_target_threshold, SM_size_threshold=SM_size_threshold,
                                                       WOEM_on=WOEM_on, WOEM_woe_threshold=WOEM_woe_threshold, WOEM_with_missing=WOEM_with_missing,
                                                       G_on=G_on, G_gini_threshold=G_gini_threshold, G_gini_decrease_threshold=G_gini_decrease_threshold,
                                                       G_gini_increase_restrict=G_gini_increase_restrict, verbose=verbose, G_with_test=G_with_test,
                                                       BL_on=BL_on, BL_conditions=current_condition,
                                                       WOEO_on=WOEO_on, WOEO_dr_threshold=WOEO_dr_threshold, WOEO_correct_threshold=WOEO_correct_threshold, WOEO_miss_is_incorrect=WOEO_miss_is_incorrect,
                                                       out_check=out_check)
            all_log=all_log.append(feature_log, ignore_index=True)
            all_gini=all_gini.append(feature_gini, ignore_index=True).dropna(axis=1, how='all')
            all_bl=all_bl.append(feature_bl, ignore_index=True)
            all_woe=all_woe.append(feature_woe, ignore_index=True).dropna(axis=1, how='all')
            all_er=all_er.append(feature_er, ignore_index=True).dropna(axis=1, how='all')

            writer=pd.ExcelWriter(out_check, engine="openpyxl")
            all_log[['feature', 'iteration','result', 'reason']].to_excel(writer, sheet_name='Log', index=False)
            worksheet = writer.sheets['Log']
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 60

            gini_columns=[x for x in all_gini.columns if x not in ['feature', 'iteration']]
            all_gini.style.apply(color_background,
                                 mn=all_gini[gini_columns].min().min(),
                                 mx=all_gini[gini_columns].max().max(),
                                 cmap='RdYlGn', subset=pd.IndexSlice[:, gini_columns]).to_excel(writer, sheet_name='Gini by Samples', index=False)
            worksheet = writer.sheets['Gini by Samples']

            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 12
            for cn in range(3, worksheet.max_column+1):
                cl = openpyxl.utils.get_column_letter(cn)
                worksheet.column_dimensions[cl].width = 12
                for cell in worksheet[cl]:
                    cell.number_format = '0.00'
            worksheet.freeze_panes = worksheet['C2']

            all_bl.style.apply(self.color_result,
                               subset=pd.IndexSlice[:,['condition_result', 'overall_result']]).to_excel(writer, sheet_name='Business Logic', index=False)
            worksheet = writer.sheets['Business Logic']
            worksheet.column_dimensions['A'].width = 40
            for cn in range(2, worksheet.max_column+1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(cn)].width = 15
            worksheet.freeze_panes = worksheet['C2']

            woe_er_columns=[x for x in all_woe.columns if x not in ['feature', 'iteration', 'group']]
            woes_values=all_woe[woe_er_columns].values.reshape(-1,).tolist()
            all_woe.style.apply(color_background,
                                mn=np.mean(woes_values)-2*np.std(woes_values),
                                mx=np.mean(woes_values)+2*np.std(woes_values),
                                cmap='RdYlGn', subset=woe_er_columns).to_excel(writer, sheet_name='WoE by Samples', index=False)
            worksheet = writer.sheets['WoE by Samples']
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 12
            for cn in range(4, worksheet.max_column+1):
                cl = openpyxl.utils.get_column_letter(cn)
                worksheet.column_dimensions[cl].width = 12
                for cell in worksheet[cl]:
                    cell.number_format = '0.00000'
            worksheet.freeze_panes = worksheet['D2']

            er_values=all_er[woe_er_columns].values.reshape(-1,).tolist()
            all_er.style.apply(color_background,
                               mn=max([0,np.mean(er_values)-2*np.std(er_values)]),
                               mx=np.mean(er_values)+2*np.std(er_values),
                               cmap='RdYlGn_r', subset=woe_er_columns).to_excel(writer, sheet_name='Event Rate by Samples', index=False)
            worksheet = writer.sheets['Event Rate by Samples']
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 12
            for cn in range(4, worksheet.max_column+1):
                cl = openpyxl.utils.get_column_letter(cn)
                worksheet.column_dimensions[cl].width = 12
                for cell in worksheet[cl]:
                    cell.number_format = '0.000%'
            worksheet.freeze_panes = worksheet['D2']

            writer.save()

            if not auto_fit_success:
                if feature not in self.excluded_feature_woes:
                    self.excluded_feature_woes[feature]=self.feature_woes[feature]
                del self.feature_woes[feature]
            else:
                self.export_groups(out=out, sep=sep, features=cycle_features[:cycle_features.index(feature)+1])
                gc.collect()
            print ('---------------------------------------', feature, 'processing ended! ---------------------------------------')

            try:
                all_log.iteration = all_log.iteration.astype(int)
                all_gini.iteration = all_gini.iteration.astype(int)
                self.stats = all_log.merge(all_gini, on = ['feature', 'iteration'], how = 'left')
            except Exception:
                pass





    def color_result(self, x):
        '''
        TECH

        Defines result cell color for excel export

        Parameters
        -----------
        x: input values

        Returns
        --------
        color description for style.apply()
        '''
        colors_=[]
        for e in x:
            if e:
                colors_.append('background-color: green')
            else:
                colors_.append('background-color: red')
        if len(colors_)==0:
            colors_=pd.Series(colors_)
        return colors_


    def stats_manual_changes(self, feature):
        '''
        Adds to self.stats information on Gini changes after merges and splits.

        Parameters
        -----------
        feature: feature name
        '''
        #if 'Train_final' not in self.stats:
        #    self.stats['Train_final'] = copy.deepcopy(self.stats['Train'])
        #if 'Test_final' not in self.stats:
        #    self.stats['Test_final'] = copy.deepcopy(self.stats['Test'])

        self.stats.loc[self.stats.feature == feature, 'Train_final'] = copy.deepcopy(self.feature_woes[feature].gini_history['Train'][-1])
        self.stats.loc[self.stats.feature == feature, 'Test_final'] = copy.deepcopy(self.feature_woes[feature].gini_history['Test'][-1])

        if 'Validate' in self.stats:
            #if 'Validate_final' not in self.stats:
            #    self.stats['Validate_final'] = copy.deepcopy(self.stats['Validate'])
            self.stats.loc[self.stats.feature == feature, 'Validate_final'] = copy.deepcopy(self.feature_woes[feature].gini_history['Validate'][-1])


    def merge(self, feature, groups_list, alpha=None, simple=None, woe_adjust=None, n_folds=None, plot_flag=True, cv=None, no_backup=False, scoring='neg_log_loss'):
        '''
        Merges user-defined groups for a feature from self.data.

        Parameters
        -------------
        feature: feature name
        groups_list: [group1, group2] - the groups of th feature to be merged
        alpha: alpha parameter for the WOE of the new group
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        no_backup: should WoE merge be considered as manual change and be backed up or not
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        print (feature, ': merging', groups_list)
        if feature not in self.feature_woes:
            print ('Error! Feature', feature, 'is absent!')
        else:
            self.feature_woes[feature].merge(groups_list, alpha=alpha, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=plot_flag, cv=cv, no_backup=no_backup, scoring=scoring)
            # for report
            self.stats_manual_changes(feature)



    def merge_by_woe(self, woe_threshold=0.05, with_missing=True, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None, out=None, sep=';', no_backup=False, scoring='neg_log_loss'):
        '''
        For each feature merges all groups that are close by WOE (for interval features only neighboring groups and missing group are checked)

        Parameters
        -----------
        woe_threshold: if woe difference between groups (neighboring groups for interval) is less then this threshold, then they are to be merged
        with_missing: should woe difference with missing group also be checked
        alpha: alpha parameter for the WOE of the new group
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        out: a path for csv/xlsx output file
        sep: the separator to be used in case of csv export
        no_backup: should WoE merge be considered as manual change and be backed up or not
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        for feature in self.feature_woes:
            print (feature, 'processing started!')
            self.feature_woes[feature].merge_by_woe(woe_threshold=woe_threshold, with_missing=with_missing,
                             alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds = n_folds, plot_flag = plot_flag, cv=cv, out=out, sep=sep,
                             no_backup=no_backup, scoring=scoring)
            print ('---------------------------------------\n')
            self.stats_manual_changes(feature)



    def merge_by_size(self, target_threshold=5, size_threshold=100, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None, no_backup=False, scoring='neg_log_loss'):
        '''
        Merges small groups (by target or size) to the closest by WoE (for interval features only neighboring groups and missing group are checked)

        Parameters
        -----------
        target_threshold: min number of targets for group to not be considered small
        size_threshold: min number of observations for group to not be considered small
        alpha: alpha parameter for the WOE of the new group
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        no_backup: should WoE merge be considered as manual change and be backed up or not
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        for feature in self.feature_woes:
            print (feature, 'processing started!')
            self.feature_woes[feature].merge_by_size(target_threshold=target_threshold, size_threshold=size_threshold, alpha = alpha, simple=simple, woe_adjust=woe_adjust,
                                                     n_folds = n_folds, plot_flag=plot_flag, cv=cv, no_backup=no_backup, scoring=scoring)
            print ('---------------------------------------\n')
            self.stats_manual_changes(feature)



    def missing_unite(self, feature, missing_process = None, alpha = None, simple=None, woe_adjust=None, n_folds = None,
                      missing_min_part = None, plot_flag=True, cv=None, scoring='neg_log_loss'):
        '''
        Adds missing values to a bin based on chosen logic (from missing_process).

        Parameters
        -----------
        feature: name of the feature to process
        missing_process: way of handling missing values
            'separate' - keep missing values in a separate bin
            'worst' - assign missing values to the worst bin
            'nearest' - assign missing values to the nearest bin by WoE
            'worst_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the worst bin, otherwise keep separate
            'best_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the best bin, otherwise keep separate
            'nearest_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the nearest bin, otherwise keep separate
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        missing_min_part: threshold for missing values part in case of missing_process=='worst_or_separate', float between 0 and 1.
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        if feature not in self.feature_woes:
            print (feature, 'does not exist! Bye.')
        else:
            self.feature_woes[feature].missing_unite(missing_process = missing_process, alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds = n_folds,
                             missing_min_part = missing_min_part, plot_flag=plot_flag, cv=cv, scoring=scoring)
            self.stats_manual_changes(feature)



    def missing_separate(self, feature, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None, scoring='neg_log_loss'):
        '''
        Separates missing values from any bin they were assigned to.

        Parameters
        -----------
        feature: name of feature to process
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        if feature not in self.feature_woes:
            print (feature, 'does not exist! Bye.')
        else:
            self.feature_woes[feature].missing_separate(alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds = n_folds, plot_flag=plot_flag,
                                                        cv=cv, scoring=scoring)
            self.stats_manual_changes(feature)


    def split(self, feature, group=None, to_add=None, scoring = 'neg_log_loss', alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None):
        '''
        Splits feature' group, only one group per feature.

        Parameters
        -------------
        feature: feature name
        group: a group to split, integer
        to_add: in case of interval - a user-defined bound for the split (the intermediate bound of the interval), only for ordered features; in case of categorical -  a user-defined new group of values, consists of values from the group to split and the other values of the group will be separated. Only for categorical features. Example: group = [1, 2, 4, 6, 9], new_group = [1, 2, 9] => the two new groups will be [1, 2, 9], [4, 6]. For the same result we could set new_group parameter = [4, 6]

        scoring: a scoring metric to use
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        '''
        if feature not in self.feature_woes:
            print (feature, 'does not exist! Skipping...')
        else:
            if group is None:
                if feature in self.categorical:
                    for g in self.feature_woes[feature].groups:
                        if to_add[0] in self.feature_woes[feature].groups[g]:
                            group = g
                            break
                else:
                    for g, v in self.feature_woes[feature].groups.items():
                        if to_add >= v[0] and to_add < v[1]:
                            group = g
                            break

            self.feature_woes[feature].split(group, to_add, scoring = scoring, alpha = alpha, simple = simple, woe_adjust = woe_adjust, n_folds = n_folds, plot_flag = plot_flag, cv = cv)
            self.stats_manual_changes(feature)


    def show_history(self, feature, figsize=(6,5)):
        '''
        Shows groups and gini history, producing WoE and Gini graphs

        Parameters
        -----------
        feature: feature to process
        figsize: size for one plot; size for complex graph will be calculated as figsize[0]*2 for width and
            figsize[1]*number of rows for height
        '''
        if feature not in self.feature_woes:
            print (feature, 'not in included features! Searching excluded features...')
            if feature not in self.excluded_feature_woes:
                print (feature, 'not in excluded features! Skipping...')
            else:
                self.excluded_feature_woes[feature].show_history(figsize=figsize)
                print ('Remember to include', feature, 'before working with it.')
        else:
            self.feature_woes[feature].show_history(figsize=figsize)



    def rollback(self, feature, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, iteration=None):
        '''
        Rolls back the last operation.

        Parameters
        -----------
        feature: feature to process
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        iteration: number of groups iteration to return to (if None, then rollback to the previous iteration)
        '''
        if feature not in self.feature_woes:
            print (feature, 'does not exist! Skipping...')
        else:
            self.feature_woes[feature].rollback(alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds = n_folds, plot_flag=plot_flag, iteration=iteration)



    #edited 03.09.2018 by Anna Goreva: changed bounds to groups
    # Anna Goreva set original_values to True 10/08/2018
    def transform(self, data, original_values=True, calc_gini=True, keep_essential=False, add_fields=None, not_found='worst', features=None, with_excluded=False):
        '''
        Transforms a Data or DataSamples object according to WOE parameters fitted. Can be used only after .fit().

        Parameters
        ------------
        data: Data or DataSamples object to transform
        original_values: should we calculate WoE by original values (not transformed to pre_woes) for categorical features
        calc_gini: should gini values be calculated on the dataframe of data object or not
        keep_essential: should only not transformed features, target and weights be kept in output Data object or not
        add_fields: list of fields to be kept in addition to transformed features, target and weights in case of keep_essential=True
        not_found: what woe to return for unknown categorical values; 'worst' - the lower woe, 'missing' - woe of group that contains missings, int - number of certain group
        features: list of features to transform (if None, then self.feature_woes are used)
        with_excluded: should excluded features also be used for transformation

        Returns
        ----------
        transformed Data or DataSamples object
        '''

        if add_fields is None:
            add_fields=[]

        if features is None:
            cycle_features=list(self.feature_woes)
            if with_excluded:
               cycle_features=cycle_features+list(self.excluded_feature_woes)
        else:
            cycle_features=features.copy()

        #corrected (+.copy()) 13.08.2018 by Yudochev Dmitry
        if type(data)==Data:
            if keep_essential:
                to_keep=cycle_features+add_fields+([data.target] if data.target is not None else [])
                if data.weights is not None:
                    to_keep.append(data.weights)
                data_to_transform = Data(data.dataframe[to_keep].copy(),  data.target, [feature + '_WOE' for feature in cycle_features], weights = data.weights, name=data.name)
            else:
                data_to_transform = Data(data.dataframe.copy(),  data.target, [feature + '_WOE' for feature in cycle_features], weights = data.weights, name=data.name)
            for feature in cycle_features:
                #added variant with original values for categorical features
                if feature in self.feature_woes:
                    data_to_transform.dataframe[feature + '_WOE'] = self.feature_woes[feature].set_avg_woes(
                                                                        data=data_to_transform.dataframe[feature],
                                                                        groups=self.feature_woes[feature].groups,
                                                                        woes=self.feature_woes[feature].woes,
                                                                        original_values = original_values and self.feature_woes[feature].categorical,
                                                                        not_found=not_found)
                elif with_excluded:
                    data_to_transform.dataframe[feature + '_WOE'] = self.excluded_feature_woes[feature].set_avg_woes(
                                                                        data=data_to_transform.dataframe[feature],
                                                                        groups=self.excluded_feature_woes[feature].groups,
                                                                        woes=self.excluded_feature_woes[feature].woes,
                                                                        original_values = original_values and self.excluded_feature_woes[feature].categorical,
                                                                        not_found=not_found)
                    #data_to_transform.dataframe[feature].apply(
                        #lambda x: self.feature_woes[feature].change_to_woe(
                            #x, self.feature_woes[feature].groups, self.feature_woes[feature].woes,
                                #original_values = original_values and self.feature_woes[feature].categorical, not_found=not_found))

            if keep_essential:
                data_to_transform.dataframe=data_to_transform.dataframe.drop(cycle_features, axis=1)
            if calc_gini:
                data_to_transform.ginis=data_to_transform.calc_gini().copy()
            return data_to_transform

        elif type(data)==DataSamples:
            datasamples_to_transform = copy.deepcopy(data)
            samples=[datasamples_to_transform.train, datasamples_to_transform.validate, datasamples_to_transform.test, datasamples_to_transform.bootstrap_base]
            for sample in samples:
                if sample is not None:
                    if keep_essential:
                        to_keep=cycle_features+add_fields+([sample.target] if sample.target is not None else [])
                        if sample.weights is not None:
                            to_keep.append(sample.weights)
                        sample.dataframe=sample.dataframe[to_keep]
                    sample.features=[feature + '_WOE' for feature in cycle_features]
                    for feature in cycle_features:
                        #added variant with original values for categorical features
                        sample.dataframe[feature + '_WOE'] = self.feature_woes[feature].set_avg_woes(data=sample.dataframe[feature],
                                                                                                     groups=self.feature_woes[feature].groups,
                                                                                                     woes=self.feature_woes[feature].woes,
                                                                                                     original_values = original_values and self.feature_woes[feature].categorical,
                                                                                                     not_found=not_found)
                            #sample.dataframe[feature].apply(
                                #lambda x: self.feature_woes[feature].change_to_woe(
                                    #x, self.feature_woes[feature].groups, self.feature_woes[feature].woes,
                                        #original_values = original_values and self.feature_woes[feature].categorical, not_found=not_found))
                    if keep_essential:
                        sample.dataframe=sample.dataframe.drop(cycle_features, axis=1)
                    if calc_gini:
                        sample.ginis=sample.calc_gini().copy()


            return datasamples_to_transform



    def exclude(self, to_exclude):
        '''
        Excludes a feature or a list of features from self.feature_woes (and adds it/them to self.excluded_feature_woes)

        Parameters
        ------------
        to_exclude: a feature name or a list of feature names

        '''

        print('Excluding', to_exclude)
        if isinstance(to_exclude, list):
            not_in_feature_woes=[x for x in to_exclude if x not in self.feature_woes]
            if len(not_in_feature_woes)>0:
                print('No', not_in_feature_woes, 'in self.feature_woes. Check the names of the features to exclude.')
                return None
            self.excluded_feature_woes.update({x:self.feature_woes[x] for x in to_exclude})
            self.feature_woes = {x:self.feature_woes[x] for x in self.feature_woes if x not in to_exclude}
        elif isinstance(to_exclude, list)==False:
            if to_exclude not in self.feature_woes:
                print('No', to_exclude, 'in self.feature_woes. Check the name of the feature to exclude.')
                return None
            self.excluded_feature_woes[to_exclude]=self.feature_woes[to_exclude]
            del self.feature_woes[to_exclude]



    def include(self, to_include):
        '''
        Includes a feature or a list of features into self.feature_woes from self.excluded_feature_woes and
        fits it/them

        Parameters
        ------------
        to_include: a feature name or a list of feature names

        '''

        print('Including', to_include)
        if isinstance(to_include, list):
            not_in_excluded_feature_woes=[x for x in to_include if x not in self.excluded_feature_woes]
            if len(not_in_excluded_feature_woes)>0:
                print('No', not_in_excluded_feature_woes, 'in self.excluded_feature_woes. Check the names of the features to include.')
                return None
            self.feature_woes.update({x:self.excluded_feature_woes[x] for x in to_include})
            self.excluded_feature_woes = {x:self.excluded_feature_woes[x] for x in self.excluded_feature_woes if x not in to_include}
        elif isinstance(to_include, list)==False:
            if to_include not in self.excluded_feature_woes:
                print('No', to_include, 'in self.excluded_feature_woes. Check the name of the feature to include.')
                return None
            self.feature_woes[to_include]=self.excluded_feature_woes[to_include]
            del self.excluded_feature_woes[to_include]



    def plot_gini(self, features = None):
        '''
        Plots Gini history for features

        Parameters
        -----------
        features: list of features to plot Gini for
        '''
        if isinstance(features, str):
            features = [features]
        elif not isinstance(features, list):
            features = list(self.feature_woes.keys())

        for feature in features:
            if feature not in list(self.feature_woes.keys()):
                print ('Achtung! Das Merkmal', feature, 'ist nicht vorhanden!')
            else:
                self.feature_woes[feature].plot_gini()



    def plot_woe(self, features = None):
        '''
        Plots WOE groups and values

        Parameters
        ------------
        features: list of features to plot WOE for
        '''
        if isinstance(features, str):
            features = [features]
        elif not isinstance(features, list):
            features = list(self.feature_woes.keys())

        for feature in features:
            if feature not in list(self.feature_woes.keys()):
                print ('Achtung! Das Merkmal', feature, 'ist nicht vorhanden!')
            else:
                self.feature_woes[feature].plot_woe()



    # edited by Anna Goreva Oct-05-10: added parameter 'sep'
    def export_groups(self, out=None, sep = ';', features=None, verbose=False):
        '''
        Transforms self.groups, self.categorical and self.missing_group to dataframe.

        Parameters
        ----------
        features: a list of features to export (if None, then all features are exported)
        out: a string with the path to exported file (if None, then no export)
        sep: separator for csv
        verbose: should any comments be printed out

        Returns
        ----------
        dataframe with binning information
        '''

        if features is None:
            cycle_features=list(self.feature_woes)
        else:
            cycle_features=list(features)

        out_df=pd.DataFrame()

        for feature in cycle_features:
            if feature in self.feature_woes:
                out_df=out_df.append(self.feature_woes[feature].groups_to_dataframe(), ignore_index=True)
            else:
                if verbose:
                    print('No', feature, 'in self.feature_woes. Skipping..')

        out_df=out_df[['feature', 'categorical', 'group', 'values', 'missing', 'woe', 'changed']]

        if out is not None:
            if out[-4:]=='.csv':
                out_df.to_csv(out, index=False, sep = sep)
            elif out[-5:]=='.xlsx' or out[-4:]=='.xls':
                out_df.to_excel(out, index=False)
            else:
                print('Unknown format for export file. Use .csv or .xlsx. Skipping export.')

        return out_df



    # Oct-04-2018
    # Anna Goreva: the method does not work! Must be fixed
    def import_groups(self, df_in, exclude_rest=False, replace=False, features=None, alpha = None, simple=None, woe_adjust=None, n_folds = None,
                      plot_flag=True, print_flag=True, fit_flag=False, sep=';', log_in=None):
        '''
        Sets groups, categorical and missing_group values from dataframe and calculates woe (by fit) for each feature in input dataframe/list.

        Parameters
        ----------
        df_in: a xlsx or csv file with 'categorical', 'group', 'values' and 'missing' fields
        exclude_rest: should the features not specified in df_in be excluded (to self.excluded_feature_woes) or not
        replace: should current binnings for self.feature_woes be replaced or not (if not, then only binnings for self.excluded_feature_woes will be replaced)
        features: a list of features to import groups for
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        print_flag: should bins and woes be printed or not
        fit_flag: should woes be calculated or taken from input dataframe
        sep: a separator value in case of csv import
        log_in: an xls file with auto fit log ('feature', 'iteration', 'result', 'reason' fields) 
            and gini values ('feature', 'iteration', 'Train', 'Validate' fields)
        '''

        if isinstance(log_in, str):
            if log_in[-5:]=='.xlsx' or log_in[-4:]=='.xls':
                gini_in=pd.read_excel(log_in, sheet_name=1)
                log_in=pd.read_excel(log_in, sheet_name=0)
                log_in.iteration = log_in.iteration.astype(int)
                gini_in.iteration = gini_in.iteration.astype(int)
                self.stats = log_in.merge(gini_in, on = ['feature', 'iteration'], how = 'left')
            else:
                print('Unknown format of import file for log and gini. Abort.')
                return None

        if isinstance(df_in, str):
            if df_in[-4:]=='.csv':
                df_in=pd.read_csv(df_in, sep=sep)
            elif df_in[-5:]=='.xlsx' or df_in[-4:]=='.xls':
                df_in=pd.read_excel(df_in)
            else:
                print('Unknown format of import file. Abort.')
                return None

        if features is None:
            cycle_features=list(df_in.feature.unique())
        else:
            cycle_features=list(features)

        for feature in cycle_features:
            if feature in self.feature_woes:
                if replace:
                    print('Replacing binning for', feature, '..')
                    self.feature_woes[feature].groups_from_dataframe(df_in[df_in['feature']==feature],
                                                                     alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds = n_folds,
                                                                     plot_flag=plot_flag, print_flag=print_flag, fit_flag=fit_flag)
                else:
                    print('Feature', feature, 'was not excluded (to import binning set replace=True). Skipping..')
            elif feature in self.excluded_feature_woes:
                print('Feature', feature, 'was excluded. Returning to included and replacing binning..')
                self.feature_woes.update({feature:self.excluded_feature_woes[feature]})
                del self.excluded_feature_woes[feature]
                self.feature_woes[feature].groups_from_dataframe(df_in[df_in['feature']==feature],
                                                                 alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds = n_folds,
                                                                     plot_flag=plot_flag, print_flag=print_flag, fit_flag=fit_flag)
            else:
                print('No', feature, 'in current WOE object. Skipping..')

        if exclude_rest:
            print('Excluding features not specified in input data:')
            features_to_check=list(self.feature_woes)
            for feature in features_to_check:
                if feature not in cycle_features:
                    print('.. dropping', feature)
                    self.excluded_feature_woes[feature]=self.feature_woes[feature]
                    del self.feature_woes[feature]
#----------------------------------------------------------------------------------------------------------





class FeatureWOE:
    '''
    WOE-transformation for a single factor, contains WOEs and bounds of intervals.
    For each factor a new FeatureWOE object should be created.
    '''

    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    def __init__(self, datasamples, feature, categorical=False, alpha = 10, simple=False, n_folds=5, woe_adjust=0.5, cv=7):
        '''
        Initialization.

        Parameters
        -----------
        datasamples: a DataSamples object (processed train data from it is stored in self.data)
        feature: the feature name
        categorical: is the feature categorical or not
        alpha: regularization coefficient for WOE
        simple: should WoE be calculated just on input data without folds or not
        n_folds: number of folds for WOE calculation
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        '''
        self.feature = feature
        self.categorical = categorical
        if categorical:
            self.groups = {0:datasamples.train.dataframe[feature].unique().tolist()}
        else:
            self.groups = {0:[-np.inf, np.inf]} # for categorical - {group_number: [list of values]}, for ordered -{group_number: [left_bound, right_bound]}
        # -1 - group number for missings
        self.woes = {0:0} # group_number:woe
        # special values for categorical features
        self.pre_woes = {}
        self.pre_groups = {}
        self.alpha = alpha
        # edited on 21-Jun-2018 by Anna Goreva
        self.gini_history = {'Train':[], 'Validate':[], 'Test':[]}
        # edited 14.08.2018 by Yudochev Dmitry
        self.simple=simple
        self.n_folds=n_folds
        self.woe_adjust=woe_adjust
        self.cv=cv
        # edited 22.08.2018 by Yudochev Dmitry
        self.missing_group = 0
        self.backup_groups = {-1: np.nan}
        self.missing_process = 'separate'
        self.missing_min_part = None
        # edited 23.10.2018 by Yudochev Dmitry
        self.groups_history = []
        self.woes_history = []
        #if woe_object is not None:
        #    self.woe_object=weakref.ref(woe_object)
        self.datasamples=datasamples
        data=datasamples.train

        if data.weights == None:
            self.data = Data(data.dataframe[[feature, data.target]].copy(), target = data.target, features = data.features, weights = data.weights, name='Train')
            self.backup_data = Data(data.dataframe[[feature, data.target]].copy(), target = data.target, features = data.features, weights = data.weights, name='Train')
        else:
            self.data = Data(data.dataframe[[feature, data.target, data.weights]].copy(), target = data.target, features = data.features, weights = data.weights, name='Train')
            self.backup_data = Data(data.dataframe[[feature, data.target, data.weights]].copy(), target = data.target, features = data.features, weights = data.weights, name='Train')



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    #added 10.08.2018 by Yudochev Dmitry
    def calc_simple_woes(self, woe_adjust=0.5, original_values=False):
        '''
        TECH

        Simply calculates regularized WOE for each interval.
        Formula for regularized WOE of the i-th interval (value group) of feature:
        SnoothedWOE_i = log((n + alpha)*DefaultRate/(n*DefaultRate_i + alpha)),
        where n is number of samples, DefaultRate = N_bad/N_good, DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust).

        Parameters
        -----------
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        original_values: should we calculate WoE by original values (not transformed to pre_woes) for categorical features

        Returns
        ----------
        woes: {left_bound : woe} where left_bound is a lower bound of an interval, woe is a WOE value for the fold
         '''

        alpha = self.alpha
        data_to_woe = self.data.dataframe

        if self.data.weights == None:
            #edited 20.08.2018 by Yudochev Dmitry to include missing values in calculation
            N_g = (1-data_to_woe[self.data.target]).sum()
            N_b = data_to_woe[self.data.target].sum()
        else:
            N_g = data_to_woe[data_to_woe[self.data.target] == 0][self.data.weights].sum()
            N_b = data_to_woe[data_to_woe[self.data.target] == 1][self.data.weights].sum()
        DR = N_b*1.0/N_g

        # WOE for each interval
        woes = {}

        # for each interval
        for group, vals in self.groups.items():
            # special group for missings
            if group == -1:
                data_i = data_to_woe.loc[data_to_woe[self.feature].isnull()]
            elif self.categorical and original_values:
                # categorical group including missings
                if group == self.missing_group:
                    data_i = data_to_woe.loc[data_to_woe[self.feature].apply(lambda x: x in vals)| data_to_woe[self.feature].isnull() ]
                #categorical group without missings
                else:
                    data_i = data_to_woe.loc[data_to_woe[self.feature].apply(lambda x: x in vals)]
            else:
                if group == self.missing_group:
                    data_i = data_to_woe.loc[(data_to_woe[self.feature] >= vals[0]) & (data_to_woe[self.feature] < vals[1]) | data_to_woe[self.feature].isnull()]
                else:
                    data_i = data_to_woe.loc[(data_to_woe[self.feature] >= vals[0]) & (data_to_woe[self.feature] < vals[1])]

                # Without sample weights
            if self.data.weights == None:
                N_b_i = data_i[self.data.target].sum()
                N_g_i = data_i.shape[0] - N_b_i
                # Weighted WOE
            else:
                N_b_i = data_i[data_i[self.data.target] == 1][self.data.weights].sum()
                N_g_i = data_i[data_i[self.data.target] == 0][self.data.weights].sum()
            if not(N_b_i==0 and N_g_i==0):
                DR_i = (N_b_i + woe_adjust)/(N_g_i + woe_adjust)
                n = N_g_i + N_b_i
                smoothed_woe_i = np.log(DR*(alpha + n)/(n*DR_i + alpha))#*DR))
                woes[group] = smoothed_woe_i
            else:
                woes[group] = None
        return woes



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    def woe_folds(self, n_folds = 5, woe_adjust=0.5, original_values=False):
        '''
        TECH

        Breaks the feature into folds for each value group (interval) and calculates regularized WOE for each fold.
        Formula for regularized WOE of the i-th interval (value group) of feature:
        SmoothedWOE_i = log((n + alpha)*DefaultRate/(n*DefaultRate_i + alpha)),
        where n is number of samples, DefaultRate = N_bad/N_good, DefaultRate_i = N_bad_i/N_good_i.

        WOE for folds is simular to cross-validation:
        1. Samples of an interval (value group) are divided into n_folds folds
        2. For each (n_folds - 1) intervals (value groups) SmoothedWOE is calculated
        3. For each fold its SmoothedWOE is the SmoothedWOE calculated on the other folds

        Example:
        Feature 'AGE', interval '25-35 years old', n_folds = 5.
        All the clients with AGE between 25 and 35 years are divided into 5 folds. For clients of the 1st fold SmoothedWOE value is SmoothedWOE calculated on 2-5 folds.

        Parameters
        -----------
        n_folds: number of folds for WOE calculation
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        original_values: should we calculate WoE by original values (not transformed to pre_woes) for categorical features

        Returns
        ----------
        woes: {left_bound : {fold_num : woe}} where left_bound is a lower bound of an interval, fold_num is a number of a fold (from 0 to n_folds-1), woe is a WOE value for the fold
        folds: {left_bound : {fold_num : fold_indexes}} where left_bound is a lower bound of an interval, fold_num is a number of a fold (from 0 to n_folds-1),  fold_indexes is indexes of samples in the fold
        '''

        alpha = self.alpha
        data_to_woe = self.data.dataframe

        if self.data.weights == None:
            #edited 20.08.2018 by Yudochev Dmitry to include missing values in calculation
            N_g = (1-data_to_woe[self.data.target]).sum()
            N_b = data_to_woe[self.data.target].sum()
        else:
            N_g = data_to_woe[data_to_woe[self.data.target] == 0][self.data.weights].sum()
            N_b = data_to_woe[data_to_woe[self.data.target] == 1][self.data.weights].sum()
        DR = N_b*1.0/N_g

        folds = {}

        # calculation of folds
        for group, vals in self.groups.items():
            # special group for missings
            if group == -1:
                data_i = data_to_woe.loc[data_to_woe[self.feature].isnull()]
            elif self.categorical and original_values:
                # categorical group including missings
                if group == self.missing_group:
                    data_i = data_to_woe.loc[data_to_woe[self.feature].apply(lambda x: x in vals)| data_to_woe[self.feature].isnull() ]
                #categorical group without missings
                else:
                    #print ('woe_folds: do, no miss - group', group)
                    #print ('woe_folds: do, no miss - self.groups', self.groups)
                    #print ('woe_folds: do, no miss - feature_values', list(data_to_woe[self.feature].drop_duplicates()))
                    data_i = data_to_woe.loc[data_to_woe[self.feature].apply(lambda x: x in vals)]
            else:
                if group == self.missing_group:
                    data_i = data_to_woe.loc[(data_to_woe[self.feature] >= vals[0]) & (data_to_woe[self.feature] < vals[1]) | data_to_woe[self.feature].isnull()]
                else:
                    data_i = data_to_woe.loc[(data_to_woe[self.feature] >= vals[0]) & (data_to_woe[self.feature] < vals[1])]

                # store samples id as interval : indexes of samples in the fold i, indexes of other samples where i in (1, ..., n_folds)

            folds[group] = {}
            if data_i.shape[0] > n_folds:
                skf = StratifiedKFold(n_folds)
                tmp = 0
                for train_index, test_index in skf.split(X = data_i, y = data_i[self.data.target]):
                    # indexes addition
                    folds[group][tmp] = [data_i.iloc[train_index].index, data_i.iloc[test_index].index]
                    tmp = tmp + 1
            else:
                folds[group][0] = [data_i.index, data_i.index]


        # WOE for each fold and interval
        woes = {}

        # for each interval
        for group, vals in self.groups.items():
            woes[group] = {}
            # special group for missings
            if group == -1:
                data_i = data_to_woe.loc[data_to_woe[self.feature].isnull()]
            elif self.categorical and original_values:
                # categorical group including missings
                if group == self.missing_group:
                    data_i = data_to_woe.loc[data_to_woe[self.feature].apply(lambda x: x in vals)| data_to_woe[self.feature].isnull() ]
                #categorical group without missings
                else:
                    data_i = data_to_woe.loc[data_to_woe[self.feature].apply(lambda x: x in vals)]
            else:
                if group == self.missing_group:
                    data_i = data_to_woe.loc[(data_to_woe[self.feature] >= vals[0]) & (data_to_woe[self.feature] < vals[1]) | data_to_woe[self.feature].isnull()]
                else:
                    data_i = data_to_woe.loc[(data_to_woe[self.feature] >= vals[0]) & (data_to_woe[self.feature] < vals[1])]

            #for each fold
            for fold in folds[group]:
                indexes_fold = folds[group][fold][0]
                data_fold = data_i.loc[indexes_fold]
                if self.data.weights == None:
                    N_b_i = data_fold[self.data.target].sum()
                    N_g_i = data_fold.shape[0] - N_b_i
                    # Weighted WOE
                else:
                    N_b_i = data_fold[data_fold[self.data.target] == 1][self.data.weights].sum()
                    N_g_i = data_fold[data_fold[self.data.target] == 0][self.data.weights].sum()

                if not(N_b_i==0 and N_g_i==0):
                    DR_i = (N_b_i + woe_adjust)/(N_g_i + woe_adjust)
                    n = N_g_i + N_b_i
                    smoothed_woe_i = np.log(DR*(alpha + n)/(n*DR_i + alpha))#*DR))
                    woes[group][fold] = smoothed_woe_i

        #removing bounds with no data cooresponding to them (in case of empty dictionary for folds)
        woes={x:woes[x] for x in woes if woes[x]!={}}
        return woes, folds
        # woes: {left_bound : {fold_num : woe}}
        # folds: {left_bound : {fold_num : fold_indexes}}



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    #edited 10.08.2018 by Yudochev Dmitry
    def calc_woe_folds(self, n_folds = 5, woe_adjust=0.5, original_values=False):
        '''
        TECH

        Calculates WOE for each sample according to folds

        Parameters
        -----------
        n_folds: number of folds for WOE calculation
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        original_values: should we calculate WoE by original values (not transformed to pre_woes) for categorical features

        Returns
        ----------
        woes - a dictionary {group: {fold_number: woe}},
        result - a list of values transformed to woe by folds
        '''
        alpha = self.alpha
        if alpha == None:
            print ('Achtung bitte! Keine Alpha... Bis dann! :) ')
            return None

        result = []
        woes, folds = self.woe_folds(n_folds, woe_adjust=woe_adjust, original_values=original_values)

        if len(woes) < 2:
            print ('Achtung bitte! Keine WOE Gruppen! ')
            #print ('woes', woes)
            #return None

        # for each sample finds its interval (values group), fold and, consequently, WOE
        #tmp = 0
        for index, row in self.data.dataframe[[self.feature, self.data.target]].iterrows():
            for group in folds:
                for fold in folds[group]:
                    if index in folds[group][fold][1]:
                        result.append(woes[group][fold])
        return woes, result



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    def optimize_alpha(self, cv = 7, scoring = 'neg_log_loss', best_criterion = 'min', n_folds = 5, simple=False,
                       original_values=False, alpha_range = range(10, 100, 10),
                       classifier = LogisticRegression(random_state = 42), woe_adjust=0.5):
        '''
        TECH

        Optimal alpha selection for WoE-transformed data

        Parameters
        -----------
        cv: number of fold for score calculation via cross-validation
        scoring: the quality metric
        best_criterion: 'min' - for minimizing the quality measure, 'max' - for maximizing the measure
        n_folds: number of folds for WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        original_values: should we calculate WoE by original values (not transformed to pre_woes) for categorical features
        alpha_range: range optimal alpha to choose from
        classifier: classifier to use for quality estimation
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)

        Returns
        --------
        optimal alpha

        '''

        scores={}
        for alpha in alpha_range:
            self.alpha = alpha
            # edited 15.11.2018 by Anna Goreva
            if simple:
                result = pd.DataFrame(self.set_avg_woes(woes=self.calc_simple_woes(woe_adjust=woe_adjust,
                                                                              original_values=original_values),
                                                        original_values=original_values),
                                      columns = [self.feature + '_woe'])

            else:
                result = pd.DataFrame(self.calc_woe_folds(n_folds, woe_adjust=woe_adjust,
                                                          original_values=original_values)[1],
                                      columns = [self.feature + '_woe'])

            scores[alpha] = np.mean(cross_val_score(classifier, result, self.data.dataframe[self.data.target], cv = cv,
                                              scoring = scoring))

        best_alpha = -1000
        # edited 14.08.2018 by Yudochev Dmitry
        if best_criterion == 'min':
            best_alpha = min(scores, key=scores.get)
        elif best_criterion == 'max':
            best_alpha = max(scores, key=scores.get)

        if best_alpha > -1000:
            print ('best alpha: ', best_alpha)
            self.alpha = best_alpha
            return best_alpha
        else:
            return None



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    def average_fold_woe(self, woe_folds):
        '''
        TECH

        WOE calculation for an interval: average WOE through all the folds

        Parameteres:
        -------------
        woe_folds: woes for folds

        Returns
        ------------
        woes for interval bounds
        '''
        result = {}
        for group in woe_folds:
            result[group] = np.mean(np.array([woe_folds[group][fold] for fold in woe_folds[group]]))
        return result



    def set_avg_woes(self, data=None, groups=None, woes=None, original_values=False, not_found='worst'):
        '''
        TECH

        Replaces all values of a feature to related WOE

        Parameters
        -----------
        data: a Series, containing initial values of feature
        groups: a dictionary with groups description
        woes: a dictionary with WOEs for groups
        original_values: should we calculate WoE by original values (not transformed to pre_woes) for categorical features
        not_found: what WoE to use in case of values, not described in groups
            'worst': using minimal WoE value from woes dictionary
            'missing': using WoE value of the group, containing missing values
            an int: using WoE of the group with the specified number

        Returns
        -----------
        a Series of WOE-transformed feature values
        '''

        if data is None:
            data=self.data.dataframe[self.feature]
        if groups is None:
            groups=self.groups
        if woes is None:
            woes=self.woes

        result=pd.Series([np.nan]*len(data), name=self.feature+'_WOE', index=data.index)
        for g, vals in groups.items():
            if original_values and self.categorical:
                result.loc[data.isin(vals)]=woes[g]
            else:
                if isinstance(vals,list):
                    result.loc[(vals[0]<=data)&(data<vals[1])]=woes[g]

        if self.missing_group in woes and not(pd.isnull(woes[self.missing_group])):
            result.loc[pd.isnull(data)]=woes[self.missing_group]

        not_found_woe=None
        if not_found == 'missing':
            if self.missing_group not in woes or pd.isnull(woes[self.missing_group]):
                print (self.feature, ': ERROR! No woes for missings!')
                return None
            else:
                not_found_woe=woes[self.missing_group]
        elif not_found == 'worst':
            woes_copy = {}
            for (k, v) in woes.items():
                if pd.isnull(v) == False:
                    woes_copy[k] = v
            min_woe_group = min(woes_copy, key = woes_copy.get)
            not_found_woe=woes[min_woe_group]
        elif not_found in woes:
            not_found_woe=woes[not_found]
        else:
            print (self.feature, ': ERROR! No woe for unknown values')
            return None

        result=result.fillna(not_found_woe)
        return result



    def set_groups(self, data=None, groups=None, woes=None, original_values=False, not_found='worst'):
        '''
        TECH

        Replaces all values of a feature to related group

        Parameters
        -----------
        data: a Series, containing initial values of feature
        groups: a dictionary with groups description
        woes: a dictionary with WOEs for groups
        original_values: should we calculate WoE by original values (not transformed to pre_woes) for categorical features
        not_found: what group to use in case of values, not described in groups
            'worst': using group with minimal WoE value from woes dictionary
            'missing': using group, containing missing values
            an int: using group with the specified number

        Returns
        -----------
        a Series of corresponding groups for input values
        '''

        if data is None:
            data=self.data.dataframe[self.feature]
        if groups is None:
            groups=self.groups
        if woes is None:
            woes=self.woes

        result=pd.Series([np.nan]*len(data), name=self.feature+'_WOE', index=data.index)
        for g, vals in groups.items():
            if original_values and self.categorical:
                result.loc[data.isin(vals)]=g
            else:
                if isinstance(vals,list):
                    result.loc[(vals[0]<=data)&(data<vals[1])]=g

        if self.missing_group in woes and not(pd.isnull(woes[self.missing_group])):
            result.loc[pd.isnull(data)]=self.missing_group

        not_found_group=None
        if not_found == 'missing':
            if self.missing_group not in woes or pd.isnull(woes[self.missing_group]):
                print (self.feature, ': ERROR! No woes for missings!')
                return None
            else:
                not_found_group=self.missing_group
        elif not_found == 'worst':
            woes_copy = {}
            for (k, v) in woes.items():
                if pd.isnull(v) == False:
                    woes_copy[k] = v
            min_woe_group = min(woes_copy, key = woes_copy.get)
            not_found_group=min_woe_group
        elif not_found in woes:
            not_found_group=not_found
        else:
            print (self.feature, ': ERROR! No woe for unknown values')
            return None

        result=result.fillna(not_found_group)
        return result



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    def print_woe(self):
        '''
        TECH

        Prints WOE parameters in a standard and convenient way

        '''
        if self.data.weights is None:
            obs=self.data.dataframe['group'].value_counts(dropna=False)
            targets=self.data.dataframe[self.data.dataframe[self.data.target]==1]['group'].value_counts(dropna=False)
        else:
            to_calc=self.data.dataframe[['group', self.data.target, self.data.weights]]
            to_calc[self.data.target]=to_calc[self.data.target]*to_calc[self.data.weights]
            obs=self.data.dataframe[['group', self.data.weights]].groupby('group').sum()
            targets=self.data.dataframe[['group', self.data.target]].groupby('group').sum()

        targets = targets.reindex(targets.index.union(obs.index)).fillna(0).astype(int)

        for group in self.groups:
            if self.woes[group] is not None:
                if group == self.missing_group and self.missing_group!=-1:
                    print ('WOE:', self.woes[group], '  bounds: missings, ', self.groups[group], ' obs (in train):', obs[group], ' targets (in train):', targets[group])
                else:
                    print ('WOE:', self.woes[group], '  bounds:', self.groups[group], ' obs (in train):', obs[group], ' targets (in train):', targets[group])



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    def plot_woe(self):
        '''
        Plots WOEs and the sample distribution by WOE groups
        '''
        with plt.style.context(('seaborn-deep')):


            fig = plt.figure()
            ax = fig.add_subplot(111)

            woes_df=pd.DataFrame(self.woes, index=['woe']).transpose().reset_index().rename({'index':'group'},axis=1)#.sort_values('bounds')
            woes_df['bounds'] = woes_df.group.apply(lambda x: self.groups[x])

            #display(woes_df)
            if not self.categorical:
                woes_df.loc[woes_df.group != -1, 'bounds'] = woes_df.loc[woes_df.group != -1, 'bounds'].apply(lambda x: '[' + str(round(x[0], 2)) + ', ' + str(round(x[1], 2)) + ')')
            if self.missing_group != -1 and not self.categorical:
                woes_df.loc[woes_df.group == self.missing_group, 'bounds'] = woes_df.loc[woes_df.group == self.missing_group, 'bounds'] + '*'
            if self.categorical:
                woes_df=woes_df[pd.isnull(woes_df.woe)==False].sort_values('woe')

            woes_df['label']=woes_df['bounds'].apply(lambda x: str(x)[:30]+'..' if len(str(x))>30 else str(x)) + '\n group ' + woes_df.group.apply(lambda x: str(x) + '\n with missings' if self.missing_group == x else str(x))
            to_calc=self.data.dataframe.copy()
            if self.data.weights is None:
                stats=to_calc[['group', self.data.target]].groupby('group').agg(['sum', 'size'])
            else:
                to_calc[self.data.target]=to_calc[self.data.target]*to_calc[self.data.weights]
                stats=to_calc[['group', self.data.target, self.data.weights]].groupby('group').sum()
            stats.columns=['target', 'amount']
            woes_df=woes_df.merge(stats.reset_index(), on='group')

            ax.set_ylabel('Observations')
            ax.set_xticks(range(woes_df.shape[0]))
            ax.set_xticklabels(woes_df['label'], rotation=30, ha="right")

            ax.bar(range(woes_df.shape[0]), woes_df['amount'], zorder=0)
            ax.annotate('Targets:', xy=(0, 1), xycoords=('axes fraction', 'axes fraction'), xytext=(-42, 5), textcoords='offset pixels', color='red', size=11)
            for i in range(woes_df.shape[0]):
                ax.annotate(str(woes_df['target'][i]), xy=(i, 1), xycoords=('data', 'axes fraction'),
                            xytext=(0, 5), textcoords='offset pixels', color='red', size=11, ha='center')
            ax.grid(False)
            ax.grid(axis='y', zorder=1, alpha=0.6)
            ax2 = ax.twinx()
            ax2.set_ylabel('WoE')
            ax2.grid(False)
            # red is for the WOE values
            ax2.plot(range(woes_df.shape[0]), woes_df['woe'], 'ro-', linewidth=2.0, zorder=4)
            plt.suptitle(self.feature, fontsize = 16)
            fig.autofmt_xdate()
            #f.tight_layout()
            plt.show()

            if self.missing_group != -1 and self.categorical==False:
                print('*with missing')



    def auto_fit(self, datasamples, alpha = 10, groups_range=range(6,1,-1), verbose=True, scoring = 'neg_log_loss', cv = 7,
                 alpha_scoring = 'neg_log_loss', alpha_best_criterion = 'min', alpha_range = None, alpha_classifier = LogisticRegression(random_state = 42),
                 n_folds = 5, simple = False, woe_adjust = 0.5, plot_flag = True,
                 max_depth = None, min_samples_leaf = None, missing_process = 'separate', missing_min_part=0.05,
                 SM_on=True, SM_target_threshold=5, SM_size_threshold=100,
                 WOEM_on=True, WOEM_woe_threshold=0.05, WOEM_with_missing=False,
                 G_on=True, G_gini_threshold=5, G_gini_decrease_threshold=0.2, G_gini_increase_restrict=True, G_with_test=False,
                 BL_on=True, BL_conditions='',
                 WOEO_on=True, WOEO_dr_threshold=0.005, WOEO_correct_threshold=0.85, WOEO_miss_is_incorrect=True,
                 out_check=None):
        '''
        Attempts to find suitable binning satisfying all conditions and passing all checks, adjusted by user parameters
        Values for groups amount are taken from groups_range in the specified order, feature is being fitted,
        then gini, business logic and WoE order checks are ran. If all checks are passed, then current binning stays,
        otherwise the feature is fitted with the next value of groups amount and checks are carried out again.

        Parameters
        -----------
        datasamples: a DataSamples object, containing bootstrap samples for WOEOrder and Gini checks (can be omitted)
        groups_range: list of integers that will be used as max_leaf_nodes value in FeatureWOE.fit method in the specified order
        verbose: if comments and graphs from checks should be printed
        out_check: a path for xlsx output file to export check results

        Fit options:
        -----------
        alpha: regularization paramter, can be set by user or calculated automatically
        scoring: a measure for cross-validation used used for optimal WOE splitting
        cv: the numbder of folds for cross-validation used ufor optimal WOE splitting
        alpha_scoring: a measure for cross-validation used used for optimal alpha search
        alpha_best_criterion: if 'min', the lower alpha_scoring - the better, if 'max', the higher the measure - the better
        n_folds: number of folds for WOE calculation
        alpha_range: a range of values accesible for the optimal alpha choice
        alpha_classifier: a calssifier used for the optimal alpha search
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        plot_flag: should bins be plotted or not
        max_depth: the maximum of the DecisionTree depth used for optimal WOE splitting
        min_samples_leaf: the minimum of the DecisionTree leaf size used for optimal WOE splitting (one value)
        missing_process: way of handling missing values
            'separate' - keep missing values in a separate bin
            'worst' - assign missing values to the worst bin
            'nearest' - assign missing values to the nearest bin by WoE
            'worst_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the worst bin, otherwise keep separate
        missing_min_part: threshold for missing values part in case of missing_process=='worst_or_separate'

        Size merge options (SM):
        -----------------------
        SM_on: flag to turn on/off WoE merge process
        SM_target_threshold: min number of targets for group to not be considered small
        SM_size_threshold: min number of observations for group to not be considered small

        WoE merge options (WOEM):
        -----------------------
        WOEM_on: flag to turn on/off merging by WoE process
        WOEM_woe_threshold: if woe difference between groups (neighboring groups for interval) is less then this threshold, then they are to be merged
        WOEM_with_missing: should woe difference with missing group also be checked

        Gini check options (G):
        ----------------------
        G_on: flag to turn on/off Gini check
        G_gini_threshold: gini on train and validate/95% bootstrap should be greater then this
        G_gini_decrease_threshold: gini decrease from train to validate/95% bootstrap deviation from mean to mean should be greater then this
        G_gini_increase_restrict: if gini increase should also be restricted
        G_with_test: should features be also checked on test sample (and be excluded in case of failure)

        Business logic check options (BL):
        ---------------------------------
        BL_on: flag to turn on/off business logic check
        BL_conditions_dict: adress for excel-file with business logic conditions (columns 'variable' and 'condition' are mandatory)

        WoE order check options (WOEO):
        ------------------------------
        WOEO_on: flag to turn on/off WoE order check
        WOEO_dr_threshold: if WoE order is not correct, then default rate difference between swaped bins is checked
        WOEO_correct_threshold: what part of checks on bootstrap should be correct for feature to pass the check
        WOEO_woe_adjust: woe adjustment factor (for Default_Rate_i formula)
        WOEO_miss_is_incorrect: is there is no data for a bin on bootstrap sample, should it be treated as error or not

        Returns
        ----------
        A boolean value: True, if successful binning was found, else False
        and if out_check then dataframes with log, gini, business logic, WoE and ER information for export

        '''

        feature_log=pd.DataFrame()
        feature_gini=pd.DataFrame()
        feature_bl=pd.DataFrame()
        feature_woe=pd.DataFrame()
        feature_er=pd.DataFrame()

        groups_range=list(groups_range)
        groups_numbers_tried=[]
        iteration=0
        for groups_number in groups_range:
            iteration=iteration+1
            print('Attempting to fit', groups_number, 'groups')
            self.fit(alpha = alpha, new_groups = True, scoring = scoring, cv = cv, alpha_scoring = alpha_scoring,
                     alpha_best_criterion = alpha_best_criterion, n_folds = n_folds, alpha_range = alpha_range,
                     alpha_classifier = alpha_classifier, simple = simple, woe_adjust = woe_adjust, plot_flag = plot_flag,
                     max_leaf_nodes = [groups_number], max_depth = max_depth, min_samples_leaf = min_samples_leaf,
                     missing_process = missing_process, missing_min_part=missing_min_part)
            groups_array=np.array(groups_range)
            fact_number_groups=len([x for x in self.woes if pd.isnull(self.woes[x])==False])
            groups_numbers_tried.append(fact_number_groups)

            for i in groups_array[(groups_array>fact_number_groups)&(np.isin(groups_array, groups_numbers_tried)==False)]:
                groups_range.remove(i)

            if SM_on:
                self.merge_by_size(target_threshold=SM_target_threshold, size_threshold=SM_size_threshold, alpha = alpha, simple=simple,
                                  woe_adjust=woe_adjust, n_folds = n_folds, plot_flag=plot_flag, no_backup=False, scoring=scoring)
            else:
                print('... Merge by size: skipped')

            if WOEM_on:
                self.merge_by_woe(woe_threshold=WOEM_woe_threshold, with_missing=WOEM_with_missing, alpha = alpha, simple=simple,
                                  woe_adjust=woe_adjust, n_folds = n_folds, plot_flag=plot_flag, cv=cv, no_backup=False, scoring=scoring)
            else:
                print('... Merge by WoE: skipped')

            fact_number_groups_after_merge=len([x for x in self.woes if pd.isnull(self.woes[x])==False])
            if fact_number_groups_after_merge==1:
                print('After the attempt with', groups_number, 'groups only one group remains. Continue cycle..')
                feature_log=feature_log.append(pd.Series({'feature':self.feature, 'iteration':iteration,'result':'Failure', 'reason':'After the attempt with '+str(groups_number)+' groups only one group remains'}), ignore_index=True)
                continue
            else:
                print('Actual number of groups fitted (before merge) is ', fact_number_groups)
                print('--------------------------------------- Checking ---------------------------------------')
                if G_on:
                    gini_correct, add_gini=GiniChecker().work(self, datasamples, gini_threshold=G_gini_threshold, gini_decrease_threshold=G_gini_decrease_threshold,
                                                                 gini_increase_restrict=G_gini_increase_restrict, verbose=verbose, with_test=G_with_test, out=out_check)
                    add_gini.update({'feature':self.feature, 'iteration':iteration})
                    feature_gini=feature_gini.append(pd.DataFrame(add_gini, index=[0]), ignore_index=True)[['feature', 'iteration', 'Train', 'Validate', 'Test']+['Bootstrap'+str(i) for i in range(len(self.datasamples.bootstrap))]].dropna(axis=1, how='all')
                else:
                    print('... Gini check skipped')
                    gini_correct=True
                if not gini_correct:
                    print('... Gini check failed (groups number =', fact_number_groups, '). Continue cycle..')
                    print('----------------------------------------------------------------------------------------')
                    feature_log=feature_log.append(pd.Series({'feature':self.feature, 'iteration':iteration,'result':'Failure', 'reason':'Gini check failed (groups number = '+str(fact_number_groups)+')'}), ignore_index=True)
                    continue
                else:
                    if G_on:
                        print('... Gini check passed')
                    if BL_on:
                        business_logic_correct, add_bl=BusinessLogicChecker().work(self, conditions=BL_conditions, verbose=verbose, out=out_check)
                        add_bl['overall_result']=business_logic_correct
                        add_bl['iteration']=iteration
                        feature_bl=feature_bl.append(add_bl, ignore_index=True)[['feature', 'iteration', 'categorical', 'condition', 'fact', 'condition_result', 'overall_result']]
                    else:
                        print('... Business logic check skipped')
                        business_logic_correct=True
                    if not business_logic_correct:
                        print('... Business logic check failed (groups number =', fact_number_groups, '). Continue cycle..')
                        print('----------------------------------------------------------------------------------------')
                        feature_log=feature_log.append(pd.Series({'feature':self.feature, 'iteration':iteration,'result':'Failure', 'reason':'Business logic check failed (groups number = '+str(fact_number_groups)+')'}), ignore_index=True)
                        continue
                    else:
                        if BL_on:
                            print('... Business logic check passed')
                        if WOEO_on:
                            woe_order_correct, add_woe, add_er=WOEOrderChecker().work(self, datasamples, dr_threshold=WOEO_dr_threshold, correct_threshold=WOEO_correct_threshold,
                                                                                      woe_adjust=woe_adjust, miss_is_incorrect=WOEO_miss_is_incorrect, verbose=verbose, out=out_check)
                            add_woe['feature']=self.feature
                            add_woe['iteration']=iteration
                            add_er['feature']=self.feature
                            add_er['iteration']=iteration
                            feature_woe=feature_woe.append(add_woe, ignore_index=True)[['feature', 'iteration', 'group', 'Train', 'Validate', 'Test']+['Bootstrap'+str(i) for i in range(len(self.datasamples.bootstrap))]].dropna(axis=1, how='all')
                            feature_er=feature_er.append(add_er, ignore_index=True)[['feature', 'iteration', 'group', 'Train', 'Validate', 'Test']+['Bootstrap'+str(i) for i in range(len(self.datasamples.bootstrap))]].dropna(axis=1, how='all')
                        else:
                            print('... WoE order check skipped')
                            woe_order_correct=True
                        if not woe_order_correct:
                            print('... WoE order check failed (groups number =', fact_number_groups, '). Continue cycle..')
                            print('----------------------------------------------------------------------------------------')
                            feature_log=feature_log.append(pd.Series({'feature':self.feature, 'iteration':iteration,'result':'Failure', 'reason':'WoE order check failed (groups number = '+str(fact_number_groups)+')'}), ignore_index=True)
                            continue
                        else:
                            if WOEO_on:
                                print('... WoE order check passed')
                            print('After the attempt with', groups_number, 'groups (actual = ',fact_number_groups_after_merge,') all active checks passed successfully. Breaking cycle.')
                            print('----------------------------------------------------------------------------------------')
                            feature_log=feature_log.append(pd.Series({'feature':self.feature, 'iteration':iteration,'result':'Success', 'reason':''}), ignore_index=True)
                            if out_check:
                                return True, feature_log[['feature', 'iteration','result', 'reason']], feature_gini, feature_bl, feature_woe, feature_er
                            else:
                                return True
        else:
            print('After all attempts no suitable binning was found. Shame.')
            if out_check:
                return False, feature_log[['feature', 'iteration','result', 'reason']], feature_gini, feature_bl, feature_woe, feature_er
            else:
                return False



    # rounding boundaries
    # added by Anna Goreva 25-Oct-2018
    # edited on 01-Nov-2018
    def set_round_groups(self, round_digits, migration_coefficient = 0.001, with_acceptable_migration=False):
        '''
        Rounds boundaries of groups. Checks if the rounding parameter is valid and extends it if necessary.
        Checks if groups do not collide and if groups' samples remain stable.

        Parameters
        -----------
        round_digits: the number of digits we would like to leave after comma
        migration_coefficient: threshold for the part of a group's sample that can migrate to the other groups
        with_acceptable_migration: is it ok to allow migration between groups after rounding
        '''

        if not self.categorical:
            if with_acceptable_migration:
                # min interval between groups' boundaries
                min_diff_b = min([1 if k < 0
                                  else (v[1] - self.data.dataframe[self.feature].min() if -np.inf in v
                                        else (self.data.dataframe[self.feature].max() - v[0] if np.inf in v
                                              else v[1] - v[0])) for (k, v) in self.groups.items()])
                #print ('Minimal difference between group bounds:', min_diff_b)
                change_rounds = False
                while min_diff_b < (.1)**round_digits:
                    round_digits += 1
                    change_rounds = True


                # checks changes in groups' volumes in case of rounding
                change_rounds_2 = True
                while change_rounds_2 and len(self.groups) > 1:
                    change_rounds_2 = False
                    for group in self.groups:
                        if group != -1:
                            left_from=round(self.groups[group][0], round_digits) if round(self.groups[group][0], round_digits) <= self.groups[group][0] else self.groups[group][0]
                            left_to=self.groups[group][0] if round(self.groups[group][0], round_digits) <= self.groups[group][0] else round(self.groups[group][0], round_digits)

                            right_from=round(self.groups[group][1], round_digits) if round(self.groups[group][1], round_digits) <= self.groups[group][1] else self.groups[group][1]
                            right_to=self.groups[group][1] if round(self.groups[group][1], round_digits) <= self.groups[group][1] else round(self.groups[group][1], round_digits)

                            migration = sum(self.data.dataframe[self.feature].apply(
                                                lambda x: (left_from <= x < left_to) or (right_from <= x < right_to)))
                            if migration/self.data.dataframe.shape[0] >= migration_coefficient and not change_rounds_2:
                                change_rounds_2 = True
                                round_digits += 1
                                #print (self.groups[group], round_digits)
                            #print (self.feature, self.groups[group], migration.shape[0]/self.data.dataframe.shape[0])

                #debug
                #print ('Rounding: group size check ended')

                # rounding
                rounded_groups = {}
                for (k, v) in self.groups.items():
                    if k >= 0:
                        rounded_groups[k] = [round(v[0], round_digits), round(v[1], round_digits)]
                    else:
                        rounded_groups[k] = v

                if change_rounds:
                    print ('The rounding parameter is too large, setting to', round_digits)

                print ('Rounding parameter:', round_digits)
                #print ('Rounded groups:', rounded_groups)

                self.groups = rounded_groups
            else:
                exact_edges=[]
                rounded_edges=[]
                for (k, v) in self.groups.items():
                    if k>0:
                        exact_edges.append(v[0])
                        before_split = self.data.dataframe[self.feature][self.data.dataframe[self.feature]<v[0]].max()
                        rounded_split = (before_split+v[0])/2
                        precision = len(str(rounded_split).split('.')[1])
                        previous_rounded_split = None
                        while rounded_split>before_split and rounded_split<v[0] and previous_rounded_split!=rounded_split:
                            previous_rounded_split=rounded_split
                            final_precision = precision
                            precision-=1
                            rounded_split = int(rounded_split) if precision==0 \
                                                else int((rounded_split)*(10**precision))/(10**precision)
                        candidate_split=int((before_split+v[0])/2) if final_precision==0 \
                                                 else int(((before_split+v[0])/2)*(10**final_precision))/(10**final_precision)

                        if final_precision<len(str(v[0]).split('.')[1])-(len(str(v[0]).replace('.',''))-len(str(v[0]).replace('.','').rstrip('0'))) and \
                           (self.data.dataframe[self.feature]<v[0]).sum()==(self.data.dataframe[self.feature]<candidate_split).sum():
                            rounded_edges.append(candidate_split)
                        else:
                            rounded_edges.append(v[0])

                rounded_groups = {}
                for (k, v) in self.groups.items():
                    if k >= 0:
                        rounded_groups[k] = [-np.inf if v[0]==-np.inf else rounded_edges[exact_edges.index(v[0])],
                                              np.inf if v[1]==np.inf  else rounded_edges[exact_edges.index(v[1])]]
                    else:
                        rounded_groups[k] = v

                if rounded_edges!=exact_edges:
                    print ('Rounding edges:', str([exact_edges[i] for i in range(len(exact_edges)) if exact_edges[i]!=rounded_edges[i]]),
                           'to', str([rounded_edges[i] for i in range(len(rounded_edges)) if exact_edges[i]!=rounded_edges[i]]))
                    self.groups = rounded_groups


    def woe_fit(self, simple, n_folds, woe_adjust, do_pd):
        '''
        TECH

        Calculates WOE for FeatureWOE due to simple parameter

        Parameters
        -----------
        simple: should WoE be calculated just on input data without folds or not
        '''
        woes = {}
        if simple:
            woes = self.calc_simple_woes(woe_adjust=woe_adjust, original_values = do_pd)
        else:
            try:
                woes = self.average_fold_woe(self.calc_woe_folds(n_folds = n_folds, woe_adjust=woe_adjust, original_values = do_pd)[0])
            except ValueError:
                print ('ValueError for WOE calculation. Please check n_folds parameter and group sizes. Turning to simple WOE calculation... Hope it works.')
                woes = self.calc_simple_woes(woe_adjust=woe_adjust, original_values = do_pd)
        return woes



    def categorical_to_interval(self, alpha, scoring, cv, alpha_scoring, alpha_best_criterion, n_folds, alpha_range, alpha_classifier, simple, woe_adjust):
        '''
        TECH

        Transforms categorical features into interval oned via WOE calculation for each category.

        Parameters
        -------------
        alpha: regularization paramter, can be set by user or calculated automatically
        scoring: a measure for cross-validation used used for optimal WOE splitting
        cv: the numbder of folds for cross-validation used ufor optimal WOE splitting
        alpha_scoring: a measure for cross-validation used used for optimal alpha search
        alpha_best_criterion: if 'min', the lower alpha_scoring - the better, if 'max', the higher the measure - the better
        n_folds: number of folds for WOE calculation
        alpha_range: a range of values accesible for the optimal alpha choice
        alpha_classifier: a calssifier used for the optimal alpha search
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        '''
        # WOE for each category: optimal alpha - woe by folds - final woe calculation
        print ('Preliminary data transformation: from categorical feature to WOE')

        woe_groups=copy.deepcopy(self.groups)
        self.groups = {}

        # turning categorical values into separate groups for pre-woe
        group = 0
        # edited on 07/09/2018 by Anna Goreva
        if self.data.dataframe[self.feature].isnull().any():
            self.groups[-1] = [np.nan]
        for value in sorted(self.data.dataframe[self.feature].dropna().unique().tolist()):
            self.groups[group] = [value]
            group = group + 1


        if alpha == None:
            alpha_opt1 = self.optimize_alpha(cv = cv, scoring = alpha_scoring, best_criterion = alpha_best_criterion,
                                        n_folds = n_folds, alpha_range = alpha_range, original_values=True,
                                        classifier = alpha_classifier, simple=simple, woe_adjust=woe_adjust)
        else:
            alpha_opt1 = alpha
        self.alpha = alpha_opt1


        self.pre_woes = self.woe_fit(simple, n_folds, woe_adjust, True)

        self.data.dataframe[self.feature] = self.set_avg_woes(woes=self.pre_woes, original_values=True)
        self.pre_groups = copy.deepcopy(self.groups)
        print ('Now I can work with ', self.feature, 'as a continious one')
        # Now the feature is considered numerical
        self.groups = copy.deepcopy(woe_groups)



    def get_tree_splits(self, dtree):
        '''
        TECH

        Returns list of thresholds of the deision tree.

        Parameters
        ---------------
        tree: DecisionTreeClassifier object

        Returns
        ---------------
        boundaries: list of thresholds
        '''
        children_left = dtree.tree_.children_left
        children_right = dtree.tree_.children_right
        threshold = dtree.tree_.threshold
        # boundaries of split
        boundaries = [-np.inf, np.inf]


        # determination of groups
        for i in range(dtree.tree_.node_count):
            if children_left[i] != children_right[i]:
                boundaries.append(threshold[i])

        return sorted(boundaries)



    def bounds_to_groups(self, bounds):
        '''
        TECH

        Transforms list of bounds to feature_woe.groups - only for interval features.

        Parameters
        -------------
        bounds: list of group bounds

        Returns:
        -------------
        groups: dictionary {group_number : group_bounds}
        '''
        if np.nan in bounds or None in bounds:
            print ('Error! NaN in bounds!!!')
            return None
        groups = {}
        bounds = sorted(bounds)
        for i in range(len(bounds)-1):
            groups[i] = [bounds[i], bounds[i+1]]

        return groups



    def missings_for_woe(self, missing_process, missing_min_part=0.05):
        '''
        TECH

        Processes missings for fit() according to 'missing_process' parameter, calculates missing_group for not categorical.
        Works only for interval features.

        Parameters
        -------------
        missing_process: parameter 'missing_process' from self.fit()
        '''
        woes_no_miss = self.calc_simple_woes()
        # find suitable group for missings
        # for interval features:
        woe_miss = woes_no_miss[-1]
        del woes_no_miss[-1]
        # for intervals: return inf to the last group
        self.groups[max(self.groups)][1] = np.inf
        if missing_process=='worst':
            self.missing_group = min(woes_no_miss, key = woes_no_miss.get)
        elif missing_process=='nearest':
            nearest_group, nearest_woe = self.find_nearest(woe_miss, woes_no_miss)
            self.missing_group = nearest_group
        elif missing_process=='worst_or_separate':
            if (self.data.weights is None and self.data.dataframe[self.feature].isnull().value_counts()[True]/self.data.dataframe[self.feature].shape[0]<missing_min_part) or \
                (self.data.weights is not None and self.data.dataframe[self.data.dataframe[self.feature].isnull()][self.data.weights].sum()/self.data.dataframe[self.data.weights].sum()<missing_min_part):
                self.missing_group = min(woes_no_miss, key = woes_no_miss.get)
        elif missing_process=='nearest_or_separate':
            if (self.data.weights is None and self.data.dataframe[self.feature].isnull().value_counts()[True]/self.data.dataframe[self.feature].shape[0]<missing_min_part) or \
                (self.data.weights is not None and self.data.dataframe[self.data.dataframe[self.feature].isnull()][self.data.weights].sum()/self.data.dataframe[self.data.weights].sum()<missing_min_part):
                nearest_group, nearest_woe = self.find_nearest(woe_miss, woes_no_miss)
                self.missing_group = nearest_group
        elif missing_process=='best_or_separate':
            if (self.data.weights is None and self.data.dataframe[self.feature].isnull().value_counts()[True]/self.data.dataframe[self.feature].shape[0]<missing_min_part) or \
                (self.data.weights is not None and self.data.dataframe[self.data.dataframe[self.feature].isnull()][self.data.weights].sum()/self.data.dataframe[self.data.weights].sum()<missing_min_part):
                self.missing_group = max(woes_no_miss, key = woes_no_miss.get)
        if self.missing_group != -1:
            del self.groups[-1]
        else:
            self.groups[-1] = np.nan



    def find_nearest(self, x, values):
        '''
        TECH

        Finds in values the nearest one for x. If 'values' is dict then returns the nearest value and its key and
        if 'values' is list then returns the nearest value and its index

        Parameters
        --------------
        x: float, the value to process
        values: dict or list of possible nearest values

        Returns
        -------------


        '''
        if isinstance(values, dict):
            diff = {}
            for (k, v) in values.items():
                diff[abs(v-x)] = k
            return diff[min(diff)], values[diff[min(diff)]]
        elif isinstance(values, list) or  isinstance(values, np.ndarray):
            diff = abs(np.array(values) - x)
            return diff.index(min(diff)), values[diff.index(min(diff))]



    def fit_gridsearch(self, x_train, cv, scoring, parameters_grid, y_train = None):
        '''
        TECH

        Searches for the best decision tree for groups. Used in self.fit().

        Parameters
        ------------
        x_train: pd.Series to fit on
        cv: number of folds for GridSearchCV
        scoring: scoring parameter for cross-validation
        parameters_grid: parameters for gridsearch
        y_train: pd.Series with targets for fit

        Returns
        ------------
        The best decision tree
        '''
        if y_train is None:
            y_train = self.data.dataframe[self.data.target]

        dtree = DecisionTreeClassifier()

        if cv==1:
            x_train2=x_train.append(x_train, ignore_index=True)
            y_train2=y_train.append(y_train, ignore_index=True)
            if self.data.weights is not None:
                w_train=np.array(self.data.dataframe[self.data.weights].append(self.data.dataframe[self.data.weights], ignore_index=True))
        elif self.data.weights is not None:
            w_train = np.array(self.data.dataframe[self.data.weights])

        test_fold = [-1 for x in range(x_train.shape[0])]+[0 for x in range(x_train.shape[0])]
        gridsearch = GridSearchCV(dtree, parameters_grid, scoring=scoring,
                                  cv=cv if cv!=1 else PredefinedSplit(test_fold=test_fold),
                                  fit_params = None if self.data.weights is None else {'sample_weight': w_train})

        gridsearch.fit(x_train[:, None] if cv!=1 else x_train2[:, None], y_train if cv!=1 else y_train2)

        return gridsearch.best_estimator_



    def categorical_recover(self):
        '''
        TECH

        Recovers self.groups and self.woes for categorical non-predefined features because such features are pre-processed
        '''
        if self.categorical:
            final_groups = {}
            for group, vals in self.groups.items():
                if isinstance(vals, list):
                    for category, pre_woe in self.pre_woes.items():
                        if not pd.isnull(pre_woe):
                            if pre_woe >= vals[0] and pre_woe < vals[1]:
                                if pd.isnull(self.pre_groups[category]):
                                    self.missing_group = group
                                else:
                                    if group in final_groups:
                                        final_groups[group] = final_groups[group] + self.pre_groups[category]
                                    else:
                                        final_groups[group] = self.pre_groups[category]

            if self.missing_group not in final_groups:
                final_groups[self.missing_group] = [np.nan]
            self.groups = {x:final_groups[x] for x in sorted(final_groups)}
            final_woes = {}
            for group in self.groups:
                if group in self.woes:
                    final_woes[group] = self.woes[group]
            self.woes = final_woes



    def initialize_parameters(self, max_leaf_nodes, min_samples_leaf, max_depth):
        '''
        TECH

        GridSearchCV parameters initialization.

        Parameters
        ------------
        max_leaf_nodes: the maximum of the DecisionTree leaves number used for optimal WOE splitting
        max_depth: the maximum of the DecisionTree depth used for optimal WOE splitting
        min_samples_leaf: the minimum of the DecisionTree leaf size used for optimal WOE splitting (one value)

        Returns
        -----------
        parameters_dict for GridSearchCV
        '''
        parameters_grid={}
        if max_leaf_nodes is not None:
            parameters_grid['max_leaf_nodes'] = max_leaf_nodes
        if min_samples_leaf is not None:
            parameters_grid['min_samples_leaf'] = [min_samples_leaf]
        else:
            parameters_grid['min_samples_leaf'] = [int(round(self.data.dataframe.shape[0]*i, 0)) for i in [.01, .005, .001]]
        if max_depth is not None:
            parameters_grid['max_depth'] = max_depth
        else:
            parameters_grid['max_depth'] = [i for i in range(1, 4)]
        return parameters_grid



    def update_gini_history(self, cv, alpha_classifier, scoring):
        '''
        TECH

        Updates gini history in self.fit()

        Parameters
        ------------
        alpha_classifier: a calssifier used for the optimal alpha search
        cv: the numbder of folds for cross-validation

        '''
        samples=[self.data, self.datasamples.validate, self.datasamples.test]

        for s in samples:
            if s is not None:
                #print(samples[i].name, 'before')
                if s.name!='Train':
                    s=self.transform(s)
                else:
                    if cv==1:
                        woe2=s.dataframe[self.feature + '_WOE'].append(s.dataframe[self.feature + '_WOE'], ignore_index=True)
                        y2=s.dataframe[s.target].append(s.dataframe[s.target], ignore_index=True)
                    cv_scoring = cross_val_score(alpha_classifier,
                                                 s.dataframe[self.feature + '_WOE'][:, None] if cv!=1 else woe2[:, None],
                                                 s.dataframe[s.target] if cv!=1 else y2, cv = cv if cv!=1 else PredefinedSplit(test_fold=[-1 for x in range(s.dataframe.shape[0])] + [0 for x in range(s.dataframe.shape[0])]),
                                                 scoring = scoring).mean()
                #print(samples[i].name, 'after')
                fpr, tpr, _ = roc_curve(s.dataframe[s.target], -s.dataframe[self.feature + '_WOE'])
                self.gini_history[s.name].append((2*auc(fpr, tpr)-1)*100)

                if s.name=='Train':
                    print (scoring, 'score:', cv_scoring)
                print ('Gini ('+ s.name+'):', self.gini_history[s.name][-1])



    def plot_gini(self):
        '''
        Plots Gini history
        '''
        fig = plt.figure(figsize=(3,3))
        ax = fig.add_subplot(111)
        for sample in self.gini_history:
            if len(self.gini_history[sample])>0:
                ax.plot([i for i in range(len(self.gini_history[sample]))], self.gini_history[sample], label=sample)
        ax.legend()
        plt.suptitle(self.feature + ': Gini changes', fontsize = 14)
        plt.show()



    # edited on Nov-19-2018 by Anna Goreva: refactoring
    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    def fit(self, alpha = 10, new_groups = True, scoring = 'neg_log_loss', cv = 7, round_digits = 1,
            alpha_scoring = 'neg_log_loss', alpha_best_criterion = 'min', n_folds = 5, alpha_range = None, alpha_classifier = LogisticRegression(random_state = 42),
            simple = None, woe_adjust = None, plot_flag = True,
            max_leaf_nodes = None, max_depth = None, min_samples_leaf = None, missing_process = 'separate', missing_min_part=0.05, no_history=False, rounding_migration_coef = .001):
        '''
        Optimizes alpha, determines optimal split into WOE intervals and calculates WOE. After that, the class is ready to
        transform unknown datasets containing the feature.

        Parameters
        -----------
        alpha: regularization paramter, can be set by user or calculated automatically
        new_groups: if True, the bounds of the feature woe groups are deleted; usefull for refitting a Feature_woe object
        scoring: a measure for cross-validation used used for optimal WOE splitting
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        round_digits: a number of digits to leave after the comma when rounding boundaries
        alpha_scoring: a measure for cross-validation used used for optimal alpha search
        alpha_best_criterion: if 'min', the lower alpha_scoring - the better, if 'max', the higher the measure - the better
        n_folds: number of folds for WOE calculation
        alpha_range: a range of values accesible for the optimal alpha choice
        alpha_classifier: a calssifier used for the optimal alpha search
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        plot_flag: should bins be plotted or not
        max_leaf_nodes: the maximum of the DecisionTree leaves number used for optimal WOE splitting
        max_depth: the maximum of the DecisionTree depth used for optimal WOE splitting
        min_samples_leaf: the minimum of the DecisionTree leaf size used for optimal WOE splitting (one value)
        missing_process: way of handling missing values
            'separate' - keep missing values in a separate bin
            'worst' - assign missing values to the worst bin
            'nearest' - assign missing values to the nearest bin by WoE
            'worst_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the worst bin, otherwise keep separate
            'best_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the best bin, otherwise keep separate
            'nearest_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the nearest bin, otherwise keep separate
        missing_min_part: threshold for missing values part in case of missing_process=='worst_or_separate'
        no_history: should new gini value be added to gini_history or not (True for rollback)
        rounding_migration_coef: part of sample that is allowed to migrate from one group to another while rounding group boundaries
        '''

        # edited 15.08.2018 by Yudochev Dmitry
        #if the are values for simple, n_folds and woe_adjust, then they will be used henceforth (in split and merge)
        #only in .fit these values can be updated for default use in split and merge
        if simple is not None:
            self.simple=simple
        elif self.simple is not None:
            simple = self.simple
        if n_folds is not None:
            self.n_folds=n_folds
        elif self.n_folds is not None:
            n_folds = self.n_folds
        if woe_adjust is not None:
            self.woe_adjust=woe_adjust
        elif self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if cv is not None:
            self.cv=cv
        elif self.cv is not None:
            cv=self.cv

        # edited 22.08.2018 by Yudochev Dmitry - same as above
        if missing_process is not None:
            self.missing_process=missing_process
        elif self.missing_process is not None:
            missing_process=self.missing_process
        if missing_min_part is not None:
            self.missing_min_part=missing_min_part
        elif self.missing_min_part is not None:
            missing_min_part=self.missing_min_part


        if alpha is not None and alpha_range is not None:
            print ('Do not understand whether I should optimize alpha or not. Please set alpha or alpha_range to None')
            return None

        if alpha is not None:
            self.alpha = alpha

        # categorical and pre-defined flag
        do_pd = False
        if new_groups:
            # History update
            if self.woes!={}:
                self.to_history()
            self.groups = {-1: np.nan}
            self.missing_group = -1
        elif (len(self.groups) > 1 or (len(self.groups) == 1 and not -1 in self.groups)) and self.categorical:
            do_pd = True
        # primary groups for categorical features
        self.pre_groups = copy.deepcopy(self.groups)

        if self.feature + '_raw' in self.data.dataframe.columns:
            self.data.dataframe[self.feature] = self.data.dataframe[self.feature + '_raw']

        # For categorical features:
        # intermediate WOEs are calculated for each category => the features turns from categorical to numerical => for the further transformations the feature is considered numerical
        if not do_pd and self.categorical:
            self.categorical_to_interval(alpha, scoring, cv, alpha_scoring, alpha_best_criterion, n_folds, alpha_range, alpha_classifier, simple, woe_adjust)

        #optimal bounds calculation
        if new_groups:
            print ('Auto binning for ', self.feature)
            print ('Searching for the best split...')

            # GridSearchCV parameters
            parameters_grid=self.initialize_parameters(max_leaf_nodes, min_samples_leaf, max_depth)

            skip_groups=False

            # should move to WOE.fit() in order to calculate ouliers for all variables' missings
            # only for interval features because for categorical missings are already in a separate group
            if self.data.dataframe[self.feature].isnull().any():
                if (min_samples_leaf is not None) and (self.data.dataframe[self.feature].dropna().shape[0]<min_samples_leaf):
                    print('Too few non-missing values:', self.data.dataframe[self.feature].dropna().shape[0])
                    skip_groups=True
            max_value = self.data.dataframe[self.feature].max()

            # flag of missings (only for interval features because of preprocessing categorical features)
            missings_to_process = self.data.dataframe[self.feature].isnull().any()
            x_train = MissingProcessor().work(self.data, parameters = {max_value + 1: self.feature}).dataframe[self.feature]

            # search for the best split
            if not skip_groups:
                # decision tree of split with pre-processed missings
                final_tree = self.fit_gridsearch(x_train, cv, scoring, parameters_grid)
                final_tree.fit(x_train[:, None],
                               self.data.dataframe[self.data.target],
                               sample_weight = None if self.data.weights is None else np.array(self.data.dataframe[self.data.weights]))

                # boundaries of split
                boundaries = self.get_tree_splits(final_tree)

                # determine whether missings should be separated: if missings present and in the last groups there are more than 1 distinct values (if only 1 then it is missing)
                if missings_to_process and x_train[x_train>= boundaries[-2]].value_counts().shape[0] > 1:
                    # separate missings
                    boundaries = sorted(boundaries + [max_value + 1])
                    # if remaining group is too small
                    if min_samples_leaf is not None and (self.data.dataframe[self.data.dataframe[self.feature].apply(lambda x: x < boundaries[-2] and x >= boundaries[-3])].shape[0] < min_samples_leaf):
                        # unite 2 last groups (non-missing)
                        # possible due to the earlier check for few non-missings
                        boundaries.remove(boundaries[-3])


                # transform boundaries to self.groups
                self.groups = self.bounds_to_groups(boundaries)
                # WOE for groups missings
                self.missing_group = -1
                if missings_to_process:
                    # moving missings to group -1
                    self.groups[-1] = copy.deepcopy(self.groups[max(self.groups)])
                    del self.groups[max(self.groups)]
                    # processing missings
                    self.missings_for_woe(missing_process, missing_min_part=missing_min_part)


            self.set_round_groups(round_digits, rounding_migration_coef)


        self.groups = {x:self.groups[x] for x in sorted(self.groups)}

        print ('Primary groups:', self.groups)

        if simple==False:
            print ('Final WOEs calculation: optimal alpha -> WOEs by folds -> average WOEs through folds for every group')
        else:
            print ('Final WOEs calculation: optimal alpha -> simple WOEs for every group')

        # optimal alpha calculation
        if alpha == None:
            print ('Optimal alpha calculation...')
            alpha_opt = self.optimize_alpha(cv = cv, scoring = alpha_scoring, best_criterion = alpha_best_criterion,
                                            n_folds = n_folds, alpha_range = alpha_range, original_values=False,
                                            classifier = alpha_classifier, simple=simple, woe_adjust=woe_adjust)
        else:
            alpha_opt = alpha
        self.alpha = alpha_opt
        print ('alpha for WOE:', alpha_opt)

        # WOE calculation
        self.woes = self.woe_fit(simple, n_folds, woe_adjust, do_pd)

        self.data.dataframe[self.feature + '_raw'] = self.backup_data.dataframe[self.feature].copy()
        self.data.dataframe['group'] = self.set_groups(woes=self.woes, original_values = do_pd)
        self.data.dataframe[self.feature + '_WOE'] = self.set_avg_woes(woes=self.woes, original_values = do_pd)

        # for categorical features recovering the original values
        if not do_pd and self.categorical:
            print ('Primary WOEs')
            self.print_woe()
            self.categorical_recover()
        elif new_groups:
            self.pre_woes=self.woes

        if len(self.groups) != len(self.woes):
            if (-1 in self.woes and -1 not in self.groups) and pd.isnull(self.woes[-1]):
                del self.woes[-1]
            elif (-1 in self.groups and -1 not in self.woes) and pd.isnull(self.groups[-1]):
                del self.groups[-1]
            else:
                print ('WARNING! Number of groups is not certain! We have', len(self.groups), 'groups and', len(self.woes), 'woes!')

        if plot_flag:
            self.plot_woe()

        print ('\n Final WOEs')
        self.print_woe()

        if not no_history:
            self.update_gini_history(cv = cv, alpha_classifier=alpha_classifier, scoring = scoring)

        if len(self.gini_history['Train']) > 1 and plot_flag:
            self.plot_gini()



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    def merge(self, groups_list, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None, no_backup=False, scoring='neg_log_loss'):
        '''
        Merges two WOE intervals

        Parameters
        -----------
        groups_list: [group1, group2] - the groups to be merged
        alpha: alpha parameter for the WOE of the new group
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        no_backup: should WoE merge be considered as manual change and be backed up or not
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''

        # edited 15.08.2018 by Yudochev Dmitry
        if simple is None and self.simple is not None:
            simple=self.simple
        if n_folds is None and self.n_folds is not None:
            n_folds=self.n_folds
        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if cv is None and self.cv is not None:
            cv=self.cv
        # Checks for correctness of the groups to merge
        # existing groups
        for group in groups_list:
            if group not in self.groups:
                print ('Es gibt keine', group, '. Viel Gluek.')
                return None
        # only 2 groups per merge call
        if len(groups_list) != 2:
            print ('Please enter 2 groups.')
            return None


        # only neighbouring groups for ordered features
        if not self.categorical:
            if -1 not in groups_list:
                if self.groups[groups_list[0]][0] != self.groups[groups_list[1]][1] and self.groups[groups_list[1]][0] != self.groups[groups_list[0]][1]:
                    print ('Please enter neighbouring groups. Good luck.')
                    return None

        if no_backup==False:
            self.to_history()

        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        # merging groups in self.groups
        min_group = min(groups_list)
        max_group = max(groups_list)

        if not isinstance(self.groups[min_group], list):
            self.missing_group = max_group
            del self.groups[min_group]
        else:
            if self.categorical:
                self.groups[min_group] = self.groups[min_group] + self.groups[max_group]
                # commented by Anna Goreva
                #del self.groups[max_group]
            else:
                self.groups[min_group] = [min(self.groups[min_group][0], self.groups[max_group][0]), max(self.groups[min_group][1], self.groups[max_group][1])]

            #edited 25.09.2018 by Yudochev Dmitry - union with non-missing groups still can affect self.missing_group
            if self.missing_group==max_group:
                self.missing_group=min_group
            elif self.missing_group>max_group:
                self.missing_group=self.missing_group-1

            shift_flag = False
            new_groups = copy.deepcopy(self.groups)
            for group in sorted(new_groups):
                if group == max_group:
                    shift_flag = True
                elif group > max_group:
                    shift_flag = True
                    self.groups[group - 1] = self.groups[group].copy()
            if shift_flag:
                del self.groups[max(self.groups)]

        print (self.groups)
        pre_woes_backup = list(self.data.dataframe[self.feature])
        self.data.dataframe[self.feature] = self.data.dataframe[self.feature + '_raw'].copy()

        self.fit(alpha = alpha, new_groups=False, alpha_range = None, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=False, cv=cv, scoring=scoring)

        gc.collect()

        self.data.dataframe[self.feature] = pre_woes_backup
        if plot_flag:
            self.plot_woe()
            if len(self.gini_history['Train']) > 1:
                self.plot_gini()



    def to_history(self):
        '''
        TECH

        Writes current state of self to history.
        '''
        self.groups_history.append(dict(missing_group=self.missing_group, groups=copy.deepcopy(self.groups)))
        self.woes_history.append(self.woes.copy())



    def insert_subgroup(self, group, new_group, alpha, simple, woe_adjust, n_folds, scoring='neg_log_loss'):
        '''
        TECH

        Makes new group and calculates WOE for categorical features

        Parameters
        --------------
        new_group: a user-defined new group of values, consists of values from the group to split and the other values of the group will be separated. Only for categorical features. Example: group = [1, 2, 4, 6, 9], new_group = [1, 2, 9] => the two new groups will be [1, 2, 9], [4, 6]. For the same result we could set new_group parameter = [4, 6]
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        pre_woes_backup = list(self.data.dataframe[self.feature])

        self.to_history()
        self.groups[max(self.groups) + 1] = new_group
        self.groups[group] = [i for i in self.groups[group] if i not in new_group]
        self.data.dataframe[self.feature] = self.data.dataframe[self.feature + '_raw'].copy()

        self.fit(new_groups=False, alpha = alpha, alpha_range = None, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=True, scoring=scoring)
        self.data.dataframe[self.feature] = pre_woes_backup



    def insert_new_bound(self, group, add_bound, alpha, simple, woe_adjust, n_folds, scoring='neg_log_loss'):
        '''
        TECH

        Inserts new bound and calculates WOE for interval features

        Parameters
        ------------
        group: the group to insert the new bound into
        add_bound: the new bound to insert
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        self.to_history()

        new_group_num = group + 1

        tmp_groups = copy.deepcopy(self.groups)
        tmp_woes = copy.deepcopy(self.woes)
        for g in sorted(self.groups):
            if g > new_group_num:
                tmp_groups[g] = self.groups[g - 1].copy()
                tmp_woes[g] = self.woes[g - 1]
        tmp_groups[max(self.groups) + 1] = self.groups[max(self.groups)].copy()
        tmp_woes[max(self.woes) + 1] = self.woes[max(self.woes)]
        if self.missing_group>group:
            self.missing_group=self.missing_group+1
        self.groups = tmp_groups
        self.woes = tmp_woes
        self.groups[new_group_num] = [add_bound, self.groups[group][1]]
        self.groups[group][1] = add_bound
        self.fit(new_groups=False, alpha = alpha, alpha_range = None, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=True, scoring=scoring)



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    # edited on Nov-21-2018 by Anna Goreva: refactoring
    def split(self, group = None, to_add = None, scoring = 'neg_log_loss', alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None):
        '''
        Splits a WOE interval into two.

        Parameters
        -----------
        group: a group to split, integer
        to_add: in case of interval - a user-defined bound for the split (the intermediate bound of the interval), only for ordered features; in case of categorical -  a user-defined new group of values, consists of values from the group to split and the other values of the group will be separated. Only for categorical features. Example: group = [1, 2, 4, 6, 9], new_group = [1, 2, 9] => the two new groups will be [1, 2, 9], [4, 6]. For the same result we could set new_group parameter = [4, 6]
        scoring: a scoring metric to use
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        '''

        # edited 15.08.2018 by Yudochev Dmitry
        if simple is None and self.simple is not None:
            simple=self.simple
        if n_folds is None and self.n_folds is not None:
            n_folds=self.n_folds
        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if cv is None and self.cv is not None:
            cv=self.cv

        # edited 22.08.2018 by Yudochev Dmitry - same as above
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        if group == -1:
            print ('Invalid group!')
            return None


        if isinstance(to_add, int) or isinstance(to_add, float):
            if self.categorical:
                print ('The feature is categorical so to_add should be a list of values for the new group. Good luck!')
                return None
            elif not (to_add >= self.groups[group][0] and to_add < self.groups[group][1]):
                print('New bound is out-of-range for the specified group. Bye-bye.')
                return None
            else:
                self.insert_new_bound(group, to_add, alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, scoring=scoring)
        elif isinstance(to_add, list) or isinstance(to_add, np.ndarray):
            if not self.categorical:
                print ('The feature is not categorical so to_add must be a float. Good luck!')
                return None
            else:
                for n in to_add:
                    if n not in self.groups[group]:
                        print('Invalid new_group!')
                        return None
                if self.groups[group] == to_add:
                    print ('Error: new_group contains all the values of group', group)
                    return None
                else:
                    self.insert_subgroup(group, to_add, alpha, simple, woe_adjust, n_folds, scoring=scoring)

                    gc.collect()

        # if no pre-defined bounds or groups
        else:
            print ('Splitting started! Feature', self.feature, 'group:', group)
            if self.categorical:
                samples_to_process = self.data.dataframe.loc[self.data.dataframe[self.feature + '_raw'].apply(lambda x: x in self.groups[group])].copy()
            else:
                samples_to_process = self.data.dataframe.loc[self.data.dataframe[self.feature] >= self.groups[group][0]].loc[self.data.dataframe[self.feature] < self.groups[group][1]].copy()

            if len(samples_to_process[self.data.target].unique())>1:
                parameters_grid = {'min_samples_split': [int(round(self.data.dataframe.shape[0]*i, 0)) for i in [.01, .001, .0001]], 'max_depth' : [1]}

                tmp_categorical = self.categorical
                self.categorical = False
                # optimal split
                try:
                    final_tree = self.fit_gridsearch(samples_to_process[self.feature], cv, scoring, parameters_grid, samples_to_process[self.data.target])
                    final_tree.fit(samples_to_process[self.feature][:, None], samples_to_process[self.data.target],
                                   sample_weight = None if self.data.weights is None else np.array(samples_to_process[self.data.weights]))

                except Exception:
                    print ('Fitting with cross-validation failed! Possible cause: too few representatives of one of the target classes.')
                    print ('Try setting the bound yourself')
                    self.categorical = tmp_categorical
                    return None


                tree_splits = self.get_tree_splits(final_tree)
                #corrected by Yudochev Dmitry 13.09.2018 - getting thresholds by comparing children of each node
                #children_left = final_tree.tree_.children_left
                #children_right = final_tree.tree_.children_right
                #threshold = final_tree.tree_.threshold

                #tree_splits = []


                #for i in range(final_tree.tree_.node_count):
                #    if children_left[i] != children_right[i]:
                #        tree_splits.append(threshold[i])

                self.categorical = tmp_categorical
                # if no splits found
                if len(tree_splits) == 0:
                    print ('No good binning found :(')
                    print ('Try setting the bound yourself')
                    #self.categorical = tmp_categorical
                    return None
                else:
                    add_bound = sorted(tree_splits)[0]
                    print ('Additional bound ', add_bound)
                    #adding the new bound (woe for categorical) to groups
                    if self.categorical:
                        # find group by woe bound...
                        # since in self.data.dataframe[feature] we have woes calculated for each categorical value...
                        new_group = list(self.data.dataframe.loc[self.data.dataframe[self.feature] < add_bound].loc[self.data.dataframe[self.feature + '_raw'].apply(lambda x: x in self.groups[group])][self.feature + '_raw'].drop_duplicates())
                        print ('new_group:', new_group)
                        self.insert_subgroup(group, new_group, alpha, simple, woe_adjust, n_folds, scoring=scoring)
                    else:
                        self.insert_new_bound(group, add_bound, alpha = alpha, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, scoring=scoring)

                    gc.collect()
            else:
                print('All observations in the specified group have the same target value =', samples_to_process[self.data.target].unique()[0])



    def merge_by_woe(self, woe_threshold=0.05, with_missing=True, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None, no_backup=False, scoring='neg_log_loss'):
        '''
        Merges all groups, close by WOE (for interval features only neighboring groups and missing group are checked)

        Parameters
        -----------
        woe_threshold: if woe difference between groups (neighboring groups for interval) is less then this threshold, then they are to be merged
        with_missing: should woe difference with missing group also be checked
        alpha: alpha parameter for the WOE of the new group
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        no_backup: should WoE merge be considered as manual change and be backed up or not
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''

        # edited 15.08.2018 by Yudochev Dmitry
        if simple is None and self.simple is not None:
            simple=self.simple
        if n_folds is None and self.n_folds is not None:
            n_folds=self.n_folds
        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if cv is None and self.cv is not None:
            cv=self.cv

        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        #print('Checking', self.feature)

        if len([x for x in self.woes if pd.isnull(self.woes[x])==False])>1:
            to_check_woe=True
        else:
            to_check_woe=False
            print('... Merge by WoE: only 1 group is present')

        while to_check_woe:
            to_check_woe=False

            groups_dna=[x for x in self.woes if x!=-1]
            min_woe_dif=None
            if with_missing and -1 in self.woes:
                if self.woes[-1] is not None:
                    if min_woe_dif is None or abs(self.woes[groups_dna[0]]-self.woes[-1])<min_woe_dif:
                        min_woe_dif=abs(self.woes[groups_dna[0]]-self.woes[-1])
                        min_group=-1
                        max_group=groups_dna[0]
            for i in range(len(groups_dna)-1):
                if min_woe_dif is None or abs(self.woes[groups_dna[i]]-self.woes[groups_dna[i+1]])<min_woe_dif:
                    min_woe_dif=abs(self.woes[groups_dna[i]]-self.woes[groups_dna[i+1]])
                    min_group=groups_dna[i]
                    max_group=groups_dna[i+1]
                if with_missing and -1 in self.woes:
                    if self.woes[-1] is not None:
                        if min_woe_dif is None or abs(self.woes[groups_dna[i+1]]-self.woes[-1])<min_woe_dif:
                            min_woe_dif=abs(self.woes[groups_dna[i+1]]-self.woes[-1])
                            min_group=-1
                            max_group=groups_dna[i+1]

            if min_woe_dif is not None and min_woe_dif<woe_threshold:
                print('... Merge by WoE: merging', min_group, 'and', max_group, 'groups with WoE difference =', min_woe_dif)
                to_check_woe=True

                self.merge([min_group, max_group], alpha = alpha, simple = simple, woe_adjust = woe_adjust, n_folds = n_folds, plot_flag = plot_flag, cv=cv,
                           no_backup = no_backup, scoring=scoring)
            else:
                print('... Merge by WoE: no groups to be merged were found')



    def merge_by_size(self, target_threshold=5, size_threshold=100, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None, no_backup=False, scoring='neg_log_loss'):
        '''
        Merges small groups (by target or size) to the closest by WoE (for interval features only neighboring groups and missing group are checked)

        Parameters
        -----------
        target_threshold: min number of targets for group to not be considered small
        size_threshold: min number of observations for group to not be considered small
        alpha: alpha parameter for the WOE of the new group
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        no_backup: should WoE merge be considered as manual change and be backed up or not
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''

        # edited 15.08.2018 by Yudochev Dmitry
        if simple is None and self.simple is not None:
            simple=self.simple
        if n_folds is None and self.n_folds is not None:
            n_folds=self.n_folds
        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if cv is None and self.cv is not None:
            cv=self.cv

        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        #print('Checking', self.feature)

        if len([x for x in self.woes if pd.isnull(self.woes[x])==False])>1:
            to_check_size=True
        else:
            to_check_size=False
            print('... Merge by size: only 1 group is present')

        while to_check_size:
            to_check_size=False

            woes_dna={x:self.woes[x] for x in self.woes if self.woes[x] is not None and (x!=-1 or self.categorical)}
            if len(woes_dna)<=1:
                print('... Merge by size: only 1 non-missing group is present')
                break

            if self.data.weights is not None:
                group_stats=self.data.dataframe[[self.feature+'_WOE', self.data.target, self.data.weights]]
                group_stats[self.data.target]=group_stats[self.data.target]*group_stats[self.data.weights]
                group_stats=group_stats.groupby(self.feature+'_WOE').sum().rename({self.data.target:'target', self.data.weights:'amount'}, axis=1)
            else:
                group_stats=self.data.dataframe[[self.feature+'_WOE', self.data.target]].groupby(self.feature+'_WOE').agg(['sum', 'size'])
                group_stats.columns=group_stats.columns.droplevel()
                group_stats=group_stats.rename({'sum':'target', 'size':'amount'}, axis=1)

            targets={x:group_stats.loc[self.woes[x]]['target'] for x in sorted(woes_dna, key=woes_dna.get if self.categorical else None) if x!=-1 or self.categorical}
            amounts={x:group_stats.loc[self.woes[x]]['amount'] for x in sorted(woes_dna, key=woes_dna.get if self.categorical else None) if x!=-1 or self.categorical}

            #display(group_stats)
            #display(targets)
            #display(amounts)

            min_woe_dif=None

            '''
            for i in range(len(targets)):
                if (target_threshold>=1 and targets[list(targets.keys())[i]]<target_threshold) or \
                (target_threshold<1 and targets[list(targets.keys())[i]]/amounts[list(targets.keys())[i]]<target_threshold) or \
                amounts[list(targets.keys())[i]]<size_threshold:
                    if i==0:
                        nearest_group=list(targets.keys())[i+1]
                        woe_dif=abs(self.woes[list(targets.keys())[i]]-self.woes[list(targets.keys())[i+1]])
                    elif i==len(targets)-1:
                        nearest_group=list(targets.keys())[i-1]
                        woe_dif=abs(self.woes[list(targets.keys())[i]]-self.woes[list(targets.keys())[i-1]])
                    else:
                        if abs(self.woes[list(targets.keys())[i]]-self.woes[list(targets.keys())[i+1]])<abs(self.woes[list(targets.keys())[i]]-self.woes[list(targets.keys())[i-1]]):
                            nearest_group=list(targets.keys())[i+1]
                            woe_dif=abs(self.woes[list(targets.keys())[i]]-self.woes[list(targets.keys())[i+1]])
                        else:
                            nearest_group=list(targets.keys())[i-1]
                            woe_dif=abs(self.woes[list(targets.keys())[i]]-self.woes[list(targets.keys())[i-1]])
                    if min_woe_dif is None or min_woe_dif>woe_dif:
                        group_from=list(targets.keys())[i]
                        group_to=nearest_group
                        min_woe_dif=woe_dif
            '''
            targets_list = list(targets.keys())

            for i, t in enumerate(targets):
                if (target_threshold>=1 and targets[t]<target_threshold) or \
                (target_threshold<1 and targets[t]/amounts[t]<target_threshold) or \
                amounts[t]<size_threshold:
                    if i==0:
                        nearest_group=targets_list[i+1]
                        woe_dif=abs(self.woes[t]-self.woes[targets_list[i+1]])
                    elif i==len(targets)-1:
                        nearest_group=targets_list[i-1]
                        woe_dif=abs(self.woes[t]-self.woes[targets_list[i-1]])
                    else:
                        if abs(self.woes[t]-self.woes[targets_list[i+1]])<abs(self.woes[t]-self.woes[targets_list[i-1]]):
                            nearest_group=targets_list[i+1]
                            woe_dif=abs(self.woes[t]-self.woes[targets_list[i+1]])
                        else:
                            nearest_group=targets_list[i-1]
                            woe_dif=abs(self.woes[t]-self.woes[targets_list[i-1]])
                    if min_woe_dif is None or min_woe_dif>woe_dif:
                        group_from=t
                        group_to=nearest_group
                        min_woe_dif=woe_dif



            if min_woe_dif is not None:
                print('... Merge by size: merging', group_from, 'and', group_to, 'groups with WoE difference =', min_woe_dif)
                to_check_size=True

                self.merge([group_from, group_to], alpha = alpha, simple = simple, woe_adjust = woe_adjust, n_folds = n_folds, plot_flag = plot_flag, cv=cv,
                           no_backup = no_backup, scoring=scoring)
            else:
                print('... Merge by size: no groups to be merged were found')



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    #added 22.08.2018 by Yudochev Dmitry
    def missing_unite(self, missing_process = None, alpha = None, simple=None, woe_adjust=None, n_folds = None,
                         missing_min_part = None, plot_flag=True, cv=None, scoring='neg_log_loss'):
        '''
        Adds missing values to a bin based on chosen logic (from missing_process).

        Parameters
        -----------
        missing_process: way of handling missing values
            'separate' - keep missing values in a separate bin
            'worst' - assign missing values to the worst bin
            'nearest' - assign missing values to the nearest bin by WoE
            'worst_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the worst bin, otherwise keep separate
            'best_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the best bin, otherwise keep separate
            'nearest_or_separate' - if missing values part in sample is less then missing_min_part, the assign to the nearest bin, otherwise keep separate
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        missing_min_part: threshold for missing values part in case of missing_process=='worst_or_separate', float between 0 and 1.
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        if simple is None and self.simple is not None:
            simple=self.simple
        if n_folds is None and self.n_folds is not None:
            n_folds=self.n_folds
        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if cv is None and self.cv is not None:
            cv=self.cv
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        #edited by Anna Goreva Oct-04-2018
        if missing_process is None:
            print ('You did not set the missing_process parameter!')
            if self.missing_process is not None:
                missing_process=self.missing_process
            else:
                print ('Error! Please use "separate", "worst", "nearest", "worst_or_separate", "best_or_separate" or "nearest_or_separate". Bye.')
                return None
        if missing_min_part is None and self.missing_min_part is not None:
            missing_min_part=self.missing_min_part

        if self.data.dataframe[self.feature].isnull().any():
            if self.missing_group != -1:
                print ('Missings are in group number', self.missing_group)
                print ('Please use missing_separate() first. Good luck!')
                return None

            elif missing_process == 'separate':
                print ('Missings are already separate, please "worst", "nearest", "worst_or_separate", "best_or_separate" or "nearest_or_separate" for missing_process parameter.')
                return None

            else:
                self.to_history()

                if missing_process=='worst':
                    woes_dna={x:self.woes[x] for x in self.woes if x!=-1}
                    self.missing_group=min(woes_dna, key=woes_dna.get)
                    print ('worst group', self.missing_group)
                elif missing_process=='nearest':
                    woes_dna_diff={x:abs(self.woes[x]-self.woes[-1]) for x in self.woes if x!=-1}
                    self.missing_group=min(woes_dna_diff, key=woes_dna_diff.get)
                    print ('nearest group', self.missing_group)
                elif missing_process=='worst_or_separate':
                    if (self.data.weights is None and self.data.dataframe[self.feature].isnull().value_counts()[True]/self.data.dataframe[self.feature].shape[0]<missing_min_part) or \
                        (self.data.weights is not None and self.data.dataframe[self.data.dataframe[self.feature].isnull()][self.data.weights].sum()/self.data.dataframe[self.data.weights].sum()<missing_min_part):
                        woes_dna={x:self.woes[x] for x in self.woes if x!=-1}
                        self.missing_group=min(woes_dna, key=woes_dna.get)
                    print ('worst or separate', self.missing_group)
                elif missing_process=='nearest_or_separate':
                    if (self.data.weights is None and self.data.dataframe[self.feature].isnull().value_counts()[True]/self.data.dataframe[self.feature].shape[0]<missing_min_part) or \
                        (self.data.weights is not None and self.data.dataframe[self.data.dataframe[self.feature].isnull()][self.data.weights].sum()/self.data.dataframe[self.data.weights].sum()<missing_min_part):
                        woes_dna_diff={x:abs(self.woes[x]-self.woes[-1]) for x in self.woes if x!=-1}
                        self.missing_group=min(woes_dna_diff, key=woes_dna_diff.get)
                    print ('nearest or separate', self.missing_group)
                elif missing_process=='best_or_separate':
                    if (self.data.weights is None and self.data.dataframe[self.feature].isnull().value_counts()[True]/self.data.dataframe[self.feature].shape[0]<missing_min_part) or \
                        (self.data.weights is not None and self.data.dataframe[self.data.dataframe[self.feature].isnull()][self.data.weights].sum()/self.data.dataframe[self.data.weights].sum()<missing_min_part):
                        woes_dna={x:self.woes[x] for x in self.woes if x!=-1}
                        self.missing_group=max(woes_dna, key=woes_dna.get)
                    print ('best or separate', self.missing_group)

                if self.missing_group != -1 and -1 in self.groups:
                    del self.groups[-1]
                    if -1 in self.woes:
                        del self.woes[-1]
                self.fit(new_groups=False, alpha = alpha, alpha_range = None, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=plot_flag, cv=cv, scoring=scoring)

                gc.collect()



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    #added 22.08.2018 by Yudochev Dmitry
    def missing_separate(self, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, cv=None, scoring='neg_log_loss'):
        '''
        Separates missing values from any bin they were assigned to.

        Parameters
        -----------
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        cv: the numbder of folds for cross-validation used for optimal WOE splitting
        scoring: a measure for cross-validation used used for optimal WOE splitting
        '''
        if simple is None and self.simple is not None:
            simple=self.simple
        if n_folds is None and self.n_folds is not None:
            n_folds=self.n_folds
        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if cv is None and self.cv is not None:
            cv=self.cv
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        self.to_history()

        self.missing_group = -1
        if -1 not in self.groups:
            if self.categorical:
                self.groups[-1] = [np.nan]
            else:
                self.groups[-1] = np.nan
        if -1 not in self.woes:
            self.woes[-1] = None

        self.fit(new_groups=False, alpha = alpha, alpha_range = None, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=plot_flag, cv=cv, scoring=scoring)

        gc.collect()



    #edited 29.08.2018 by Anna Goreva: changed bounds to groups
    # GAI set original_values to True 10/08/2018
    def transform(self, data, original_values=True, not_found='worst'):
        '''
        Transforms a Data object according to WOE parameters fitted. Can be used only after .fit().

        Parameters
        ------------
        data: Data object to transform
        original_values: should we calculate WoE by original values (not transformed to pre_woes) for categorical features
        not_found: what woe to return for unknown categorical values; 'worst' - the lower woe, 'missing' - woe of group that contains missings, int - number of certain group

        Returns
        ----------
        transformed Data object
        '''
        #corrected (+.copy()) 13.08.2018 by Yudochev Dmitry
        data_to_transform = Data(data.dataframe.copy(),  data.target, self.feature + '_WOE', weights = data.weights, name=data.name)
        #added variants for original values (in <feature>_RAW or just <feature> column)
        if original_values and self.categorical:
            #if self.feature + '_raw' in data_to_transform.dataframe.columns:
            data_to_transform.dataframe[self.feature + '_WOE'] = self.set_avg_woes(data=(data_to_transform.dataframe[self.feature + '_raw'] \
                                                                                           if self.feature + '_raw' in data_to_transform.dataframe.columns \
                                                                                           else data_to_transform.dataframe[self.feature]),
                                                                                   original_values=True, not_found=not_found)
                #data_to_transform.dataframe[self.feature + '_raw'].apply(lambda x: self.change_to_woe(x, self.groups, self.woes, True, not_found=not_found))
            #else:
                #data_to_transform.dataframe[self.feature + '_WOE'] = self.set_avg_woes(data=data_to_transform.dataframe[self.feature],
                #                                                                       original_values=True, not_found=not_found)
                #data_to_transform.dataframe[self.feature].apply(lambda x: self.change_to_woe(x, self.groups, self.woes, True, not_found=not_found))
        else:
            data_to_transform.dataframe[self.feature + '_WOE'] = self.set_avg_woes(data=data_to_transform.dataframe[self.feature], not_found=not_found)
            #data_to_transform.dataframe[self.feature].apply(lambda x: self.change_to_woe(x, self.groups, self.woes, not_found=not_found))

        return data_to_transform



    def groups_to_dataframe(self):
        '''
        TECH

        Transforms self.groups, self.categorical and self.missing_group to dataframe.

        Returns
        ----------
        dataframe with binning information
        '''

        export_groups=[]
        # searching for WOE for each interval of values
        for group in [x for x in self.woes if self.woes[x] is not None]:
            export_groups.append(dict(feature=self.feature,
                                      categorical=self.categorical,
                                      group=group,
                                      values=self.groups[group],
                                      missing=(group==self.missing_group)*1,
                                      woe=self.woes[group],
                                      changed=len(self.groups_history)))

        return pd.DataFrame(export_groups)



    def show_history(self, figsize=(6,5)):
        '''
        Shows groups and gini history, producing WoE and Gini graphs

        Parameters
        ------------
        figsize: size for one plot; size for complex graph will be calculated as figsize[0]*2 for width and
            figsize[1]*number of rows for height
        '''
        g2=2
        g1=int((len(self.groups_history)+1)/g2)+((len(self.groups_history)+1)%g2>0)

        f, axes = plt.subplots(g1, g2, figsize=(figsize[0]*g2,figsize[1]*(g1 if g1>1 else 0.8)), gridspec_kw=dict(wspace=0.3, hspace=0.35))
        if g1==1:
            axes=np.array([axes])
        for h in range(len(self.groups_history)+1):
            if h!=len(self.groups_history):
                current_groups=self.groups_history[h]['groups']
                current_missing_group=self.groups_history[h]['missing_group']
                current_woes=self.woes_history[h]
            else:
                current_groups=self.groups
                current_missing_group=self.missing_group
                current_woes=self.woes

            with plt.style.context(('seaborn-deep')):

                woes_df=pd.DataFrame(current_woes, index=['woe']).transpose().reset_index().rename({'index':'group'},axis=1)#.sort_values('bounds')
                woes_df['bounds'] = woes_df.group.apply(lambda x: current_groups[x])

                #display(woes_df)
                if not self.categorical:
                    woes_df.loc[woes_df.group != -1, 'bounds'] = woes_df.loc[woes_df.group != -1, 'bounds'].apply(lambda x: '[' + str(round(x[0], 2)) + ', ' + str(round(x[1], 2)) + ')')
                if current_missing_group != -1 and not self.categorical:
                    woes_df.loc[woes_df.group == current_missing_group, 'bounds'] = woes_df.loc[woes_df.group == current_missing_group, 'bounds'] + '*'
                if self.categorical:
                    woes_df=woes_df[pd.isnull(woes_df.woe)==False].sort_values('woe')

                woes_df['label']=woes_df['bounds'].apply(lambda x: str(x)[:30]+'..' if len(str(x))>30 else str(x)) + '\n group ' + woes_df.group.apply(lambda x: str(x) + '\n with missings' if current_missing_group == x else str(x))
                to_calc=self.data.dataframe.copy()
                if h!=len(self.groups_history):

                    to_calc['group']=self.set_groups(data=to_calc[self.feature+'_raw'],
                                                     groups=current_groups, woes=current_woes,
                                                     original_values = True, not_found='worst')
                        #to_calc[self.feature+'_raw'].apply(
                                                    #lambda x: self.change_to_group(
                                                        #x, current_groups, current_woes,
                                                            #original_values = True, not_found='worst'))
                if self.data.weights is None:
                    stats=to_calc[['group', self.data.target]].groupby('group').agg(['sum', 'size'])
                else:
                    to_calc[self.data.target]=to_calc[self.data.target]*to_calc[self.data.weights]
                    stats=to_calc[['group', self.data.target, self.data.weights]].groupby('group').sum()
                stats.columns=['target', 'amount']
                woes_df=woes_df.merge(stats.reset_index(), on='group')

                axes[int(h/g2), h%g2].set_ylabel('Observations')
                axes[int(h/g2), h%g2].set_xticks(range(woes_df.shape[0]))
                axes[int(h/g2), h%g2].set_xticklabels(woes_df['label'], rotation=30, ha="right")

                axes[int(h/g2), h%g2].bar(range(woes_df.shape[0]), woes_df['amount'], zorder=0)
                axes[int(h/g2), h%g2].annotate('Targets:', xy=(0, 1), xycoords=('axes fraction', 'axes fraction'), xytext=(-42, 5), textcoords='offset pixels', color='red', size=11)
                if h!=len(self.groups_history):
                    iteration_text='Iteration '+str(h)+' (Gini='+str(round(self.gini_history['Train'][h],3))+')'
                else:
                    iteration_text='Actual (Gini='+str(round(self.gini_history['Train'][h],3))+')'
                axes[int(h/g2), h%g2].annotate(iteration_text,
                                     xy=(0.5, 1), xycoords=('axes fraction', 'axes fraction'),
                                     xytext=(0, 20), textcoords='offset pixels', color='black', ha='center', size=12)
                for i in range(woes_df.shape[0]):
                    axes[int(h/g2), h%g2].annotate(str(woes_df['target'][i]), xy=(i, 1), xycoords=('data', 'axes fraction'),
                                xytext=(0, 5), textcoords='offset pixels', color='red', size=11, ha='center')
                axes[int(h/g2), h%g2].grid(False)
                axes[int(h/g2), h%g2].grid(axis='y', zorder=1, alpha=0.6)
                ax2 = axes[int(h/g2), h%g2].twinx()
                ax2.set_ylabel('WoE')
                ax2.grid(False)
                # red is for the WOE values
                ax2.plot(range(woes_df.shape[0]), woes_df['woe'], 'ro-', linewidth=2.0, zorder=4)
                #plt.suptitle(s.feature, fontsize = 16)
                #f.autofmt_xdate()
        if h%2==0:
            axes[-1, -1].set_visible(False)
        axes[0,0].annotate(self.feature, (1, 1), xycoords='axes fraction', ha='center',
                           xytext=(45, 40), textcoords='offset pixels', fontsize = 16)
        f.tight_layout()
        plt.show()

        if len(self.gini_history['Train']) > 1:
            fig = plt.figure(figsize=(13,3))
            ax = fig.add_subplot(111)
            for sample in self.gini_history:
                if len(self.gini_history[sample])>0:
                    ax.plot([i for i, s in enumerate(self.gini_history[sample])], self.gini_history[sample], label=sample)
            ax.legend()
            plt.suptitle(self.feature+' Gini changes', fontsize = 14)
            plt.show()



    def rollback(self, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, iteration=None):
        '''
        Rolls back the last operation.

        Parameters
        -----------
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        iteration: number of groups iteration to return to (if None, then rollback to the previous iteration)
        '''
        if simple is None and self.simple is not None:
            simple=self.simple
        if n_folds is None and self.n_folds is not None:
            n_folds=self.n_folds
        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        # edited by Anna Goreva 01-Nov-2018
        if iteration is None and len(self.groups_history)>0:
            self.missing_group = self.groups_history[-1]['missing_group']
            self.groups = self.groups_history[-1]['groups']
            self.woes = self.woes_history[-1]
            self.groups_history=copy.deepcopy(self.groups_history[:-1])
            self.woes_history=copy.deepcopy(self.woes_history[:-1])

            for sample in self.gini_history:
                if len(self.gini_history[sample])>0:
                    self.gini_history[sample] = self.gini_history[sample][:-1].copy()
            self.fit(new_groups=False, alpha = alpha, alpha_range = None, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=plot_flag, no_history=True)

            gc.collect()
        elif iteration is not None and iteration<len(self.groups_history):
            self.missing_group = self.groups_history[iteration]['missing_group']
            self.groups = self.groups_history[iteration]['groups']
            self.woes = self.woes_history[iteration]
            self.groups_history=copy.deepcopy(self.groups_history[:iteration])
            self.woes_history=copy.deepcopy(self.woes_history[:iteration])

            for sample in self.gini_history:
                if len(self.gini_history[sample])>0:
                    self.gini_history[sample] = self.gini_history[sample][:iteration+1].copy()
            self.fit(new_groups=False, alpha = alpha, alpha_range = None, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=plot_flag, no_history=True)

            gc.collect()
        else:
            print ('Sorry, no changes detected or iteration found. Nothing to rollback.')
            return None



    def check_values(self, s):
        '''
        TECH

        Checks if any string element of the list contains comma. This method is used in parsing imported dataframes with groups, borders and woes.

        Returns
        --------
        False if there is a comma
        '''
        quotes = s.count("'")
        if quotes > 0:
            commas = s.count(',')
            if commas != (quotes/2)-1:
                return False
        return True



    def str_to_list(self, s):
        '''
        TECH

        Parses ['values'] from a dataframe constructed by self.groups_to_dataframe().
        '''
        s = str(s)

        if pd.isnull(s) or s == '[nan]':
            return np.nan
        if self.check_values(s):
            v = (re.split('[\'|"]? *, *[\'|"]?', s[1:-1]))
            if v[0][0] in ("'", '"'):
                v[0]=v[0][1:]
            if v[-1][-1] in ("'", '"'):
                v[-1]=v[-1][:-1]
            return [float(x) if (x[-3:] == 'inf' or (min([y.isdigit() for y in x.split('.')]) and x.count('.') < 2)) else (x if x!='' else np.nan) for x in v]
        else:
            print ('Error in string', s, '! Delete commas from feature values!')
            return None



    def groups_from_dataframe(self, df_in, alpha = None, simple=None, woe_adjust=None, n_folds = None, plot_flag=True, print_flag=True, fit_flag=False):

        '''
        TECH

        Sets self.groups, self.categorical and self.missing_group values from dataframe and calculates woe (by fit).

        Parameters
        ----------
        df_in: a DataFrame with 'categorical', 'group', 'values' and 'missing' fields
        alpha: alpha parameter for the WOE calculation
        simple: should WoE be calculated just on input data without folds or not
        woe_adjust: adjustment parameter for interval's DR: DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust)
        n_folds: number of folds for WOE calculation
        plot_flag: should bins and Gini changes be plotted or not
        print_flag: should bins and woes be printed or not
        fit_flag: should woes be calculated or taken from input dataframe
        '''

        if simple is None and self.simple is not None:
            simple=self.simple
        if n_folds is None and self.n_folds is not None:
            n_folds=self.n_folds
        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        values = list(df_in['values'])
        groups = list(df_in['group'])
        missing = list(df_in.missing)
        if 'woe' not in df_in:
            fit_flag=True

        if 'categorical' in df_in:
            categorical = list(df_in.categorical)[0]
        else:
            categorical = self.categorical

        to_convert=False
        for v in values:
            if isinstance(v, str):
                to_convert=True

        missing_group=-1

        if isinstance(categorical, str):
            categorical=(categorical in ['True', 'ИСТИНА'])

        if categorical==False:
            if to_convert:
                values_corrected=[]
                for v in values:
                    if str(v)[0]=='[':
                        values_corrected.append([float(x) if x!='' else np.nan for x in str(v)[1:-1].replace(" ", "").split(',')])
                    else:
                        values_corrected.append(float(v))
                values=values_corrected.copy()
            groups_dict={-1:np.nan}

        else:
            if to_convert:
                values_corrected=[]
                for v in values:
                    if self.feature+'_raw' in self.data.dataframe:
                        check_type=self.feature+'_raw'
                    else:
                        check_type=self.feature
                    if self.data.dataframe[check_type].dtype==object:
                        v_=re.split('[\'|"] *, *[\'|"]', v[1:-1])
                    else:
                        v_=re.split(' *, *', v[1:-1])

                    if v_[0][0] in ("'", '"'):
                        v_[0]=v_[0][1:]
                    if v_[-1][-1] in ("'", '"'):
                        v_[-1]=v_[-1][:-1]
                    values_corrected.append(np.array(v_).astype(self.data.dataframe[check_type].dtype).tolist())
                values=values_corrected.copy()
            groups_dict={-1:[np.nan]}


        for i, v in enumerate(values):#range(len(values)):
            groups_dict[groups[i]]=v
            if missing[i]==1:
                missing_group=groups[i]
                if missing_group!=-1:
                    del groups_dict[-1]

        #print('groups_dict=',groups_dict)
        #groups_dict={x:groups_dict[x] for x in sorted(groups_dict)}
        groups_dict={k:v for k, v in sorted(groups_dict.items())}


        if self.feature + '_raw' in self.data.dataframe:
            self.data.dataframe[self.feature]=self.data.dataframe[self.feature + '_raw']

        self.categorical=categorical
        self.groups=groups_dict
        self.missing_group=missing_group

        if fit_flag:
            self.fit(new_groups=False, alpha = alpha, alpha_range = None, simple=simple, woe_adjust=woe_adjust, n_folds=n_folds, plot_flag=plot_flag)
        else:
            self.woes = {}
            for group in self.groups:
                if df_in[df_in.group == group].shape[0]==0:
                    self.woes[group] = None
                else:
                    self.woes[group] = df_in[df_in.group == group].woe.values[0]
            self.data.dataframe[self.feature + '_raw'] = self.data.dataframe[self.feature]
            self.data.dataframe['group'] = self.set_groups(woes=self.woes, original_values = True)
            df_transformed = self.transform(self.data)
            self.data = df_transformed
            if plot_flag:
                self.plot_woe()
            if print_flag:
                self.print_woe()

        #change categories to their WoE values
        if self.categorical:
            self.groups = {-1: np.nan}
            self.missing_group = -1
            self.categorical_to_interval(alpha, None, None, None, None,
                                        n_folds, None, None, simple, woe_adjust)
        self.groups=groups_dict
        self.missing_group=missing_group