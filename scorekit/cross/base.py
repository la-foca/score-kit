# -*- coding: utf-8 -*-
import pandas as pd
import math as m
import numpy as np
import matplotlib
import matplotlib.colors as colors
import matplotlib.pyplot as plt
from sklearn.tree import DecisionTreeClassifier
import datetime
import gc
import copy
try:
    from tqdm import tqdm_notebook
except Exception:
    print('No tqdm package was found. This package is used for visualizing progress bar during trees fitting. So no amazingly beautiful progress bars for you (although everything will work)..')
    def tqdm_notebook(features, disable=False, desc=''):
        return features

import networkx as nx
from operator import itemgetter
from sklearn.model_selection import ParameterGrid
import openpyxl

from ..data import Data, DataSamples
from .._utils import color_background
from ..processor import BusinessLogicChecker, WOEOrderChecker, GiniChecker
from ..woe import WOE


class Crosses:
    '''
    Crosses generation for groups of features of a specific DataSamples object (using train sample).
    WOE transformation of all/specified features of a specific DataSamples object (using train sample).
    Can be used for trasformation of other Data/DataSamples objects.

    Parameters
    -----------
    datasamples: DataSamples object to build trees
    categorical: user-defined list of categorical features
    '''
    def __init__(self, datasamples,  categorical = None):

        if categorical is None:
            categorical=[]

        self.datasamples = datasamples
        self.categorical = categorical

        # decision_trees - {N: DecisionTree}
        self.decision_trees = {}
        self.excluded_decision_trees = {}
        self.prefix=None



    def auto_fit(self, groups=None, criterion=None, max_depth = None, max_leaf_nodes=None, min_samples_leaf=None, min_targets_leaf=None,
                 missing_process=None, bins=200, bins_detailed=200, bins_type='distr', interval_split_old=False,
                 use_features_once=True, alpha = 0, woe_adjust = 0.5, verbose=True,
                 BL_on=True, WOEO_on=True, G_on=True, to_correct=True, max_corrections=None,
                 BL_dictionary=None, WOEO_er_threshold=0.01, WOEO_correct_threshold=0.85, WOEO_miss_is_incorrect=True,
                 WOEM_on=True, WOEM_woe_threshold=0.05,
                 G_gini_threshold=5, G_gini_decrease_threshold=0.2, G_gini_increase_restrict=True, G_with_test=False,
                 selection_dict={'mean':0.75, 'std':0.75, 'best':'validate'}, out=None):
        '''
        Automatically build trees for each existing Decision Tree in self.decision_trees, going through available
        parameters combinations, checking business logic (if BL_on), WoE order stability (if WOEO_on) and Gini
        stability (if G_on). If to_correct==True, then in case of errors during business logic and WoE order
        attempts are made to correct them by uniting nodes/groups and pruning. After all trees are built and checked
        the best tree is chosen from available correct candidates using logic, defined by selection_dict

        Parameters
        -----------
        groups: a list of lists, containing features to build new trees on
        criterion: a list of criterions used in building tree candidates. Available values are 'gini' and 'entropy'
        max_depth: a list of maximal depth values used in building tree candidates
        max_leaf_nodes: list of leaves number values used in building tree candidates
        min_samples_leaf: a value of minimal observations number (or a minimal part of train sample) for a leaf
            used in building tree candidates
        min_targets_leaf: a value of minimal targets number for a leaf used in building tree candidates
        missing_process: defines how missing values should be treated in case of interval features:
            'separate' - algorithm will always try to separate missing values from other values
            'worst' - algorithm will split non missing values and assign missings to the child with the greatest event rate
            'nearest' - algorithm will split non missing values and assign missings to the child with the closest event rate
            'best' - algorithm will check 'separate' and 'nearest' variants and choose the one with the least criterion value
        bins: a number of bins to initially split feature by
        bins_detailed: a number of bins to split initially found best split's vicinity by (form the split wight before
            the best one to the split right after the best one)
        bins_type: a type of binning process:
            'distr' - bins will be defined by splitting sample on equally sized parts (for now - ignoring weights)
            'scale' - bins will be defined by splitting sample on parts by equally sized scale intervals (ignoring weights)
        interval_split_old: a flag to use the deprecated splitting procedure for interval features (using sklearn's
            DecisionTreeClassifier and ignoring min_targets_leaf option)
        use_features_once: should it be impossible to split by the same feature twice along the path to any leaf or not
        alpha: regularization parameter, used in WoE calculation
        woe_adjust: adjustment parameter for group's ER: EventRate_i = (N_target_i+woe_adjust)/(N_nontarget_i+woe_adjust)
        verbose: log detalization level. If 0 or False only major steps are printed out, in 1 or True, then building trees and
            processing checks is displayed, if 2, then also graphs for checks and trees are visualized
        to_correct: should there be any attempts to correct trees for them to pass the BL and WOEO checks by uniting or
            pruning nodes (both checks) and by uniting groups (for WOEO check in case of uniting nodes wasn't successful)
        max_corrections: maximal number of corrections for each check
        selection_dict: a dictionary, describing the logic used for the best tree selection. This logic is described like
            {<type>: <value>, <type>: <value>, ..., <type>: <value>, 'best': <type>}. Possible types: 'train', 'validate',
            'test' (gini on corresponding samples), 'mean', 'std' (mean and std of gini values on bootstrap samples).
            <type>: <value> selects best trees of remaining candidates by corresponding type, amount is defined by value
            (if value is less then 1, then it is used like a part of available candidates, otherwise it is treated like
            an exact amount). Lastly 'best':<type> is used to get the best tree by corresponfing type.
            Best trees by 'train', 'validate', 'test' and 'mean' have the maximal corresponding values, best trees by 'std'
            has the minimal value of std.
            Example: {'mean':0.75, 'std':0.75, 'best':'validate'} means 'take 75% of candidates with the biggest mean gini
            in bootstrap, then 75% of remaining candidates with the least std in bootstrap and in the end take the best tree
            by gini in validate sample'.

        Business logic check options (BL):
        ---------------------------------
        BL_on: flag to turn on/off business logic check
        BL_dictionary: a DataFrame, a dictionary or a string with a path to conditions dictionary for business logic check
            (in case of a DataFrame or a string the field with features' names should be called 'feature', 'variable' or 'var')

        WoE order check options (WOEO):
        ------------------------------
        WOEO_on: flag to turn on/off WoE order check
        WOEO_er_threshold: if WoE order is not correct, then event rate difference between swaped bins is checked
        WOEO_correct_threshold: what part of checks on bootstrap should be correct for tree to pass the check
        WOEO_miss_is_incorrect: if there is no data for a bin on bootstrap sample, should it be treated as an error or not

        WoE merge options (WOEM):
        -----------------------
        WOEM_on: flag to turn on/off merging by WoE process
        WOEM_woe_threshold: if woe difference between leaves is less then this threshold, then they are to be merged (pruned or groupped)

        Gini check options (G):
        ----------------------
        G_on: flag to turn on/off Gini check
        G_gini_threshold: gini on train and validate/95% bootstrap should be greater then this (gini is multiplied by 100)
        G_gini_decrease_threshold: gini decrease from train to validate/95% bootstrap deviation from mean to mean should be greater then this
        G_gini_increase_restrict: if gini increase should also be restricted (by G_gini_decrease_threshold)
        G_with_test: should trees be also checked on test sample (and be excluded in case of failure)
        '''
        if interval_split_old:
            print("Attention! The old version of interval splitting is now deprecated. It uses sklearn's DecisionTreeClassifier and ignores the min_targets_leaf option!")

        if groups is None:
            cycle_groups=[list(self.datasamples.train.features)]
        else:
            cycle_groups=groups.copy()

        for group in cycle_groups:
            if sum([x not in self.datasamples.train.features for x in group])>0:
                print('No',str([x for x in group if x not in self.datasamples.train.features]), 'features in datasamples.train.features. Skipping group..')
            else:
                dt_number = 0 if self.decision_trees == {} else max(self.decision_trees)+1
                self.decision_trees[dt_number] = DecisionTree(self.datasamples, group, [x for x in group if x in self.categorical])

        if BL_dictionary is None:
            BL_dictionary={}

        if out is None:
            out=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")+'_trees_backup.xlsx'

        times=[]
        for dt_number in self.decision_trees:
            print ('--------------------------------------- Tree #'+str(dt_number), 'processing started! ---------------------------------------')
            current_time=datetime.datetime.now()
            print('-- Starting at:                  '+str(current_time))
            times.append(current_time)
            if len(times)>1:
                time_per_feature=np.mean([times[i+1]-times[i] for i in range(len(times)-1)])
                print('-- Estimated time of completion: '+str(datetime.datetime.now()+time_per_feature*(len(self.decision_trees)-list(self.decision_trees).index(dt_number)))+' (average per tree = '+str(time_per_feature)+')')

            self.decision_trees[dt_number].auto_fit(criterion=criterion, max_depth = max_depth, max_leaf_nodes=max_leaf_nodes,
                                                    min_samples_leaf=min_samples_leaf, min_targets_leaf=min_targets_leaf,
                                                    missing_process=missing_process, bins=bins, bins_detailed=bins_detailed, bins_type=bins_type, interval_split_old=interval_split_old,
                                                    use_features_once=use_features_once, alpha = alpha, woe_adjust = woe_adjust,
                                                    verbose=verbose, to_correct=to_correct, max_corrections=max_corrections,
                                                    BL_on=BL_on, BL_dictionary=BL_dictionary,
                                                    WOEO_on=WOEO_on, WOEO_er_threshold=WOEO_er_threshold,
                                                    WOEO_correct_threshold=WOEO_correct_threshold, WOEO_miss_is_incorrect=WOEO_miss_is_incorrect,
                                                    WOEM_on=WOEM_on, WOEM_woe_threshold=WOEM_woe_threshold,
                                                    G_on=G_on, G_gini_threshold=G_gini_threshold, G_gini_decrease_threshold=G_gini_decrease_threshold,
                                                    G_gini_increase_restrict=G_gini_increase_restrict, G_with_test=G_with_test,
                                                    selection_dict=selection_dict)
            if self.decision_trees[dt_number].tree is None:
                self.exclude(dt_number)
            else:
                self.export_trees(out)
            #gc.collect()
            print ('--------------------------------------- Tree #'+str(dt_number), 'processing ended! ---------------------------------------')
        gc.collect()



    def auto_fit_rest(self, group=None, max_trees=None, criterion=None, max_depth = None, max_leaf_nodes=None, min_samples_leaf=None,
                      min_targets_leaf=None, missing_process=None, bins=200, bins_detailed=200, bins_type='distr', interval_split_old=False,
                      use_features_once=True, alpha = 0, woe_adjust = 0.5, verbose=True,
                      BL_on=True, WOEO_on=True, G_on=True, to_correct=True, max_corrections=None,
                      BL_dictionary=None, WOEO_er_threshold=0.01, WOEO_correct_threshold=0.85, WOEO_miss_is_incorrect=True,
                      WOEM_on=True, WOEM_woe_threshold=0.05,
                      G_gini_threshold=5, G_gini_decrease_threshold=0.2, G_gini_increase_restrict=True, G_with_test=False,
                      selection_dict={'mean':0.75, 'std':0.75, 'best':'validate'}, out=None, correlated_exclusion_binning=True,
                      correlated_exclusion_threshold=0.6):
        '''
        Automatically build trees in cycle, based on features from the input list. Each iteration new tree is generated on
        features, that weren't used for the previous tree. Each tree is built going through available parameters combinations,
        checking business logic (if BL_on), WoE order stability (if WOEO_on) and Gini stability (if G_on). If to_correct==True,
        then in case of errors during business logic and WoE order attempts are made to correct them by uniting nodes/groups
        and pruning. After all trees are built and checked the best tree is chosen from available correct candidates using logic,
        defined by selection_dict.
        Cycle is interrupted if new tree is built only using one feature.

        Parameters
        -----------
        group: a list of features to start building consecutive trees from
        max_trees: a maximal number of trees to build
        criterion: a list of criterions used in building tree candidates. Available values are 'gini' and 'entropy'
        max_depth: a list of maximal depth values used in building tree candidates
        max_leaf_nodes: list of leaves number values used in building tree candidates
        min_samples_leaf: a value of minimal observations number (or a minimal part of train sample) for a leaf
            used in building tree candidates
        min_targets_leaf: a value of minimal targets number for a leaf used in building tree candidates
        missing_process: defines how missing values should be treated in case of interval features:
            'separate' - algorithm will always try to separate missing values from other values
            'worst' - algorithm will split non missing values and assign missings to the child with the greatest event rate
            'nearest' - algorithm will split non missing values and assign missings to the child with the closest event rate
            'best' - algorithm will check 'separate' and 'nearest' variants and choose the one with the least criterion value
        bins: a number of bins to initially split feature by
        bins_detailed: a number of bins to split initially found best split's vicinity by (form the split wight before
            the best one to the split right after the best one)
        bins_type: a type of binning process:
            'distr' - bins will be defined by splitting sample on equally sized parts (for now - ignoring weights)
            'scale' - bins will be defined by splitting sample on parts by equally sized scale intervals (ignoring weights)
        interval_split_old: a flag to use the deprecated splitting procedure for interval features (using sklearn's
            DecisionTreeClassifier and ignoring min_targets_leaf option)
        use_features_once: should it be impossible to split by the same feature twice along the path to any leaf or not
        alpha: regularization parameter, used in WoE calculation
        woe_adjust: adjustment parameter for group's ER: EventRate_i = (N_target_i+woe_adjust)/(N_nontarget_i+woe_adjust)
        verbose: log detalization level. If 0 or False only major steps are printed out, in 1 or True, then building trees and
            processing checks is displayed, if 2, then also graphs for checks and trees are visualized
        to_correct: should there be any attempts to correct trees for them to pass the BL and WOEO checks by uniting or
            pruning nodes (both checks) and by uniting groups (for WOEO check in case of uniting nodes wasn't successful)
        max_corrections: maximal number of corrections for each check
        selection_dict: a dictionary, describing the logic used for the best tree selection. This logic is described like
            {<type>: <value>, <type>: <value>, ..., <type>: <value>, 'best': <type>}. Possible types: 'train', 'validate',
            'test' (gini on corresponding samples), 'mean', 'std' (mean and std of gini values on bootstrap samples).
            <type>: <value> selects best trees of remaining candidates by corresponding type, amount is defined by value
            (if value is less then 1, then it is used like a part of available candidates, otherwise it is treated like
            an exact amount). Lastly 'best':<type> is used to get the best tree by corresponfing type.
            Best trees by 'train', 'validate', 'test' and 'mean' have the maximal corresponding values, best trees by 'std'
            has the minimal value of std.
            Example: {'mean':0.75, 'std':0.75, 'best':'validate'} means 'take 75% of candidates with the biggest mean gini
            in bootstrap, then 75% of remaining candidates with the least std in bootstrap and in the end take the best tree
            by gini in validate sample'.
        correlated_exclusion_binning: a WOE class object used to determine significantly correlated features (with used ones)
            after WOE-transformation to exclude from the rest of features. If omitted no such exclusion is performed.
        correlated_exclusion_threshold: a threshold value for Pearson correlcation coefficient (for absolute value)

        Business logic check options (BL):
        ---------------------------------
        BL_on: flag to turn on/off business logic check
        BL_dictionary: a DataFrame, a dictionary or a string with a path to conditions dictionary for business logic check
            (in case of a DataFrame or a string the field with features' names should be called 'feature', 'variable' or 'var')

        WoE order check options (WOEO):
        ------------------------------
        WOEO_on: flag to turn on/off WoE order check
        WOEO_er_threshold: if WoE order is not correct, then event rate difference between swaped bins is checked
        WOEO_correct_threshold: what part of checks on bootstrap should be correct for tree to pass the check
        WOEO_miss_is_incorrect: if there is no data for a bin on bootstrap sample, should it be treated as an error or not

        WoE merge options (WOEM):
        -----------------------
        WOEM_on: flag to turn on/off merging by WoE process
        WOEM_woe_threshold: if woe difference between leaves is less then this threshold, then they are to be merged (pruned or groupped)

        Gini check options (G):
        ----------------------
        G_on: flag to turn on/off Gini check
        G_gini_threshold: gini on train and validate/95% bootstrap should be greater then this (gini is multiplied by 100)
        G_gini_decrease_threshold: gini decrease from train to validate/95% bootstrap deviation from mean to mean should be greater then this
        G_gini_increase_restrict: if gini increase should also be restricted (by G_gini_decrease_threshold)
        G_with_test: should trees be also checked on test sample (and be excluded in case of failure)
        '''

        if interval_split_old:
            print("Attention! The old version of interval splitting is now deprecated. It uses sklearn's DecisionTreeClassifier and ignores the min_targets_leaf option!")

        if group is None:
            initial_group=list(self.datasamples.train.features)
        else:
            initial_group=group.copy()

        if sum([x not in self.datasamples.train.features for x in initial_group])>0:
            print('No',str([x for x in initial_group if x not in self.datasamples.train.features]), 'features in datasamples.train.features. Return None..')
            return None
        else:
            self.decision_trees={}
            self.decision_trees[0] = DecisionTree(self.datasamples, initial_group, [x for x in initial_group if x in self.categorical])

        if BL_dictionary is None:
            BL_dictionary={}

        if out is None:
            out=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")+'_trees_backup.xlsx'

        if correlated_exclusion_binning is not None and isinstance(correlated_exclusion_binning, WOE):
            print('Transforming train sample according to binning information from correlated_exclusion_binning..')
            to_check_correlation=correlated_exclusion_binning.transform(self.datasamples.train, keep_essential=True,
                                                                        calc_gini=False, with_excluded=True).dataframe


        times=[]
        to_fit=True
        dt_number=0
        while to_fit:
            print ('--------------------------------------- Tree #'+str(dt_number), 'processing started! ---------------------------------------')
            current_time=datetime.datetime.now()
            print('-- Starting at:                  '+str(current_time))
            times.append(current_time)
            if len(times)>1:
                time_per_feature=np.mean([times[i+1]-times[i] for i in range(len(times)-1)])
                print('-- average per tree = '+str(time_per_feature))

            self.decision_trees[dt_number].auto_fit(criterion=criterion, max_depth = max_depth, max_leaf_nodes=max_leaf_nodes,
                                                    min_samples_leaf=min_samples_leaf, min_targets_leaf=min_targets_leaf,
                                                    missing_process=missing_process, bins=bins, bins_detailed=bins_detailed, bins_type=bins_type, interval_split_old=interval_split_old,
                                                    use_features_once=use_features_once, alpha = alpha, woe_adjust = woe_adjust,
                                                    verbose=verbose, to_correct=to_correct, max_corrections=max_corrections,
                                                    BL_on=BL_on, BL_dictionary=BL_dictionary,
                                                    WOEO_on=WOEO_on, WOEO_er_threshold=WOEO_er_threshold,
                                                    WOEO_correct_threshold=WOEO_correct_threshold, WOEO_miss_is_incorrect=WOEO_miss_is_incorrect,
                                                    WOEM_on=WOEM_on, WOEM_woe_threshold=WOEM_woe_threshold,
                                                    G_on=G_on, G_gini_threshold=G_gini_threshold, G_gini_decrease_threshold=G_gini_decrease_threshold,
                                                    G_gini_increase_restrict=G_gini_increase_restrict, G_with_test=G_with_test,
                                                    selection_dict=selection_dict)

            print ('--------------------------------------- Tree #'+str(dt_number), 'processing ended! ---------------------------------------')
            gc.collect()

            tree_df=self.decision_trees[dt_number].tree
            all_features=self.decision_trees[dt_number].features
            if tree_df is None:
                print('No tree passed active checks. Breaking cycle.')
                self.exclude(dt_number)
                break
            else:
                used_features=list(tree_df.columns[:tree_df.columns.get_loc('node')])
                print('Selected tree used '+str(used_features)+' features.')
                if len(used_features)<2:
                    print('Current tree used less then 2 features. Breaking cycle.')
                    self.exclude(dt_number)
                    break
                else:
                    self.export_trees(out)
                    dt_number=max(self.decision_trees)+1

                    rest_features=[x for x in all_features if x not in used_features]

                    if correlated_exclusion_binning is not None and isinstance(correlated_exclusion_binning, WOE):
                        for f1 in used_features:
                            correlated_features={}
                            for f2 in rest_features:
                                if f1+'_WOE' in to_check_correlation and f2+'_WOE' in to_check_correlation:
                                    current_corr=to_check_correlation[[f1+'_WOE', f2+'_WOE']].corr().iloc[0,1]
                                    if abs(current_corr)>correlated_exclusion_threshold:
                                        rest_features.remove(f2)
                                        correlated_features[f2]=round(current_corr,4)
                            if len(correlated_features)>0:
                                print('Significant correlation with '+f1+' (removing features from the rest): '+str(correlated_features))

                    if len(rest_features)==0:
                        print('No more features available. Breaking cycle.')
                        break
                    elif max_trees is not None and dt_number>=max_trees:
                        print('Maximal number of trees reached. Breaking cycle.')
                        break
                    else:
                        print('Trying to build another tree using the rest features:', str(rest_features))
                        self.decision_trees[dt_number]=DecisionTree(self.datasamples, rest_features, [x for x in rest_features if x in self.categorical])



    def transform(self, data, trees=None, not_found='worst', calc_gini=True, keep_essential=False, add_fields=None, reset_features=False,
                  exclude_trees_input=False, prefix=None):#, ret_values=['group_woe']):
        '''
        Transforms a Data or DataSamples object according to built DecisionTrees. Can be used only after .fit().

        Parameters
        ------------
        data: Data or DataSamples object to transform
        trees: a list of trees' numbers to calculate
        not_found: what woe to return for unknown categorical values; 'worst' - the lowest woe, int - number of certain group
        ret_values: a list of fields from tree description DataFrame to add. Available fields:
            'node', 'parent_node', 'depth', 'amount', 'nontarget', 'target',
            'er', 'woe', 'group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'
        calc_gini: should gini values be calculated on the dataframe of data object or not
        keep_essential: should only not transformed features, target and weights be kept in output Data/DataSamples object or not
        add_fields: list of fields to be kept in addition to transformed features, target and weights in case of keep_essential=True
        reset_features: should crosses be added to input features or rewrite them
        exclude_trees_input: should WoE-transformed features based on trees input features be excluded or not
        prefix: the prefix of new fields (for each tree its' number will be added after the prefix)

        Returns
        ----------
        transformed Data or DataSamples object
        '''
        if prefix is not None:
            self.prefix=prefix
        elif self.prefix is not None:
            prefix=self.prefix
        else:
            prefix='TREE'
            self.prefix=prefix

        if add_fields is None:
            add_fields=[]

        if trees is None:
            trees=list(self.decision_trees)
        else:
            not_exist=[x for x in trees if x not in self.decision_trees]
            if len(not_exist)>0:
                print('No', str(not_exist), 'trees were found. Abort.')
                return None

        result=copy.deepcopy(data)
        if isinstance(data, Data):
            new_features=([] if reset_features else data.features) + [prefix+str(i)+'_WOE' for i in trees]

            for dt_number in trees:
                tree_df=self.decision_trees[dt_number].tree
                used_features=list(tree_df.columns[:tree_df.columns.get_loc('node')])
                if exclude_trees_input:
                    new_features=[x for x in new_features if x[:-4] not in used_features]
                #result.dataframe[['TREE'+str(dt_number)+'_'+x.upper() for x in ret_values]]=self.decision_trees[dt_number].transform(data.keep(used_features).dataframe, ret_values=ret_values)
                #result.features=result.features+['tree'+str(dt_number)+'_'+x.upper() for x in ret_values]
                if not_found=='worst':
                    not_found_woe=tree_df[tree_df['leaf']]['group_woe'].min()
                else:
                    not_found_woe=tree_df[tree_df['group']==not_found]['group_woe'].min()
                result.dataframe[prefix+str(dt_number)+'_WOE']=self.decision_trees[dt_number].transform(data.keep(used_features).dataframe, ret_values=['group_woe']).fillna(not_found_woe)

            if keep_essential:
                result=result.keep(new_features+add_fields)
            result.features=new_features
            if calc_gini:
                result.ginis=result.calc_gini().copy()
            return result
        elif isinstance(data, DataSamples):
            samples=[result.train, result.validate, result.test, result.bootstrap_base]
            for sample in samples:
                if sample is not None:
                    new_features=([] if reset_features else sample.features) + [prefix+str(i)+'_WOE' for i in trees]

                    for dt_number in trees:
                        tree_df=self.decision_trees[dt_number].tree
                        used_features=list(tree_df.columns[:tree_df.columns.get_loc('node')])
                        if exclude_trees_input:
                            new_features=[x for x in new_features if x[:-4] not in used_features]
                        #sample.dataframe[['TREE'+str(dt_number)+'_'+x.upper() for x in ret_values]]=self.decision_trees[dt_number].transform(sample.keep(used_features).dataframe, ret_values=ret_values)
                        #sample.features=sample.features+['TREE'+str(dt_number)+'_'+x.upper() for x in ret_values]
                        if not_found=='worst':
                            not_found_woe=tree_df[tree_df['leaf']]['group_woe'].min()
                        else:
                            not_found_woe=tree_df[tree_df['group']==not_found]['group_woe'].min()
                        sample.dataframe[prefix+str(dt_number)+'_WOE']=self.decision_trees[dt_number].transform(sample.keep(used_features).dataframe, ret_values=['group_woe']).fillna(not_found_woe)

                    if keep_essential:
                        if sample.weights is not None:
                            keep_technical=[sample.target, sample.weights]
                        else:
                            keep_technical=[sample.target]
                        sample.dataframe=sample.dataframe[new_features+add_fields+keep_technical].copy()
                    sample.features=new_features
                    if calc_gini:
                        sample.ginis=sample.calc_gini().copy()
            return result



    def exclude(self, to_exclude):
        '''
        Excludes a tree or a list of trees (by number) from self.decision_trees (and adds it/them to self.excluded_decision_trees)

        Parameters
        ------------
        to_exclude: a tree number or a list of trees' numbers

        '''

        print('Excluding tree #'+str(to_exclude))
        if isinstance(to_exclude, list)==False:
            to_exclude=[to_exclude]
        not_in_decision_trees=[x for x in to_exclude if x not in self.decision_trees]
        if len(not_in_decision_trees)>0:
            print('No', not_in_decision_trees, 'in self.decision_trees. Check the numbers of trees to exclude.')
            return None
        self.excluded_decision_trees.update({x:self.decision_trees[x] for x in to_exclude})
        self.decision_trees = {x:self.decision_trees[x] for x in self.decision_trees if x not in to_exclude}



    def include(self, to_include):
        '''
        Includes a tree or a list of trees (by number) from self.excluded_decision_trees (and adds it/them to self.decision_trees)

        Parameters
        ------------
        to_include: a tree number or a list of trees' numbers

        '''

        print('Including tree #'+str(to_include))
        if isinstance(to_include, list)==False:
            to_include=[to_include]
        not_in_excluded_decision_trees=[x for x in to_include if x not in self.excluded_decision_trees]
        if len(not_in_excluded_decision_trees)>0:
            print('No', not_in_excluded_decision_trees, 'in self.excluded_decision_trees. Check the numbers of trees to include.')
            return None
        self.decision_trees.update({x:self.excluded_decision_trees[x] for x in to_include})
        self.excluded_decision_trees = {x:self.excluded_decision_trees[x] for x in self.excluded_decision_trees if x not in to_include}



    def draw_trees(self, order=None, figsize=(12,8), draw_node_labels=True, draw_node_labels_list=None,
                  draw_leaves_by_group=True, shift_node_labels=(0,-0.04), margins=(0.05,0.15)):
        '''
        Visualize selected trees and print trees' parameters and stats

        Parameters
        -----------
        order: type of gini values to sort tree candidates by
        figsize: a tuple with sizes for a tree visualization (width, height)
        draw_node_labels: a flag for printing any additional information for nodes (except nodes' numbers)
        draw_node_labels_list: a list of fields from input_df to print on graph (in addition to features'
            conditions). Available fields:
                'node', 'parent_node', 'depth', 'amount', 'nontarget', 'target',
                'er', 'woe', 'group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'
        draw_leaves_by_group: a flag for leaves visualization according to groups (if any nodes were united
            as a group)
        shift_node_labels: a tuple with shift values for node information (0,0 corresponds to the centers of nodes)
        margins: a tuple with margins' sizes for the graph (if 0,0, then nodes' information and leaves figures may be cut)
        '''

        if draw_node_labels_list is None:
            draw_node_labels_list=[]

        trees_ginis=[[x, self.decision_trees[x].ginis['Train'], self.decision_trees[x].ginis['Validate'], self.decision_trees[x].ginis['Test'],
                      np.mean([self.decision_trees[x].ginis['Bootstrap'+str(g)] for g in range(len(self.datasamples.bootstrap))]),
                      np.std([self.decision_trees[x].ginis['Bootstrap'+str(g)] for g in range(len(self.datasamples.bootstrap))])] for x in self.decision_trees]
        if order is not None:
            order_types=['train', 'validate', 'test', 'mean', 'std']
            trees_ginis.sort(key=itemgetter(order_types.index(order)+1), reverse=(order.lower()!='std'))
            print('Trees in order of', order, 'gini:\n')

        for number in range(len(trees_ginis)):
            x=trees_ginis[number]
            print('\nTree #'+str(trees_ginis[number][0]))
            print('\nParameters')
            print('-----------')
            print(str(self.decision_trees[x[0]].parameters))
            print('\nGini')
            print('-----------')
            train_gini=round(x[1],3) if x[1] is not None else None
            validate_gini=round(x[2],3) if x[2] is not None else None
            test_gini=round(x[3],3) if x[3] is not None else None
            print(('Train: '+str(train_gini)).ljust(16)+('Validate: '+str(validate_gini)).ljust(19)+\
                  ('Test '+str(test_gini)).ljust(14)+('Mean: '+str(round(x[4],3))).ljust(15)+'Std: '+str(round(x[5],3)))
            self.decision_trees[x[0]].draw_tree(figsize=figsize, draw_node_labels=draw_node_labels,
                                                draw_node_labels_list=draw_node_labels_list, draw_leaves_by_group=draw_leaves_by_group,
                                                shift_node_labels=shift_node_labels, margins=margins)



    def export_trees(self, out=None):
        '''
        Export all candidates and their gini values across self.decision_trees
        in DataFrame and Excel file

        Parameters
        -----------
        out: a string with the path to export file (.xlsx or .xls extension required)
        '''
        all_features=sorted(list(set([x for y in self.decision_trees for x in self.decision_trees[y].features])))
        tree_columns=list(self.decision_trees[list(self.decision_trees)[0]].tree.columns[self.decision_trees[list(self.decision_trees)[0]].tree.columns.get_loc('node'):])
        gini_columns=list(self.decision_trees[list(self.decision_trees)[0]].ginis)
        out_trees=pd.DataFrame(columns=['tree', 'candidate', 'selected']+all_features+tree_columns)
        out_ginis=pd.DataFrame(columns=['tree', 'candidate', 'selected']+gini_columns)
        for i in self.decision_trees:
            if self.decision_trees[i].tree is not None:
                current_out_trees, current_out_ginis=self.decision_trees[i].export_trees()
                for f in all_features:
                    if f not in current_out_trees:
                        current_out_trees[f]=None
                current_out_trees['tree']=i
                out_trees=out_trees.append(current_out_trees, ignore_index=True)
                current_out_ginis['tree']=i
                out_ginis=out_ginis.append(current_out_ginis, ignore_index=True)
        out_trees=out_trees[['tree', 'candidate', 'selected']+all_features+tree_columns].dropna(how='all', axis=1)
        used_features=[x for x in all_features if x in out_trees]
        out_trees[used_features]=out_trees[used_features].astype(str)
        out_ginis=out_ginis[['tree', 'candidate', 'selected']+gini_columns].dropna(how='all', axis=1)
        if out is not None and isinstance(out, str):
            if out[-5:]=='.xlsx' or out[-4:]=='.xls':
                writer = pd.ExcelWriter(out, engine='openpyxl')
                #woes_values=woes_df[df_columns].values.reshape(-1,).tolist()
                #woes_df.style.apply(color_background,
                #                    mn=np.mean(woes_values)-2*np.std(woes_values),
                #                    mx=np.mean(woes_values)+2*np.std(woes_values),
                #                    cmap='RdYlGn', subset=df_columns,
                #                    high=out_woe_high, low=out_woe_low)
                out_trees.to_excel(writer, sheet_name='Trees', index=False)
                #er_values=er_df[df_columns].values.reshape(-1,).tolist()
                out_ginis.style.apply(color_background,
                                  mn=0,
                                  mx=out_ginis.max().max(),
                                  cmap='RdYlGn', subset=gini_columns,
                                  high=0, low=0).to_excel(writer, sheet_name='Gini by Samples', index=False)
                # Get the openpyxl objects from the dataframe writer object.
                for cn in range(1, writer.sheets['Trees'].max_column+1):
                    cl = openpyxl.utils.get_column_letter(cn)
                    writer.sheets['Trees'].column_dimensions[cl].width = 30 if writer.sheets['Trees'][cl][0].value in used_features else 12
                #for x in writer.sheets['Trees'].columns:
                #    writer.sheets['Trees'].column_dimensions[x[0].column].width = 30 if x[0].value in used_features else 12
                for cn in range(1, writer.sheets['Gini by Samples'].max_column+1):
                    cl = openpyxl.utils.get_column_letter(cn)
                    writer.sheets['Gini by Samples'].column_dimensions[cl].width = 12
                #for x in writer.sheets['Gini by Samples'].columns:
                #    writer.sheets['Gini by Samples'].column_dimensions[x[0].column].width = 12
                writer.save()
            else:
                print('Unknown format for export file. Use .xlsx. Skipping export.')

        #return out_trees, out_ginis



    def import_trees(self, data, replace=False, exclude_rest=False, alpha=None, woe_adjust=None):
        '''
        Import all trees with their candidates from data, which can be a pandas.DataFrame or a string with
        the path to Excel file,  containing trees descriptions

        Parameters
        -----------
        data: a string with the path to import file or a pandas.DataFrame with trees
        replace: should trees from import data replace existing trees (if their numbers are equals) or should these
            trees be added with changed numbers (to keep old trees)
        exclude_rest: should the trees not presented in import data be excluded or not
        '''
        if isinstance(data, str):
            if data[-5:]=='.xlsx' or data[-4:]=='.xls':
                data=pd.read_excel(data)
            else:
                print('Unknown format of import file (use .xlsx or .xls). Abort.')
                return None

        new_trees=[]
        for tree in sorted(data['tree'].unique().tolist()):
            tree_features=[f for f in data[data['tree']==tree].dropna(how='all', axis=1) if f not in \
                               ['tree', 'candidate', 'selected', 'node', 'parent_node', 'depth', 'leaf', 'amount',
                                'nontarget', 'target', 'split_feature', 'er', 'woe', 'group', 'group_target',
                                'group_nontarget', 'group_amount', 'group_woe']]
            if (tree in self.decision_trees and replace) or tree not in self.decision_trees:
                if tree in self.decision_trees:
                    print('Replacing', tree,'tree ..')
                    self.exclude(tree)
                else:
                    print('Adding', tree,'tree ..')
                new_tree=tree
            elif (tree in self.decision_trees and not replace):
                new_tree=max(self.decision_trees)+1
                print('Tree number', tree, 'already exists and replace==False. Changing number of tree to', new_tree)

            new_trees.append(new_tree)

            self.decision_trees[new_tree]=DecisionTree(self.datasamples, tree_features, [x for x in tree_features if x in self.categorical], alpha=alpha, woe_adjust=woe_adjust)
            self.decision_trees[new_tree].import_trees(data[data['tree']==tree], replace=replace, exclude_rest=exclude_rest)

        if exclude_rest:
            to_exclude=[x for x in self.decision_trees if x not in new_trees]
            if len(to_exclude)>0:
                self.exclude(to_exclude)




class DecisionTree:
    '''
    Decision tree, that can be built on interval and categorical features (without any transformations),
    with two main methods:
        auto_fit - trying different parameters sets, checking quality, stability and business logic of built trees and
            choosing the best of correct trees
        fit - building a single tree on shosen parameters without any checks

    Also available: visualization, uniting nodes, groups and pruning
    '''

    def __init__(self, datasamples=None, features=None, categorical=None, alpha=None, woe_adjust=None):
        '''
        Initialization.

        Parameters
        -----------
        datasamples: a DataSamples object (trees are built on datasamples.train sample)
        features: a list of features names, used in training
        categorical: a list of categorical features names
        '''

        if categorical is None:
            categorical=[]

        self.datasamples=datasamples
        self.features=features
        self.categorical=categorical

        self.trees_initial={}
        self.trees_correct={}
        self.trees_ginis={}
        self.trees_parameters={}

        self.tree=None
        self.ginis={}
        self.parameters={}

        self.alpha=alpha
        self.woe_adjust=woe_adjust



    def auto_fit(self, criterion=None, max_depth = None, max_leaf_nodes=None, min_samples_leaf=None, min_targets_leaf=None,
                 missing_process='best', bins=200, bins_detailed=200, bins_type='distr', interval_split_old=False,
                 use_features_once=True, alpha = 0, woe_adjust = 0.5, verbose=True,
                 WOEM_on=True, WOEM_woe_threshold=0.05,
                 BL_on=True, WOEO_on=True, G_on=True, to_correct=True, max_corrections=None,
                 BL_dictionary=None, WOEO_er_threshold=0.01, WOEO_correct_threshold=0.85, WOEO_miss_is_incorrect=True,
                 G_gini_threshold=5, G_gini_decrease_threshold=0.2, G_gini_increase_restrict=True, G_with_test=False,
                 selection_dict={'mean':0.75, 'std':0.75, 'best':'validate'}):
        '''
        Automatically build trees going through available parameters combinations, checking business logic (if BL_on), WoE order
        stability (if WOEO_on) and Gini stability (if G_on). If to_correct==True, then in case of errors during business logic and WoE
        order attempts are made to correct them by uniting nodes/groups and pruning.
        After all trees are built and checked the best tree is chosen from available correct candidates using logic,
        defined by selection_dict

        Parameters
        -----------
        criterion: a list of criterions used in building tree candidates. Available values are 'gini' and 'entropy'
        max_depth: a list of maximal depth values used in building tree candidates
        max_leaf_nodes: list of leaves number values used in building tree candidates
        min_samples_leaf: a value of minimal observations number (or a minimal part of train sample) for a leaf
            used in building tree candidates
        min_targets_leaf: a value of minimal targets number for a leaf used in building tree candidates
        missing_process: defines how missing values should be treated in case of interval features:
            'separate' - algorithm will always try to separate missing values from other values
            'worst' - algorithm will split non missing values and assign missings to the child with the greatest event rate
            'nearest' - algorithm will split non missing values and assign missings to the child with the closest event rate
            'best' - algorithm will check 'separate' and 'nearest' variants and choose the one with the least criterion value
        bins: a number of bins to initially split feature by
        bins_detailed: a number of bins to split initially found best split's vicinity by (form the split wight before
            the best one to the split right after the best one)
        bins_type: a type of binning process:
            'distr' - bins will be defined by splitting sample on equally sized parts (for now - ignoring weights)
            'scale' - bins will be defined by splitting sample on parts by equally sized scale intervals (ignoring weights)
        interval_split_old: a flag to use the deprecated splitting procedure for interval features (using sklearn's
            DecisionTreeClassifier and ignoring min_targets_leaf option)
        use_features_once: should it be impossible to split by the same feature twice along the path to any leaf or not
        alpha: regularization parameter, used in WoE calculation
        woe_adjust: adjustment parameter for group's ER: EventRate_i = (N_target_i+woe_adjust)/(N_nontarget_i+woe_adjust)
        verbose: log detalization level. If 0 or False only major steps are printed out, in 1 or True, then building trees and
            processing checks is displayed, if 2, then also graphs for checks and trees are visualized
        to_correct: should there be any attempts to correct trees for them to pass the BL and WOEO checks by uniting or
            pruning nodes (both checks) and by uniting groups (for WOEO check in case of uniting nodes wasn't successful)
        max_corrections: maximal number of corrections for each check
        selection_dict: a dictionary, describing the logic used for the best tree selection. This logic is described like
            {<type>: <value>, <type>: <value>, ..., <type>: <value>, 'best': <type>}. Possible types: 'train', 'validate',
            'test' (gini on corresponding samples), 'mean', 'std' (mean and std of gini values on bootstrap samples).
            <type>: <value> selects best trees of remaining candidates by corresponding type, amount is defined by value
            (if value is less then 1, then it is used like a part of available candidates, otherwise it is treated like
            an exact amount). Lastly 'best':<type> is used to get the best tree by corresponfing type.
            Best trees by 'train', 'validate', 'test' and 'mean' have the maximal corresponding values, best trees by 'std'
            has the minimal value of std.
            Example: {'mean':0.75, 'std':0.75, 'best':'validate'} means 'take 75% of candidates with the biggest mean gini
            in bootstrap, then 75% of remaining candidates with the least std in bootstrap and in the end take the best tree
            by gini in validate sample'.

        Business logic check options (BL):
        ---------------------------------
        BL_on: flag to turn on/off business logic check
        BL_dictionary: a DataFrame, a dictionary or a string with a path to conditions dictionary for business logic check
            (in case of a DataFrame or a string the field with features' names should be called 'feature', 'variable' or 'var')

        WoE order check options (WOEO):
        ------------------------------
        WOEO_on: flag to turn on/off WoE order check
        WOEO_er_threshold: if WoE order is not correct, then event rate difference between swaped bins is checked
        WOEO_correct_threshold: what part of checks on bootstrap should be correct for tree to pass the check
        WOEO_miss_is_incorrect: if there is no data for a bin on bootstrap sample, should it be treated as an error or not

        WoE merge options (WOEM):
        -----------------------
        WOEM_on: flag to turn on/off merging by WoE process
        WOEM_woe_threshold: if woe difference between leaves is less then this threshold, then they are to be merged (pruned or groupped)

        Gini check options (G):
        ----------------------
        G_on: flag to turn on/off Gini check
        G_gini_threshold: gini on train and validate/95% bootstrap should be greater then this (gini is multiplied by 100)
        G_gini_decrease_threshold: gini decrease from train to validate/95% bootstrap deviation from mean to mean should be greater then this
        G_gini_increase_restrict: if gini increase should also be restricted (by G_gini_decrease_threshold)
        G_with_test: should trees be also checked on test sample (and be excluded in case of failure)
        '''

        if woe_adjust is not None:
            self.woe_adjust=woe_adjust
        if alpha is not None:
            self.alpha=alpha

        if BL_dictionary is None:
            BL_dictionary={}

        datasamples=self.datasamples
        train_data=datasamples.train

        all_amount=train_data.dataframe.shape[0] if train_data.weights is None else train_data.dataframe[train_data.weights].sum()
        if min_samples_leaf is not None and min_samples_leaf<1:
            min_samples_leaf=all_amount*min_samples_leaf

        parameters_grid={}
        parameters_grid['criterion'] = criterion if criterion is not None else ['gini', 'entropy']
        parameters_grid['min_samples_leaf'] = [min_samples_leaf] if min_samples_leaf is not None \
                                                    else [int(round(all_amount*i, 0)) for i in [.02, .01, .005]]
        parameters_grid['max_depth'] = max_depth if max_depth is not None else [i for i in range(1, 3)]
        if max_leaf_nodes is not None:
            parameters_grid['max_leaf_nodes'] = max_leaf_nodes
        if min_targets_leaf is not None:
            parameters_grid['min_targets_leaf'] = [min_targets_leaf]

        parameters={}
        trees_initial={}
        trees_correct={}
        trees_ginis={}
        for pi in range(len(ParameterGrid(parameters_grid))):
            p=ParameterGrid(parameters_grid)[pi]
            parameters[pi]=p

            current_criterion=p['criterion'] if 'criterion' in p else 'gini'
            current_max_depth=p['max_depth'] if 'max_depth' in p else 20
            current_min_samples_leaf=p['min_samples_leaf'] if 'min_samples_leaf' in p else 0
            current_max_leaf_nodes=p['max_leaf_nodes'] if 'max_leaf_nodes' in p else 100
            current_min_targets_leaf=p['min_targets_leaf'] if 'min_targets_leaf' in p else 0
            print('... Building', pi, 'candidate:', str(p))
            tree_df=self.fit(current_criterion, current_max_depth, current_min_samples_leaf, current_max_leaf_nodes,
                             current_min_targets_leaf, use_features_once=use_features_once,
                             alpha=alpha, woe_adjust=woe_adjust, missing_process=missing_process, verbose=verbose,
                             to_self=False, bins=bins, bins_detailed=bins_detailed, bins_type=bins_type, interval_split_old=interval_split_old)
            trees_initial[pi]=tree_df
            print('... Building', pi, 'candidate complete!')
            if verbose>True:
                self.draw_tree(tree_df)

            print('--------------------------------------- Checking ---------------------------------------')
            if tree_df['depth'].max()<2:
                print('......',pi,'candidate has the depth of', tree_df['depth'].max(),'. Continue candidates cycle..')
                print('----------------------------------------------------------------------------------------')
                continue
            elif BL_on:
                bl_correct, tree_df=BusinessLogicChecker().work_tree(self, tree_df, BL_dictionary, max_corrections=max_corrections,
                                                              to_correct=to_correct, verbose=verbose)
            else:
                print('... Business logic check skipped')
                bl_correct=True

            if not bl_correct:
                print('... Business logic check for', pi, 'candidate failed. Continue candidates cycle..')
                print('----------------------------------------------------------------------------------------')
                continue
            elif tree_df['depth'].max()<2:
                print('...... After all checks with corrections',pi,'candidate has the depth of', tree_df['depth'].max(),'. Continue candidates cycle..')
                print('----------------------------------------------------------------------------------------')
                continue
            else:
                if BL_on:
                    print('... Business logic check passed')
                if WOEO_on:
                    woeo_correct, tree_df=WOEOrderChecker().work_tree(self, tree_df, er_threshold=WOEO_er_threshold, correct_threshold=WOEO_correct_threshold,
                                                                      miss_is_incorrect=WOEO_miss_is_incorrect, to_correct=to_correct,
                                                                      max_corrections=max_corrections, verbose=verbose)
                else:
                    print('... WoE order check skipped')
                    woeo_correct=True

                if not woeo_correct:
                    print('... WoE order check for', pi, 'candidate failed. Continue candidates cycle..')
                    print('----------------------------------------------------------------------------------------')
                    continue
                elif tree_df['depth'].max()<2:
                    print('...... After all checks with corrections',pi,'candidate has the depth of', tree_df['depth'].max(),'. Continue candidates cycle..')
                    print('----------------------------------------------------------------------------------------')
                    continue
                else:
                    if WOEO_on:
                        print('... WoE order check passed')

                    if WOEM_on:
                        tree_df=self.merge_by_woe(tree_df, woe_threshold=WOEM_woe_threshold)
                    else:
                        print('... Merge by WoE: skipped')

                    if tree_df['group'].dropna().unique().shape[0]<2:
                        print('...... After all checks with corrections',pi,'candidate has less then 2 groups. Continue candidates cycle..')
                        print('----------------------------------------------------------------------------------------')
                        continue

                    gini_correct, gini_values=GiniChecker().work_tree(self, tree_df, verbose=G_on and verbose, out=True,
                                                              gini_threshold=G_gini_threshold,
                                                              gini_decrease_threshold=G_gini_decrease_threshold,
                                                              gini_increase_restrict=G_gini_increase_restrict, with_test=G_with_test)
                    if not G_on:
                        print('... Gini check skipped')
                        gini_correct=True
                    if G_on and not gini_correct:
                        print('... Gini check for', pi, 'candidate failed. Continue candidates cycle..')
                        print('----------------------------------------------------------------------------------------')
                        continue

                    if  gini_correct:
                        if G_on:
                            print('... Gini check passed')
                        trees_correct[pi]=tree_df
                        trees_ginis[pi]=gini_values
                        print('For', pi, 'candidate all active checks passed successfully. Continue candidates cycle..')
                        if verbose>True and trees_initial[pi].shape[0]!=tree_df.shape[0]:
                            print('See changed candidate below:')
                            self.draw_tree(tree_df)
                        print('----------------------------------------------------------------------------------------')

        self.trees_initial=copy.deepcopy(trees_initial)
        self.trees_correct=copy.deepcopy(trees_correct)
        self.trees_ginis=copy.deepcopy(trees_ginis)
        self.trees_parameters=copy.deepcopy(parameters)

        if selection_dict is not None:
            self.tree_select(selection_dict)



    def merge_by_woe(self, input_df=None, woe_threshold=0.05):
        '''
        Merges groups, close by WOE

        Parameters
        -----------
        input_df: an input DataFrame, containing tree description
        woe_threshold: if woe difference between groups (neighboring groups for interval) is less then this threshold, then they are to be merged

        Returns
        -----------
        a DataFrame containing tree description with united groups
        '''

        if input_df is None:
            to_self=True
            input_df=self.tree.copy()
        else:
            to_self=False

        tree_df=input_df.copy()

        if tree_df['group'].dropna().unique().shape[0]>1:
            to_check_woe=True
        else:
            to_check_woe=False
            print('... Merge by WoE: only 1 group is present')

        while to_check_woe:
            to_check_woe=False

            groups_to_check=tree_df[['group', 'group_woe']].drop_duplicates().sort_values('group_woe')
            min_woe_dif=None

            for i in range(groups_to_check.shape[0]-1):
                if min_woe_dif is None or abs(groups_to_check.iloc[i]['group_woe']-groups_to_check.iloc[i+1]['group_woe'])<min_woe_dif:
                    min_woe_dif=abs(groups_to_check.iloc[i]['group_woe']-groups_to_check.iloc[i+1]['group_woe'])
                    group_to=groups_to_check.iloc[i]['group']
                    group_from=groups_to_check.iloc[i+1]['group']

            if min_woe_dif is not None and min_woe_dif<woe_threshold:
                print('... Merge by WoE: merging', group_to, 'and', group_from, 'groups with WoE difference =', min_woe_dif)
                to_check_woe=True

                tree_df=self.unite_groups(tree_df, [group_to, group_from])
            else:
                print('... Merge by WoE: no groups to merge were found')

        if to_self:
            self.tree=tree_df.copy()
            self.parameters.update({'manual':True})
            _, self.ginis=GiniChecker().work_tree(self, out=True)
        else:
            return tree_df


    def tree_select(self, selection={'mean':0.75, 'std':0.75, 'best':'validate'}):
        '''
        Choose the tree from available correct candidates by specified number or using logic, defined by selection dictionary

        Parameters
        -----------
        selection:
            the number of selected by user tree or
            a DataFrame, containing tree description or
            a dictionary, describing the logic used for the tree selection.
                This logic is described like {<type>: <value>, <type>: <value>, ..., <type>: <value>, 'best': <type>}.
                Possible types: 'train', 'validate', 'test' (gini on corresponding samples), 'mean', 'std' (mean and std of gini
                values on bootstrap samples).
                <type>: <value> selects best trees of remaining candidates by corresponding type, amount is defined by value
                (if value is less then 1, then it is used like a part of available candidates, otherwise it is treated like
                an exact amount).
                Lastly 'best':<type> is used to get the best tree by corresponfing type. If no 'best' part was found in the
                dictionary, then the first tree is chosen (the best by the last criterion).
                Best trees by 'train', 'validate', 'test' and 'mean' have the maximal corresponding values, best trees by 'std'
                has the minimal value of std.
                Example: {'mean':0.75, 'std':0.75, 'best':'validate'} means 'take 75% of candidates with the biggest mean gini
                in bootstrap, then 75% of remaining candidates with the least std in bootstrap and in the end take the best tree
                by gini in validate sample'.
        '''
        if self.trees_ginis!={}:
            if isinstance(selection, pd.DataFrame):
                self.tree=selection
                _, self.ginis=GiniChecker().work_tree(self, selection, out=True)
                self.parameters={'manual':True}
            else:
                if isinstance(selection, dict):
                    trees_selection=[[x, self.trees_ginis[x]['Train'], self.trees_ginis[x]['Validate'], self.trees_ginis[x]['Test'],
                                      np.mean([self.trees_ginis[x]['Bootstrap'+str(g)] for g in range(len(self.datasamples.bootstrap))]),
                                      np.std([self.trees_ginis[x]['Bootstrap'+str(g)] for g in range(len(self.datasamples.bootstrap))])] for x in sorted(self.trees_ginis)]
                    selection_types=['train', 'validate', 'test', 'mean', 'std']
                    selected_tree_number=None
                    for select in selection:
                        for it in range(len(selection_types)):
                            if select.lower()==selection_types[it]:
                                trees_selection.sort(key=itemgetter(it+1), reverse=(select.lower()!='std'))
                                amount=int(m.ceil(selection[select]*len(trees_selection))) if selection[select]<1 else selection[select]
                                trees_selection=trees_selection[:amount]
                                print('Getting the best', selection[select], 'of candidates by',select.lower(),'gini..')
                                for x in trees_selection:
                                    train_gini=round(x[1],3) if x[1] is not None else None
                                    validate_gini=round(x[2],3) if x[2] is not None else None
                                    test_gini=round(x[3],3) if x[3] is not None else None
                                    print(('\tTree #'+str(x[0])).ljust(12)+('Train: '+str(train_gini)).ljust(16)+('Validate: '+str(validate_gini)).ljust(19)+\
                                      ('Test '+str(test_gini)).ljust(14)+('Mean: '+str(round(x[4],3))).ljust(15)+'Std: '+str(round(x[5],3)))
                        if select.lower()=='best':
                            for it in range(len(selection_types)):
                                if selection[select].lower()==selection_types[it]:
                                    trees_selection.sort(key=itemgetter(it+1), reverse=(selection[select].lower()!='std'))
                                    selected_tree_number=trees_selection[0][0]
                                    print('The best remaining candidate by',selection[select].lower(),'gini - tree #'+str(selected_tree_number))
                                    break
                    if selected_tree_number is None:
                        selected_tree_number=trees_selection[0][0]
                else:
                    selected_tree_number=selection

                self.tree=self.trees_correct[selected_tree_number].copy()
                self.ginis=self.trees_ginis[selected_tree_number].copy()
                self.parameters=self.trees_parameters[selected_tree_number].copy()
        else:
            print('No candidates are present. Abort')



    def draw_candidates(self, order=None, figsize=(12,8), draw_node_labels=True, draw_node_labels_list=None,
                  draw_leaves_by_group=True, shift_node_labels=(0,-0.04), margins=(0.05,0.15)):
        '''
        Visualize correct tree candidates and print trees' parameters and stats

        Parameters
        -----------
        order: type of gini values to sort tree candidates by
        figsize: a tuple with sizes for a tree visualization (width, height)
        draw_node_labels: a flag for printing any additional information for nodes (except nodes' numbers)
        draw_node_labels_list: a list of fields from input_df to print on graph (in addition to features'
            conditions). Available fields:
                'node', 'parent_node', 'depth', 'amount', 'nontarget', 'target',
                'er', 'woe', 'group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'
        draw_leaves_by_group: a flag for leaves visualization according to groups (if any nodes were united
            as a group)
        shift_node_labels: a tuple with shift values for node information (0,0 corresponds to the centers of nodes)
        margins: a tuple with margins' sizes for the graph (if 0,0, then nodes' information and leaves figures may be cut)
        '''

        if draw_node_labels_list is None:
            draw_node_labels_list=[]

        trees_ginis=[[x, self.trees_ginis[x]['Train'], self.trees_ginis[x]['Validate'], self.trees_ginis[x]['Test'],
                      np.mean([self.trees_ginis[x]['Bootstrap'+str(g)] for g in range(len(self.datasamples.bootstrap))]),
                      np.std([self.trees_ginis[x]['Bootstrap'+str(g)] for g in range(len(self.datasamples.bootstrap))])] for x in sorted(self.trees_ginis)]
        if order is not None:
            order_types=['train', 'validate', 'test', 'mean', 'std']
            trees_ginis.sort(key=itemgetter(order_types.index(order)+1), reverse=(order.lower()!='std'))

        for number in range(len(trees_ginis)):
            x=trees_ginis[number]
            print('\nCandidate #'+str(trees_ginis[number][0]))
            print('\nParameters')
            print('-----------')
            print(str(self.trees_parameters[trees_ginis[number][0]]))
            print('\nGini')
            print('-----------')
            train_gini=round(x[1],3) if x[1] is not None else None
            validate_gini=round(x[2],3) if x[2] is not None else None
            test_gini=round(x[3],3) if x[3] is not None else None
            print(('Train: '+str(train_gini)).ljust(16)+('Validate: '+str(validate_gini)).ljust(19)+\
                  ('Test '+str(test_gini)).ljust(14)+('Mean: '+str(round(x[4],3))).ljust(15)+'Std: '+str(round(x[5],3)))
            self.draw_tree(input_df=self.trees_correct[trees_ginis[number][0]], figsize=figsize, draw_node_labels=draw_node_labels,
                           draw_node_labels_list=draw_node_labels_list, draw_leaves_by_group=draw_leaves_by_group,
                           shift_node_labels=shift_node_labels, margins=margins)



    def add_candidate(self, input_df=None):
        '''
        Add a tree (from DataFrame or from self.tree) as a candidate

        Parameters
        -----------
        input_df: an input DataFrame, containing tree description
        '''
        if input_df is None:
            input_df=self.tree.copy()
            if sum([self.ginis==self.trees_ginis[x] for x in self.trees_ginis])>0:
                print('This candidate is already present. Abort.')
            else:
                self.trees_correct[max(self.trees_correct)+1]=self.tree
                self.trees_ginis[max(self.trees_ginis)+1]=self.ginis
                self.trees_parameters[max(self.trees_parameters)+1]=self.parameters
        else:
            _, ginis=GiniChecker().work_tree(self, input_df, out=True)
            parametes={'manual':True}
            if sum([ginis==self.trees_ginis[x] for x in self.trees_ginis])>0:
                print('This candidate is already present. Abort.')
            else:
                self.trees_correct[max(self.trees_correct)+1]=input_df
                self.trees_ginis[max(self.trees_ginis)+1]=ginis
                self.trees_parameters[max(self.trees_parameters)+1]=parametes



    def fit(self, criterion=None, max_depth=None, min_samples_leaf=None, max_leaf_nodes=None, min_targets_leaf=None, to_continue=False, missing_process='best',
            use_features_once=True, alpha=None, woe_adjust=None, verbose=False, to_self=True, except_nodes=None, train_nodes=None, bins=200, bins_detailed=200,
            bins_type='distr', interval_split_old=False, manual_train_features=None):
        '''
        Build a tree by the specified parameters

        Parameters
        -----------
        criterion: a list of criterions used in building tree candidates. Available values are 'gini' and 'entropy'
        max_depth: a list of maximal depth values used in building tree candidates
        min_samples_leaf: a value of minimal observations number (or a minimal part of train sample) for a leaf
            used in building tree candidates
        max_leaf_nodes: list of leaves number values used in building tree candidates
        min_targets_leaf: a value of minimal targets number for a leaf used in building tree candidates
        to_continue: should tree builing process continue from the current selected tree from self.tree or start from
            the tree, consisting of 1 node, containing all available data
        missing_process: defines how missing values should be treated in case of interval features:
            'separate' - algorithm will always try to separate missing values from other values
            'worst' - algorithm will split non missing values and assign missings to the child with the greatest event rate
            'nearest' - algorithm will split non missing values and assign missings to the child with the closest event rate
            'best' - algorithm will check 'separate' and 'nearest' variants and choose the one with the least criterion value
        use_features_once: should it be impossible to split by the same feature twice along the path to any leaf or not
        alpha: regularization parameter, used in WoE calculation
        woe_adjust: adjustment parameter for group's ER: EventRate_i = (N_target_i+woe_adjust)/(N_nontarget_i+woe_adjust)
        verbose: log detalization level. If 0 or False only major steps are printed out, in 1 or True, then building trees and
            processing checks is displayed, if 2, then also graphs for checks and trees are visualized
        to_self: a flag to save the built tree in self.tree
        except_nodes: a list of nodes not to split/re-split (either they stay sa leaves, or current split is fixed -
            for a case of fitting tree after using DecisionTree.split method)
        train_nodes: list of nodes to train (all other nodes are added to except_nodes list)
        bins: a number of bins to initially split feature by
        bins_detailed: a number of bins to split initially found best split's vicinity by (form the split wight before
            the best one to the split right after the best one)
        bins_type: a type of binning process:
            'distr' - bins will be defined by splitting sample on equally sized parts (for now - ignoring weights)
            'scale' - bins will be defined by splitting sample on parts by equally sized scale intervals (ignoring weights)
        interval_split_old: a flag to use the deprecated splitting procedure for interval features (using sklearn's
            DecisionTreeClassifier and ignoring min_targets_leaf option)
        manual_train_features: a list of features to manually train on (if None, then all available features will be used)

        Returns
        -----------
        a DataFrame containing built tree description
        '''
        #-----------------------------------------------Subsidiary functions--------------------------------------------------
        def best_categorical_split(df, features, min_samples_leaf, min_targets_leaf, good_bad, criterion='gini', verbose=False):
            '''
            TECH

            Get the best split for a input DataFrame by a categorical feature. For each category stats are calculated,
            categories are sorted by event rate and then two best groups (with monotonis event rate) are searched for (each time
            weighted impurity is calculated and the best split with minimal impurity is chosen).

            Parameters
            -----------
            df: an input DataFrame to find the best (with the most 'pure' nodes) split by specified criterion
            features: a list of features to try splitting input df by
            min_samples_leaf: minimal observations number for a leaf
            min_targets_leaf: minimal targets number for a leaf
            good_bad: a tuple with nontarget and target amounts for entire train sample
            criterion: a criterion to calculate nodes impurity. Available values are 'gini' and 'entropy'
            verbose: a flag for printing variants

            Returns
            -----------
            weighted impurity of chosen nodes, split feature, nodes conditions and nodes target/nontarget amounts
            '''
            best_splits_by_feature={}
            if train_data.weights is not None:
                all_category_stats=df[features+[train_data.target, train_data.weights]]
                all_category_stats['---weight---']=all_category_stats[train_data.weights]
            else:
                all_category_stats=df[features+[train_data.target]]
                all_category_stats['---weight---']=1
            all_category_stats[train_data.target]=all_category_stats[train_data.target]*all_category_stats['---weight---']
            train_obs=sum(good_bad)
            all_target=all_category_stats[train_data.target].sum()
            all_nontarget=all_category_stats['---weight---'].sum()-all_target

            for f in tqdm_notebook(features, disable=(verbose==False), desc='Categorical splitting'):
                for_category_stats=all_category_stats[[f, train_data.target, '---weight---']]
                category_stats=for_category_stats.groupby(f, as_index=False).sum()
                category_stats.columns=[f, 'target', 'amount']

                if for_category_stats[f].isnull().any():
                    miss_target, miss_all=for_category_stats[[train_data.target, '---weight---']][pd.isnull(for_category_stats[f])].sum()
                    category_stats.loc[-1]=(np.nan, miss_target, miss_all)

                category_stats['nontarget']=category_stats['amount']-category_stats['target']
                category_stats['er']=category_stats['target']/category_stats['amount']
                category_stats=category_stats.sort_values('er')

                left_nontarget=0
                left_target=0
                left_group=[]
                right_group=category_stats[f].unique().tolist()
                right_nontarget=all_nontarget
                right_target=all_target

                min_impurity=None
                for _, (add_target, add_nontarget, new_value) in category_stats[['target', 'nontarget', f]][:-1].iterrows():
                #for er in category_stats['er']:
                    impurity=None
                    left_nontarget+=add_nontarget
                    left_target+=add_target
                    right_nontarget-=add_nontarget
                    right_target-=add_target
                    left_group.append(new_value)
                    if pd.isnull(new_value):
                        right_group=[x for x in right_group if pd.isnull(x)==False]
                    else:
                        right_group.remove(new_value)

                    groups=[left_group,
                            right_group]
                    amounts=[[left_nontarget, left_target], [right_nontarget, right_target]]
                    impurity=check_split(amounts, min_samples_leaf, min_targets_leaf, train_obs, criterion)

                    if impurity is not None:
                        if verbose>True:
                                print('\t\tVariants: categorical split by', f, 'on', str(groups),', impurity = ', impurity)
                        if min_impurity is None or impurity<min_impurity:
                            min_impurity=impurity
                            min_impurity_groups=copy.deepcopy(groups)
                            min_impurity_amounts=amounts


                if min_impurity is not None:
                    min_impurity_groups=[[x for x in min_impurity_groups[0] if pd.isnull(x)]+\
                                             sorted([x for x in min_impurity_groups[0] if not pd.isnull(x)]),
                                        [x for x in min_impurity_groups[1] if pd.isnull(x)]+\
                                             sorted([x for x in min_impurity_groups[1] if not pd.isnull(x)])]
                    if verbose>True:
                        print('\t\tVariants: best categorical split by', f, 'on', str(min_impurity_groups),', impurity = ', min_impurity)
                    best_splits_by_feature[f]=(min_impurity,
                                               min_impurity_groups,
                                               min_impurity_amounts)
            final_impurity=None
            final_split=None
            for f in features:
                if f in best_splits_by_feature  and (final_impurity is None or best_splits_by_feature[f][0]<final_impurity):
                    final_impurity=best_splits_by_feature[f][0]
                    final_split=(f, best_splits_by_feature[f][1], best_splits_by_feature[f][2])

            if final_impurity is not None:
                return final_impurity, final_split[0], final_split[1], final_split[2]
            else:
                return None, None, None, None


        #deprecated by Yudochev Dmitry 18.03.2019
        def best_interval_split_old(df, features, conditions, missing_process, min_samples_leaf, good_bad, criterion='gini', verbose=False):
            '''
            TECH

            Get the best split for an input DataFrame by an interval feature. There are separate attempts for features
            without missing values (using sklearn DecisionTreeClassifier) and for features with missing values (non-missing
            values are split by DecisionTreeClassifier and then split is corrected by uniting missing values or keeping
            them separate), then the best split is chosen between them.

            Parameters
            -----------
            df: an input DataFrame to find the best (with the most 'pure' nodes) split by the specified criterion
            features: a list of features to try splitting input df by
            conditions: parent node conditions to use for generating children conditions
            missing_process: defines how missing values should be treated in case of interval features:
                'separate' - algorithm will always try to separate missing values from other values
                'worst' - algorithm will split non missing values and assign missings to the child with the greatest event rate
                'nearest' - algorithm will split non missing values and assign missings to the child with the closest event rate
                'best' - algorithm will check 'separate' and 'nearest' variants and choose the one with the least criterion value
            min_samples_leaf: minimal observations number for a leaf
            good_bad: a tuple with nontarget and target amounts for entire train sample
            criterion: a criterion to calculate nodes impurity. Available values are 'gini' and 'entropy'
            verbose: a flag for printing variants

            Returns
            -----------
            weighted impurity of chosen nodes, split feature, nodes conditions and nodes target/nontarget amounts
            '''
            target=train_data.target
            features_with_missings=[]
            features_without_missings=[]
            for f in features:
                if df[f].isnull().any():
                    features_with_missings.append(f)
                else:
                    features_without_missings.append(f)

            without_missings_impurity=None
            if len(features_without_missings)>0:
                dtree=DecisionTreeClassifier()
                dtree.set_params(max_depth=1, min_samples_leaf=min_samples_leaf)
                dtree.fit(df[features_without_missings],
                          df[target],
                          sample_weight=df[train_data.weights] if train_data.weights is not None else None)
                if dtree.tree_.node_count>1:
                    without_missings_feature=features_without_missings[dtree.tree_.feature[0]]
                    shifted_threshold=df[df[without_missings_feature]>dtree.tree_.threshold[0]][without_missings_feature].min()
                    if str(np.dtype(shifted_threshold))[:3]=='int':
                        shifted_threshold=int(shifted_threshold)
                    else:
                        shifted_threshold=float(shifted_threshold)
                    if conditions[without_missings_feature] is None:
                        without_missings_groups=[(-np.inf, shifted_threshold), (shifted_threshold, np.inf)]
                    else:
                        without_missings_groups=[(conditions[without_missings_feature][0], shifted_threshold),
                                                 (shifted_threshold, conditions[without_missings_feature][1])]
                    if train_data.weights is None:
                        without_missings_amounts=dtree.tree_.value[1:, 0].astype(int).tolist()
                    else:
                        without_missings_amounts=dtree.tree_.value[1:, 0].tolist()
                    without_missings_impurity=sum([calc_impurity(without_missings_amounts[i], criterion=criterion)* \
                                                   sum(without_missings_amounts[i])/sum(good_bad) for i in range(len(without_missings_amounts))])
                    if verbose>True:
                        print('\t\tVariants: best without missings interval split by', without_missings_feature, 'on', str(without_missings_groups),', impurity = ', without_missings_impurity)

            with_missings_impurity=None
            if len(features_with_missings)>0:
                if train_data.weights is not None:
                    all_miss_stats=df[features_with_missings+[target, train_data.weights]]
                    all_miss_stats['---weight---']=all_miss_stats[train_data.weights]
                else:
                    all_miss_stats=df[features_with_missings+[target]]
                    all_miss_stats['---weight---']=1
                all_miss_stats[target]=all_miss_stats[target]*all_miss_stats['---weight---']

                for f in features_with_missings:
                    for_miss_stats=all_miss_stats[[f, target, '---weight---']]
                    for_miss_stats['miss']=pd.isnull(for_miss_stats[f])
                    miss_stats=for_miss_stats[['miss', target, '---weight---']].groupby('miss', as_index=False).sum()
                    miss_stats.columns=['miss', 'target', 'amount']
                    miss_stats['nontarget']=miss_stats['amount']-miss_stats['target']

                    impurity=None
                    if miss_stats[miss_stats['miss']==False].shape[0]>0 and \
                       miss_stats[miss_stats['miss']==False]['amount'].values[0]>2*min_samples_leaf and \
                       missing_process in ['worst', 'nearest', 'best']:
                        df_without_missings=df[pd.isnull(df[f])==False]
                        dtree=DecisionTreeClassifier()
                        dtree.set_params(max_depth=1, min_samples_leaf=min_samples_leaf)
                        dtree.fit(df_without_missings[[f]],
                                  df_without_missings[target],
                                  sample_weight=df_without_missings[train_data.weights] \
                                                  if train_data.weights is not None else None)

                        if dtree.tree_.node_count>1:
                            shifted_threshold=df_without_missings[df_without_missings[f]>dtree.tree_.threshold[0]][f].min()
                            if str(np.dtype(shifted_threshold))[:3]=='int':
                                shifted_threshold=int(shifted_threshold)
                            else:
                                shifted_threshold=float(shifted_threshold)
                            if conditions[f] is None:
                                groups=[(-np.inf, shifted_threshold), (shifted_threshold, np.inf)]
                            else:
                                groups=[(conditions[f][0][0], shifted_threshold),
                                        (shifted_threshold, conditions[f][0][1])]
                            if train_data.weights is None:
                                amounts=dtree.tree_.value[1:, 0].astype(int).tolist()
                            else:
                                amounts=dtree.tree_.value[1:, 0].tolist()
                            if (missing_process=='worst' and \
                                    amounts[0][1]/amounts[0][0]>=amounts[1][1]/amounts[1][0]) or \
                               (missing_process in ['nearest', 'best'] and \
                                    abs(amounts[0][1]/amounts[0][0]-miss_stats[miss_stats['miss']]['target'].values[0]/miss_stats[miss_stats['miss']]['nontarget'].values[0])<= \
                                    abs(amounts[1][1]/amounts[1][0]-miss_stats[miss_stats['miss']]['target'].values[0]/miss_stats[miss_stats['miss']]['nontarget'].values[0])):
                                groups=[(groups[0], np.nan), groups[1]]
                                amounts=[[amounts[0][0]+miss_stats[miss_stats['miss']]['nontarget'].values[0],
                                          amounts[0][1]+miss_stats[miss_stats['miss']]['target'].values[0]],
                                         amounts[1]]
                            else:
                                groups=[groups[0], (groups[1], np.nan)]
                                amounts=[amounts[0],
                                         [amounts[1][0]+miss_stats[miss_stats['miss']]['nontarget'].values[0],
                                          amounts[1][1]+miss_stats[miss_stats['miss']]['target'].values[0]]]
                            impurity=sum([calc_impurity(amounts[i], criterion=criterion)*sum(amounts[i])/sum(good_bad) for i in range(len(amounts))])
                            if verbose>True:
                                print('\t\tVariants: with missings split (',missing_process,') by', f, 'on', str(groups),', impurity = ', impurity)

                            if with_missings_impurity is None or impurity<with_missings_impurity:
                                with_missings_impurity=impurity
                                with_missings_feature=f
                                with_missings_groups=groups
                                with_missings_amounts=amounts

                    if impurity is None or missing_process=='best':
                        if impurity is None and missing_process in ['worst', 'nearest', 'best'] and verbose>True:
                            print("\t\tNo worst/nearest missing split was found. Trying separate..")

                        if miss_stats[miss_stats['miss']==False].shape[0]>0 and \
                           miss_stats[miss_stats['miss']==False]['amount'].values[0]>min_samples_leaf:
                            if conditions[f] is None:
                                groups=[np.nan, (-np.inf, np.inf)]
                            else:
                                groups=[np.nan, (conditions[f][0][0], conditions[f][0][1])]
                            amounts=[[miss_stats[miss_stats['miss']].nontarget.values[0],
                                          miss_stats[miss_stats['miss']].target.values[0]],
                                     [miss_stats[miss_stats['miss']==False].nontarget.values[0],
                                          miss_stats[miss_stats['miss']==False].target.values[0]]]
                            impurity=sum([calc_impurity(amounts[i], criterion=criterion)*sum(amounts[i])/sum(good_bad) for i in range(len(amounts))])
                            if verbose>True:
                                print('\t\tVariants: with missings split (separate) by', f, 'on', str(groups),', impurity = ', impurity)

                            if with_missings_impurity is None or impurity<with_missings_impurity:
                                with_missings_impurity=impurity
                                with_missings_feature=f
                                with_missings_groups=groups
                                with_missings_amounts=amounts

                #if verbose and with_missings_impurity is not None:
                #    print('Best with missings split by', with_missings_feature, 'on', str(with_missings_groups),', impurity = ', with_missings_impurity)

            if with_missings_impurity is not None and (without_missings_impurity is None or without_missings_impurity>with_missings_impurity):
                return with_missings_impurity, with_missings_feature, with_missings_groups, with_missings_amounts
            elif without_missings_impurity is not None:
                return without_missings_impurity, without_missings_feature, without_missings_groups, without_missings_amounts
            else:
                return None, None, None, None


        def best_interval_split(df, features, conditions, missing_process, min_samples_leaf, min_targets_leaf, good_bad, criterion='gini',
                                verbose=False, bins=200, bins_detailed=200, bins_type='distr'):
            '''
            TECH

            Get the best split for an input DataFrame by an interval feature.

            Parameters
            -----------
            df: an input DataFrame to find the best (with the most 'pure' nodes) split by the specified criterion
            features: a list of features to try splitting input df by
            conditions: parent node conditions to use for generating children conditions
            missing_process: defines how missing values should be treated in case of interval features:
                'separate' - algorithm will always try to separate missing values from other values
                'worst' - algorithm will split non missing values and assign missings to the child with the greatest event rate
                'nearest' - algorithm will split non missing values and assign missings to the child with the closest event rate
                'best' - algorithm will check 'separate' and 'nearest' variants and choose the one with the least criterion value
            min_samples_leaf: minimal observations number for a leaf
            min_targets_leaf: minimal targets number for a leaf
            good_bad: a tuple with nontarget and target amounts for entire train sample
            criterion: a criterion to calculate nodes impurity. Available values are 'gini' and 'entropy'
            verbose: a flag for printing variants
            bins: a number of bins to initially split feature by
            bins_detailed: a number of bins to split initially found best split's vicinity by (form the split wight before
                the best one to the split right after the best one)
            bins_type: a type of binning process:
                'distr' - bins will be defined by splitting sample on equally sized parts (for now - ignoring weights)
                'scale' - bins will be defined by splitting sample on parts by equally sized scale intervals (ignoring weights)

            Returns
            -----------
            weighted impurity of chosen nodes, split feature, nodes conditions and nodes target/nontarget amounts
            '''

            def calc_bins(for_interval_stats, feature, bins_type, bins_number):
                '''
                TECH
                Calculating a pandas DataFrame with target and nontarget amounts for each bin with a corresponding split value

                Parameters
                -----------
                for_interval_stats: a pandas DataFrame, containing initial data with a feature to split by, target flags and weights
                feature: a feature to split sample by
                bins_type: a type of binning process:
                    'distr' - bins will be defined by splitting sample on equally sized parts (for now - ignoring weights)
                    'scale' - bins will be defined by splitting sample on parts by equally sized scale intervals (ignoring weights)
                bins_number: a maximal number of bins to split on (if there is not enough unique values to split by, then data will be groupped by these values)

                Returns
                -----------
                a DataFrame with split values and target, nontarget amounts for each bin (left edge of eash bin is a specified split value)
                '''
                if for_interval_stats[feature].unique().shape[0]>bins_number:
                    if bins_type=='distr':
                        for_interval_stats['bin']=pd.qcut(for_interval_stats[feature], bins_number, duplicates='drop')
                        for_interval_stats=for_interval_stats.drop(feature, axis=1).merge(for_interval_stats[['bin', feature]].groupby('bin', as_index=False).min(),
                                                                                          on='bin', how='left').drop('bin', axis=1)
                    elif bins_type=='scale':
                        for_interval_stats[f]=pd.cut(for_interval_stats[feature], bins_number, right=False, include_lowest=True).apply(lambda x: x.left)

                interval_stats=for_interval_stats.groupby(feature, as_index=False).sum()
                interval_stats.columns=[feature, 'target', 'amount']
                interval_stats['nontarget']=interval_stats['amount']-interval_stats['target']
                interval_stats=interval_stats.sort_values(feature).reset_index(drop=True)
                interval_stats['split']=interval_stats[feature].shift(-1)
                interval_stats.drop([feature, 'amount'], axis=1, inplace=True)

                return interval_stats

            def place_missings(initial_groups, missing_process, left_target, left_nontarget, right_target, right_nontarget, miss_target, miss_nontarget):
                '''
                TECH
                Checking where should missings be placed with current missing proccess and amounts of targets/nontargets of split

                Parameters
                -----------
                amounts: a list of lists with amounts of nontargets and targets of current split (left and right parts)
                missing_process: defines how missing values should be treated in case of interval features:
                    'separate' - algorithm will always try to separate missing values from other values
                    'worst' - algorithm will split non missing values and assign missings to the child with the greatest event rate
                    'nearest' - algorithm will split non missing values and assign missings to the child with the closest event rate
                    'best' - algorithm will check 'separate' and 'nearest' variants and choose the one with the least criterion value
                left_target: a number of targets in the left part of split
                left_nontarget: a number of nontargets in the left part of split
                right_target: a number of targets in the right part of split
                right_nontarget: a number of nontargets in the right part of split
                miss_target: a number of targets with missing value
                miss_nontarget: a number of nontargets with missing value

                Returns
                -----------
                a list of tuples, describing result groups
                and a list of lists with amounts of nontargets and targets of current split (left and right parts)
                '''
                if (missing_process=='worst' and \
                        left_target/(left_target+left_nontarget)>=right_target/(right_target+right_nontarget)) or \
                   (missing_process in ['nearest', 'best'] and \
                        abs(left_target/(left_target+left_nontarget)-miss_target/(miss_target+miss_nontarget))<= \
                        abs(right_target/(right_target+right_nontarget)-miss_target/(miss_target+miss_nontarget))):
                    amounts=[[left_nontarget+miss_nontarget, left_target+miss_target],
                             [right_nontarget, right_target]]
                    groups=[(initial_groups[0], np.nan),
                             initial_groups[1]]
                else:
                    amounts=[[left_nontarget, left_target],
                             [right_nontarget+miss_nontarget, right_target+miss_target]]
                    groups=[initial_groups[0],
                            (initial_groups[1], np.nan)]

                return groups, amounts

            #-------------------------------------------------

            features_with_missings=[]
            for f in features:
                if df[f].isnull().any():
                    features_with_missings.append(f)

            best_splits_by_feature={}
            if train_data.weights is not None:
                all_interval_stats=df[features+[train_data.target, train_data.weights]]
                all_interval_stats['---weight---']=all_interval_stats[train_data.weights]
            else:
                all_interval_stats=df[features+[train_data.target]]
                all_interval_stats['---weight---']=1
            all_interval_stats[train_data.target]=all_interval_stats[train_data.target]*all_interval_stats['---weight---']
            train_obs=sum(good_bad)
            all_target=all_interval_stats[train_data.target].sum()
            all_nontarget=all_interval_stats['---weight---'].sum()-all_target

            for f in tqdm_notebook(features, disable=(verbose==False), desc='Interval splitting'):
                with_missings=(f in features_with_missings)
                min_max=conditions[f] if conditions[f] is not None else (-np.inf, np.inf)
                if pd.isnull(min_max):
                    continue
                if pd.isnull(min_max[1]):
                    min_max=min_max[0]

                interval_stats = calc_bins(all_interval_stats[[f, train_data.target, '---weight---']], f, bins_type, bins)

                if with_missings:
                    miss_target, miss_all = all_interval_stats[[train_data.target, '---weight---']][pd.isnull(all_interval_stats[f])].sum()
                    miss_nontarget = miss_all - miss_target
                else:
                    miss_target, miss_nontarget = 0, 0

                min_impurity=None
                min_impurity_groups=None
                min_dna_impurity=None
                printed_impurity=None
                left_nontarget=0
                left_target=0
                right_nontarget=all_nontarget-miss_nontarget
                right_target=all_target-miss_target

                for detailed in range(2):
                    if detailed==1:
                        if min_impurity is None:
                            continue

                        best_split=min_impurity_groups[0][0][1] if pd.isnull(min_impurity_groups[0][1]) else min_impurity_groups[0][1]

                        before_best_split=(interval_stats.split<best_split)
                        detailed_from, detailed_from_target, detailed_from_nontarget = \
                            interval_stats[before_best_split].agg({'split':'max', 'target':'sum', 'nontarget':'sum'}) if before_best_split.any() else (-np.inf, 0, 0)

                        after_best_split=(interval_stats.split>best_split)
                        detailed_to = interval_stats[after_best_split].split.min() if after_best_split.any() else np.inf

                        interval_stats = calc_bins(all_interval_stats[[f, train_data.target, '---weight---']]\
                                                   [(all_interval_stats[f]>=detailed_from)&(all_interval_stats[f]<detailed_to)], f, bins_type, bins)

                        left_nontarget=detailed_from_nontarget
                        left_target=detailed_from_target
                        right_nontarget=all_nontarget-left_nontarget-miss_nontarget
                        right_target=all_target-left_target-miss_target

                    for _, (add_target, add_nontarget, new_split) in interval_stats[:-1].iterrows():
                        impurity=None
                        left_nontarget+=add_nontarget
                        left_target+=add_target
                        right_nontarget-=add_nontarget
                        right_target-=add_target

                        groups=[(min_max[0], new_split),
                                (new_split, min_max[1])]
                        amounts=[[left_nontarget, left_target], [right_nontarget, right_target]]
                        impurity=check_split(amounts, min_samples_leaf, min_targets_leaf, train_obs, criterion)

                        if not with_missings and impurity is not None and (min_impurity is None or impurity<min_impurity):
                            min_impurity, min_impurity_groups, min_impurity_amounts = impurity, groups, amounts
                        elif with_missings and impurity is not None and (min_dna_impurity is None or impurity<min_dna_impurity):
                            min_dna_impurity, min_dna_impurity_groups, min_dna_impurity_amounts = impurity, groups, amounts

                        if with_missings and impurity is None and missing_process in ['worst', 'nearest', 'best']:

                            groups, amounts = place_missings([(min_max[0], new_split), (new_split, min_max[1])], missing_process,
                                                             left_target, left_nontarget, right_target, right_nontarget, miss_target, miss_nontarget)
                            impurity=check_split(amounts, min_samples_leaf, min_targets_leaf, train_obs, criterion)

                            if impurity is not None and (min_impurity is None or impurity<min_impurity):
                                min_impurity, min_impurity_groups, min_impurity_amounts = impurity, groups, amounts

                    #return splits_and_impurity

                    if with_missings:
                        if verbose>2 and min_impurity is not None and (min_impurity!=printed_impurity):
                            printed_impurity=min_impurity
                            print('\t\tVariants: interval split with missings (',missing_process,') by', f, 'on', str(min_impurity_groups),', impurity = ', min_impurity)
                        if missing_process in ['worst', 'nearest', 'best'] and min_dna_impurity is not None:
                            left_nontarget=min_dna_impurity_amounts[0][0]
                            left_target=min_dna_impurity_amounts[0][1]
                            right_nontarget=min_dna_impurity_amounts[1][0]
                            right_target=min_dna_impurity_amounts[1][1]

                            groups, amounts = place_missings(min_dna_impurity_groups, missing_process,
                                                             left_target, left_nontarget, right_target, right_nontarget, miss_target, miss_nontarget)
                            impurity=check_split(amounts, min_samples_leaf, min_targets_leaf, train_obs, criterion)

                            if impurity is not None and (min_impurity is None or impurity<min_impurity):
                                min_impurity, min_impurity_groups, min_impurity_amounts = impurity, groups, amounts

                            if verbose>2 and impurity is not None and (impurity!=printed_impurity):
                                printed_impurity=impurity
                                print('\t\tVariants: interval split with missings (',missing_process,') by', f, 'on', str(groups),', impurity = ', impurity)
                    else:
                        if verbose>2 and min_impurity is not None:
                            print('\t\tVariants: interval split without missings by', f, 'on', str(min_impurity_groups),', impurity = ', min_impurity)

                if with_missings:
                    if min_impurity is None or missing_process=='best':
                        if min_impurity is None and missing_process in ['worst', 'nearest', 'best'] and verbose>2:
                            print("\t\tNo worst/nearest interval split with missings was found. Trying separate..")

                        groups=[np.nan, min_max]
                        amounts=[[miss_nontarget, miss_target],
                                 [left_nontarget+right_nontarget, left_target+right_target]]
                        impurity=check_split(amounts, min_samples_leaf, min_targets_leaf, train_obs, criterion)

                        if verbose>2:
                            print('\t\tVariants: interval split with separate missings by', f, 'on', str(groups),', impurity = ', impurity)
                        if impurity is not None and (min_impurity is None or impurity<min_impurity):
                            min_impurity, min_impurity_groups, min_impurity_amounts = impurity, groups, amounts

                if min_impurity is not None:
                    if verbose>True:
                        print('\t\tVariants: best interval split by', f, 'on', str(min_impurity_groups),', impurity = ', min_impurity)
                    best_splits_by_feature[f]=(min_impurity,
                                               min_impurity_groups,
                                               min_impurity_amounts)
            final_impurity=None
            final_split=None
            for f in features:
                if f in best_splits_by_feature  and (final_impurity is None or best_splits_by_feature[f][0]<final_impurity):
                    final_impurity=best_splits_by_feature[f][0]
                    final_split=[f, best_splits_by_feature[f][1], best_splits_by_feature[f][2]]

            if final_impurity is not None:
                if verbose>True:
                    print('\t\tBest: split before rounding by', final_split[0], 'on', str([final_split[1], final_split[2]]),', impurity = ', final_impurity)
                if pd.isnull(final_split[1][0])==False:
                    best_split=final_split[1][0][0][1] if pd.isnull(final_split[1][0][1]) else final_split[1][0][1]
                    before_best_split = all_interval_stats[final_split[0]][all_interval_stats[final_split[0]]<best_split].max()
                    rounded_best_split = (best_split+before_best_split)/2
                    precision = len(str(rounded_best_split).split('.')[1])
                    previous_rounded_best_split = None
                    while rounded_best_split>before_best_split and rounded_best_split<best_split  and previous_rounded_best_split!=rounded_best_split:
                        previous_rounded_best_split=rounded_best_split
                        final_precision = precision
                        precision-=1
                        rounded_best_split = int(rounded_best_split) if precision==0 \
                                                else int((rounded_best_split)*(10**precision))/(10**precision)
                    candidate_best_split=int((best_split+before_best_split)/2) if final_precision==0 \
                                                else int(((best_split+before_best_split)/2)*(10**final_precision))/(10**final_precision)

                    #comparing to the true precision (for example -2 for 100.0) calculated like the number of symbols after '.' minus number of trailing zeros (ignoring '.')
                    if final_precision<len(str(best_split).split('.')[1])-(len(str(best_split).replace('.',''))-len(str(best_split).replace('.','').rstrip('0'))) and \
                       (all_interval_stats[final_split[0]]<best_split).sum()==(all_interval_stats[final_split[0]]<candidate_best_split).sum():
                        rounded_best_split = candidate_best_split
                    else:
                        rounded_best_split = best_split
                    if pd.isnull(final_split[1][0][1]):
                        final_split[1] = [((final_split[1][0][0][0], rounded_best_split), np.nan), (rounded_best_split, final_split[1][1][1])]
                    elif pd.isnull(final_split[1][1][1]):
                        final_split[1] = [(final_split[1][0][0], rounded_best_split), ((rounded_best_split, final_split[1][1][0][1]), np.nan)]
                    else:
                        final_split[1] = [(final_split[1][0][0], rounded_best_split), (rounded_best_split, final_split[1][1][1])]
                return final_impurity, final_split[0], final_split[1], final_split[2]
            else:
                return None, None, None, None


        def check_split(amounts, min_samples_leaf, min_targets_leaf, train_obs, criterion):
            '''
            TECH
            Checking if split is possible due to imposed restrictions and returning impurity in case of positive decision

            Parameters
            -----------
            amounts: a list of listswith amounts of nontargets and targets of current split (left and right parts)
            min_samples_leaf: minimal observations number for a leaf
            min_targets_leaf: minimal targets number for a leaf
            train_obs: amount of observations in current sample (for weighted impurity calculation)
            criterion: a criterion to calculate nodes impurity. Available values are 'gini' and 'entropy'

            Returns
            -----------
            an impurity value (float) if split is possible
            '''
            left_obs=sum(amounts[0])
            right_obs=sum(amounts[1])

            if left_obs>=min_samples_leaf and right_obs>=min_samples_leaf and \
                (min_targets_leaf is None or (amounts[0][1]>=min_targets_leaf and amounts[1][1]>=min_targets_leaf)):

                impurity=calc_impurity(amounts[0], criterion=criterion)*left_obs/train_obs + \
                         calc_impurity(amounts[1], criterion=criterion)*right_obs/train_obs

                return impurity


        def calc_impurity(x, criterion='gini'):
            '''
            TECH
            Calculate impurity value by a tuple with nontarget and target amounts and the specified criterion

            Parameters
            -----------
            x: a tuple with nontarget and target amounts
            criterion: a criterion to calculate nodes impurity. Available values are 'gini' and 'entropy'

            Returns
            -----------
            an impurity value (float)
            '''
            if criterion=='gini':
                return 1-((x[0]/(x[0]+x[1]))**2+(x[1]/(x[0]+x[1]))**2)
            elif criterion=='entropy':
                stat=-((x[0]/(x[0]+x[1]))*np.log2(x[0]/(x[0]+x[1]))+
                           (x[1]/(x[0]+x[1]))*np.log2(x[1]/(x[0]+x[1])))
                return 0 if pd.isnull(stat) else stat
            else:
                return None


        def recursive_sort(df, node=0):
            '''
            TECH
            Recursively numerates an input DataFrame nodes by depth and values order for interval split features or event rate order
            for categorical split features and then sorts input DataFrame by this number

            Parameters
            -----------
            df: an input DataFrame containing tree description
            node: current node number

            Returns
            -----------
            ennumerated input DataFrame if recurse depth>0
            sorted input DataFrame otherwise
            '''
            if node==0:
                df['num']=0

            df.loc[df['node']==node, 'num']=df['num'].max()+1
            if not df[df['node']==node]['leaf'].values[0]:
                for child in df[df['parent_node']==node]['node'].sort_values():
                    df=recursive_sort(df, child)

            if node==0:
                df=df.sort_values('num').drop('num', axis=1).reset_index(drop=True)

            return df
        #---------------------------------------------------------------------------------------------------------------------

        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        train_data=self.datasamples.train
        categorical=self.categorical
        #target=train_data.target
        train_df=train_data.dataframe
        features=self.features

        if criterion is None:
            criterion=self.parameters['criterion']
        else:
            self.parameters['criterion']=criterion
        if max_depth is None:
            max_depth=self.parameters['max_depth']
        else:
            self.parameters['max_depth']=max_depth
        if min_samples_leaf is None:
            min_samples_leaf=self.parameters['min_samples_leaf']
        else:
            self.parameters['min_samples_leaf']=min_samples_leaf
        if min_targets_leaf is None:
            min_targets_leaf=self.parameters['min_targets_leaf']
        else:
            self.parameters['min_targets_leaf']=min_targets_leaf
        if max_leaf_nodes is None:
            max_leaf_nodes=self.parameters['max_leaf_nodes']
        else:
            self.parameters['max_leaf_nodes']=max_leaf_nodes

        if train_data.weights is not None:
            good_bad=[train_df[train_df[train_data.target]==0][train_data.weights].sum(),
                      train_df[train_df[train_data.target]==1][train_data.weights].sum()]
        else:
            good_bad=[(1-train_df[train_data.target]).sum(),
                         train_df[train_data.target].sum()]

        tree_df=pd.DataFrame(columns=features+['node', 'parent_node', 'depth', 'leaf', 'amount', 'nontarget', 'target', 'impurity',
                                               'impurity_fall', 'split_feature', 'split_groups', 'split_amounts', 'er',
                                               'woe', 'group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'])
        if not to_continue:
            dict0={x:None for x in features}
            dict0.update(dict(node=0, parent_node=None, depth=0, leaf=True,
                              amount=sum(good_bad), nontarget=good_bad[0], target=good_bad[1], impurity=calc_impurity(good_bad, criterion=criterion)))
            tree_df=tree_df.append(dict0, ignore_index=True)
        else:
            if verbose:
                print('Starting candidate building process from the currently selected candidate..')
            input_df=self.tree.copy()
            for f in features:
                if f not in input_df:
                    input_df[f]=None
            tree_df=tree_df.append(input_df)
            for field in features:
                if field not in self.tree.columns:
                    tree_df[field]=None
            tree_df['impurity']=tree_df.apply(lambda row: calc_impurity([row['nontarget'], row['target']], criterion=criterion), axis=1)

        tree_df=tree_df[features+['node', 'parent_node', 'depth', 'leaf', 'amount', 'nontarget', 'target', 'impurity',
                                               'impurity_fall', 'split_feature', 'split_groups', 'split_amounts', 'er',
                                               'woe', 'group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe']]

        leaves_count=tree_df[tree_df['leaf']].shape[0]
        if except_nodes is None:
            except_nodes=[]
        if train_nodes is not None:
            except_nodes+=[x for x in tree_df['node'] if x not in train_nodes]
            except_nodes=list(set(except_nodes))

        while leaves_count<max_leaf_nodes:
            if verbose:
                print('Current number of leaves:', leaves_count)

            min_impurity_fall=None
            if tree_df[(tree_df['leaf'])&(tree_df['depth']<max_depth+1)&\
                       (tree_df['amount']>min_samples_leaf*2)&\
                       (tree_df['target']>min_targets_leaf*2)]['node'].shape[0]>0:
                old_leaves_count=tree_df[tree_df['leaf']].shape[0]-1

                #print('except_nodes=', str(except_nodes))
                #display(tree_df)
                for node in tree_df[(tree_df['leaf'])&(tree_df['depth']<max_depth+1)&\
                                    (tree_df['amount']>min_samples_leaf*2)&\
                                    (tree_df['target']>min_targets_leaf*2)&\
                                    (tree_df['node'].isin(except_nodes)==False)&\
                                    (pd.isnull(tree_df['impurity_fall']))]['node']:
                    if verbose:
                        print('Checking node', node, '(impurity =',tree_df[tree_df['node']==node]['impurity'].values[0],', '+str(dict(tree_df[tree_df['node']==node][features].dropna(axis=1).iloc[0]))+')..')

                    current_df=self.df_filter(train_df, tree_df[tree_df['node']==node][features].iloc[0])
                    parent_node=tree_df.loc[tree_df['node']==node, 'parent_node'].values[0]
                    parent_amount=tree_df[tree_df['node']==node]['amount'].values[0]
                    parent_impurity=tree_df[tree_df['node']==node]['impurity'].values[0]
                    node_depth=tree_df.loc[tree_df['node']==node, 'depth'].values[0]

                    if manual_train_features is not None:
                        current_features=manual_train_features.copy()
                    elif use_features_once or node_depth==max_depth:
                        if node_depth>0:
                            current_features=[tree_df.loc[tree_df['node']==parent_node, 'split_feature'].values[0]]
                            if node_depth==max_depth and verbose:
                                print("\tCurrent node has maximum depth, checking only it's parent's split feature", current_features[0])
                        else:
                            current_features=[]
                        for f in features:
                            if node_depth!=max_depth and tree_df.loc[tree_df['node']==node, f].values[0] is None:
                                current_features.append(f)
                    else:
                        current_features=features.copy()

                    if tree_df.loc[tree_df['node']==node, 'parent_node'].values[0] in except_nodes and \
                        tree_df.loc[tree_df['node']==parent_node, 'split_feature'].values[0] in current_features:
                            if verbose:
                                print("\tThe current node's parent node is in except_nodes list. No split by", tree_df.loc[tree_df['node']==parent_node, 'split_feature'].values[0], 'available.')
                            current_features.remove(tree_df.loc[tree_df['node']==parent_node, 'split_feature'].values[0])

                    interval_impurity=None
                    if len([x for x in current_features if x not in categorical])>0:
                        interval_impurity, interval_feature, interval_groups, interval_amounts = \
                            best_interval_split_old(current_df,
                                                    [x for x in current_features if x not in categorical],
                                                    tree_df[tree_df['node']==node][features].iloc[0],
                                                    missing_process=missing_process, min_samples_leaf=min_samples_leaf,
                                                    good_bad=good_bad, criterion=criterion, verbose=verbose) if interval_split_old else \
                            best_interval_split(current_df,
                                                [x for x in current_features if x not in categorical],
                                                tree_df[tree_df['node']==node][features].iloc[0],
                                                missing_process=missing_process, min_samples_leaf=min_samples_leaf, min_targets_leaf=min_targets_leaf,
                                                good_bad=good_bad, criterion=criterion, verbose=verbose, bins=bins, bins_detailed=bins_detailed, bins_type=bins_type)

                        if verbose:
                            if interval_impurity is not None:
                                print('\tBest interval split by', interval_feature, 'on', str(interval_groups),', impurity = ', interval_impurity)
                            else:
                                print('\tNo interval splits were found..')
                    categorical_impurity=None
                    if len([x for x in current_features if x in categorical])>0:
                        categorical_impurity, categorical_feature, categorical_groups, categorical_amounts = \
                            best_categorical_split(current_df,
                                                  [x for x in current_features if x in categorical],
                                                  min_samples_leaf=min_samples_leaf, min_targets_leaf=min_targets_leaf, good_bad=good_bad,
                                                  criterion=criterion, verbose=verbose)
                        if verbose:
                            if categorical_impurity is not None:
                                print('\tBest categorical split by', categorical_feature, 'on', str(categorical_groups),', impurity = ', categorical_impurity)
                            else:
                                print('\tNo categorical splits were found..')

                    if categorical_impurity is not None and (interval_impurity is None or interval_impurity>categorical_impurity):
                        split_impurity=categorical_impurity
                        split_feature=categorical_feature
                        split_groups=categorical_groups
                        split_amounts=categorical_amounts
                    elif interval_impurity is not None:
                        split_impurity=interval_impurity
                        split_feature=interval_feature
                        split_groups=interval_groups
                        split_amounts=interval_amounts
                    else:
                        if verbose:
                            print('\tNo acceptable variants were found for node', node)
                        tree_df.loc[tree_df['node']==node, 'impurity_fall']=999
                        continue

                    impurity_fall=split_impurity-parent_impurity*parent_amount/sum(good_bad)

                    tree_df.loc[tree_df['node']==node, 'impurity_fall']=impurity_fall
                    tree_df.loc[tree_df['node']==node, 'split_feature']=split_feature
                    tree_df.loc[tree_df['node']==node, 'split_groups']=pd.Series([split_groups]*tree_df.shape[0])
                    tree_df.loc[tree_df['node']==node, 'split_amounts']=pd.Series([split_amounts]*tree_df.shape[0])

                    if verbose:
                        print('\tBest split for', node, 'node by', split_feature, 'will acomplish', impurity_fall,' impurity fall.')

                for node in tree_df[(tree_df['leaf'])&(tree_df['depth']<max_depth+1)&\
                                    (tree_df['amount']>min_samples_leaf*2)&\
                                    (tree_df['target']>min_targets_leaf*2)&\
                                    (tree_df['node'].isin(except_nodes)==False)&\
                                    (tree_df['impurity_fall']<0)]['node']:

                    if min_impurity_fall is None or tree_df.loc[tree_df['node']==node, 'impurity_fall'].values[0]<min_impurity_fall:
                        min_impurity_fall=tree_df.loc[tree_df['node']==node, 'impurity_fall'].values[0]
                        parent_node=tree_df.loc[tree_df['node']==node, 'parent_node'].values[0]
                        split_node=node
                        split_depth=tree_df.loc[tree_df['node']==node, 'depth'].values[0]
                        split_conditions=dict(tree_df[tree_df['node']==node][features].iloc[0])
                        groups=tree_df.loc[tree_df['node']==node, 'split_groups'].values[0]
                        split_feature=tree_df.loc[tree_df['node']==node, 'split_feature'].values[0]
                        amounts=tree_df.loc[tree_df['node']==node, 'split_amounts'].values[0]
            else:
                break

            if min_impurity_fall is not None:
                if verbose:
                    print('Final split at', split_node, 'by', split_feature, 'with groups: '+str(groups))
                leaves_count=old_leaves_count+len(groups)

                tree_df.loc[tree_df['node']==split_node, 'leaf']=False
                nodes_numbers=[tree_df['node'].max()+1, tree_df['node'].max()+2]

                if split_depth>0:
                    parent_split_feature=tree_df.loc[tree_df['node']==parent_node, 'split_feature'].values[0]

                    if split_feature==parent_split_feature:
                        nodes_numbers=[split_node, split_node+1]
                        if verbose:
                            print('\tSplitting by the same feature as the parent node. Replacing current node with its children:', split_node, '=>', str(nodes_numbers))
                        split_node=parent_node
                        split_depth=split_depth-1
                        tree_df=tree_df.drop(tree_df.loc[tree_df['node']==nodes_numbers[0]].index, axis=0)
                        tree_df.loc[tree_df['node']>nodes_numbers[0], 'node']=tree_df.loc[tree_df['node']>nodes_numbers[0], 'node']+1
                        tree_df.loc[tree_df['parent_node']>nodes_numbers[0], 'parent_node']=tree_df.loc[tree_df['parent_node']>nodes_numbers[0], 'parent_node']+1

                for i in range(len(groups)):
                    child_row=split_conditions.copy()
                    child_row[split_feature]=groups[i]
                    child_row.update({'node': nodes_numbers[i], 'leaf':True, 'parent_node':split_node, 'depth':split_depth+1,
                                  'amount':sum(amounts[i]), 'nontarget':amounts[i][0], 'target':amounts[i][1],
                                  'impurity':calc_impurity(amounts[i], criterion=criterion)})
                    tree_df=tree_df.append(child_row, ignore_index=True)
                #display(tree_df)
            else:
                if verbose:
                    print('No possible splits were found. Break.')
                break
        else:
            if verbose:
                print('Maximum number of leaves reached (',max_leaf_nodes,'). Break.')

        tree_df=recursive_sort(tree_df)
        tree_df['er']=tree_df['target']/tree_df['amount']
        tree_df['woe']=np.log(((good_bad[1]/good_bad[0])*(alpha + tree_df['amount'])/ \
                              (tree_df['amount']*((tree_df['target'] + woe_adjust)/(tree_df['nontarget'] + woe_adjust)) + alpha)).astype(float))

        tree_df[['target', 'nontarget']]=tree_df[['target', 'nontarget']].astype(tree_df['amount'].dtype)
        tree_df.loc[tree_df[tree_df['leaf']].sort_values('woe').index, 'group']=[i for i in range(tree_df[tree_df['leaf']].shape[0])]
        tree_df.loc[tree_df['leaf'], 'group_target']=tree_df.loc[tree_df['leaf'], 'target']
        tree_df.loc[tree_df['leaf'], 'group_nontarget']=tree_df.loc[tree_df['leaf'], 'nontarget']
        tree_df.loc[tree_df['leaf'], 'group_amount']=tree_df.loc[tree_df['leaf'], 'amount']
        tree_df.loc[tree_df['leaf'], 'group_woe']=tree_df.loc[tree_df['leaf'], 'woe']
        tree_df.loc[tree_df['leaf'], 'split_feature']=np.nan

        tree_df=tree_df.drop(['impurity', 'impurity_fall', 'split_groups', 'split_amounts'], axis=1).dropna(axis=1, how='all')
        if to_self:
            self.tree=tree_df

        return tree_df



    def df_filter(self, df, conditions):
        '''
        TECH
        Get sub-DataFrame for input DataFrame by input conditions (from tree description DataFrame)

        Parameters
        -----------
        df: an input DataFrame to get sub-DataFrame from
        conditions: a dict-like object with feature names as keys and conditions (intervals with/without missing values
            for interval features or lists for categorical features) as values

        Returns
        -----------
        a filtered sub-DataFrame
        '''
        df_filter=pd.Series([True]*df.shape[0], index=df.index)
        for f in conditions.index:
            if conditions[f] is not None:
                if isinstance(conditions[f], list):
                    if sum([pd.isnull(x) for x in conditions[f]])>0:
                        df_filter=df_filter & ((df[f].isin(conditions[f])|(pd.isnull(df[f]))))
                    else:
                        df_filter=df_filter & (df[f].isin(conditions[f]))
                elif not isinstance(conditions[f], list) and not isinstance(conditions[f], tuple) and \
                   pd.isnull(conditions[f]):
                    df_filter=df_filter & (pd.isnull(df[f]))
                elif isinstance(conditions[f], tuple) and pd.isnull(conditions[f][1]):
                    df_filter=df_filter & (((df[f]>=conditions[f][0][0])&(df[f]<conditions[f][0][1]))|pd.isnull(df[f]))
                elif isinstance(conditions[f], tuple):
                    df_filter=df_filter & ((df[f]>=conditions[f][0])&(df[f]<conditions[f][1]))
        '''
        new_df=df.copy()
        for f in conditions.index:
            if conditions[f] is not None:
                if isinstance(conditions[f], list):
                    if sum([pd.isnull(x) for x in conditions[f]])>0:
                        new_df=new_df[(new_df[f].isin(conditions[f])|(pd.isnull(new_df[f])))]
                    else:
                        new_df=new_df[new_df[f].isin(conditions[f])]
                elif not isinstance(conditions[f], list) and not isinstance(conditions[f], tuple) and \
                   pd.isnull(conditions[f]):
                    new_df=new_df[pd.isnull(new_df[f])]
                elif isinstance(conditions[f], tuple) and pd.isnull(conditions[f][1]):
                    new_df=new_df[((new_df[f]>=conditions[f][0][0])&(new_df[f]<conditions[f][0][1]))|pd.isnull(new_df[f])]
                elif isinstance(conditions[f], tuple):
                    new_df=new_df[(new_df[f]>=conditions[f][0])&(new_df[f]<conditions[f][1])]
        '''
        return df[df_filter]



    def split(self, node, split_feature, split_groups, to_fit=False, criterion=None, max_depth=None, min_samples_leaf=None, min_targets_leaf=None, max_leaf_nodes=None,
              missing_process='best', use_features_once=True, except_nodes=None, train_nodes=None, alpha=None, woe_adjust=None, verbose=False):
        '''
        Change currently selected tree by spliting selected node by the specified split feature and groups. Then if to_fit==True fitting of the
        selected tree continues according to its parameters (from self.parameters)

        Parameters
        -----------
        node: a number of node to re-split
        split_feature: a feature to re-split the specified node by
        split_groups: a list of groups to split the specified node by the specidied feature on. Lists depend of split feature type:
            categorical: list consists of disjoint lists of the specified feature's values
            interval: list consists of np.nan, (from, to) tuples of ((from, to), np.nan) tuples
        to_fit: should the changed tree be trained or not (user-specified split will be kept either way)
        criterion: a list of criterions used in building tree candidates. Available values are 'gini' and 'entropy'
        max_depth: a list of maximal depth values used in building tree candidates
        min_samples_leaf: a value of minimal observations number (or a minimal part of train sample) for a leaf
            used in building tree candidates
        min_targets_leaf: a value of minimal targets number for a leaf used in building tree candidates
        max_leaf_nodes: list of leaves number values used in building tree candidates
        missing_process: defines how missing values should be treated in case of interval features:
            'separate' - algorithm will always try to separate missing values from other values
            'worst' - algorithm will split non missing values and assign missings to the child with the greatest event rate
            'nearest' - algorithm will split non missing values and assign missings to the child with the closest event rate
            'best' - algorithm will check 'separate' and 'nearest' variants and choose the one with the least criterion value
        use_features_once: should it be impossible to split by the same feature twice along the path to any leaf or not
        except_nodes: list of nodes whose splitting can't be changed (leaves stay leaves, splitted node keep their split intact)
        train_nodes: list of nodes to train according to the specified parameters
        alpha: regularization parameter, used in WoE calculation
        woe_adjust: adjustment parameter for group's ER: EventRate_i = (N_target_i+woe_adjust)/(N_nontarget_i+woe_adjust)
        verbose: flag of printing detailed information

        Returns
        -----------
        a DataFrame describing the changed tree
        '''

        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        tree_df=self.tree.copy()
        if node not in tree_df['node']:
            if verbose:
                print('No such node in the selected candidate. Return input candidate')
            return tree_df
        elif split_feature not in self.features:
            if verbose:
                print('No such feature in the features available for cross. Return input candidate')
            return tree_df
        elif (sum([isinstance(g, tuple) or (isinstance(g, float) and pd.isnull(g)) for g in split_groups])>0 and split_feature in self.categorical) or \
           (sum([isinstance(g, list) for g in split_groups])>0 and split_feature not in self.categorical):
            if verbose:
                print('Specified groups are not consistent with specified split feature type.')
                print('\tFor categorical features groups should a list of lists')
                print('\tFor interval features groups should be a list of: tuples (from, to), tuples ((from, to), np.nan) or np.nan')
                print('Return input candidate')
            return tree_df
        elif split_feature not in self.categorical and sum([(isinstance(g, tuple) and pd.isnull(g[1])) or (isinstance(g, float) and pd.isnull(g)) for g in split_groups])>1:
            if verbose:
                print('More then 1 group contains np.nan (corresponding to missing values). Return input candidate')
            return tree_df
        else:

            features=self.features
            used_features=list(tree_df.columns[:tree_df.columns.get_loc('node')])
            other_fields=list(tree_df.columns[tree_df.columns.get_loc('node'):])
            train_data=self.datasamples.train.keep(features)
            train_df=train_data.dataframe
            if split_feature not in tree_df:
                tree_df[split_feature]=None
                tree_df=tree_df[used_features+[split_feature]+other_fields]
                used_features=used_features+[split_feature]

            current_df=self.df_filter(train_df, tree_df[tree_df['node']==node][used_features].iloc[0])

            parent_node_condition=tree_df.loc[tree_df['node']==node, split_feature].values[0]
            if split_feature not in self.categorical:
                parent_min=np.nan if current_df[split_feature].dropna().shape[0]==0 else -np.inf if parent_node_condition is None \
                                else parent_node_condition[0][0] if pd.isnull(parent_node_condition[1]) else parent_node_condition[0]
                parent_max=np.nan if current_df[split_feature].dropna().shape[0]==0 else np.inf if parent_node_condition is None \
                                else parent_node_condition[0][1] if pd.isnull(parent_node_condition[1]) else parent_node_condition[1]
                parent_has_missings=current_df[split_feature].isnull().any() if parent_node_condition is None else ((isinstance(parent_node_condition, tuple) and pd.isnull(parent_node_condition[1])) or pd.isnull(parent_node_condition))
                new_min=min([x for x in [g[0][0] if isinstance(g, tuple) and pd.isnull(g[1]) else g[0] if isinstance(g, tuple) else np.nan for g in split_groups] if pd.isnull(x)==False])
                new_max=max([x for x in [g[0][1] if isinstance(g, tuple) and pd.isnull(g[1]) else g[1] if isinstance(g, tuple) else np.nan for g in split_groups] if pd.isnull(x)==False])
                new_intervals={i:[split_groups[i][0][0], split_groups[i][0][1]] if pd.isnull(split_groups[i][1]) else \
                                 [split_groups[i][0], split_groups[i][1]] for i in range(len(split_groups)) if isinstance(split_groups[i], tuple)}
                new_intervals={i:new_intervals[i] for i in sorted(new_intervals, key=new_intervals.get)}
                for i in range(len(new_intervals)-1):
                    if new_intervals[list(new_intervals)[i]][1]!=new_intervals[list(new_intervals)[i+1]][0]:
                        if verbose:
                            print("Adjacent intervals don't have the common edge:",str(split_groups[list(new_intervals)[i]]),"and", str(split_groups[list(new_intervals)[i+1]]), ". Return input tree")
                        return tree_df
                split_groups=([np.nan] if np.nan in split_groups else [])+[split_groups[i] for i in new_intervals]
                has_missings=sum([(isinstance(g, tuple) and pd.isnull(g[1])) or (isinstance(g, float) and pd.isnull(g)) for g in split_groups])>0
                #print('parent_min=', parent_min,'parent_max=', parent_max,'parent_has_missings=', parent_has_missings,
                #      'new_min=', new_min,'new_max=', new_max,'has_missings=', has_missings)
                if pd.isnull(parent_min) and pd.isnull(parent_max):
                    if verbose:
                        print('Node to split has only missing values for split feature. Return input candidate')
                    return tree_df
                elif parent_min!=new_min or parent_max!=new_max:
                    if verbose:
                        print('Minimal and maximal interval edges not consistent with node to split. Return input candidate')
                    return tree_df
                elif (parent_has_missings and not has_missings) or (not parent_has_missings and has_missings):
                    if verbose:
                        print('Missings in the node to split and not in split groups or otherwise. Return input candidate')
                    return tree_df
            else:
                if parent_node_condition is None:
                    parent_node_condition=current_df[split_feature].unique().tolist()#+([np.nan] if current_df[split_feature].isnull().any() else [])
                all_values=[]
                for g in split_groups:
                    all_values+=g
                if len(all_values)!=len(set(all_values)):
                    if verbose:
                        print('Duplicates in split groups. Return input candidate')
                    return tree_df
                elif set(all_values)!=set(parent_node_condition):
                    if verbose:
                        print('Not all values or extra values are mentioned in split groups in comparison to split node values ('+str(parent_node_condition)+'). Return input candidate')
                    return tree_df

            if tree_df.loc[tree_df['node']==node, 'leaf'].values[0]==False:
                tree_df=self.prune(tree_df, node)

            tree_df.loc[tree_df['node']==node, 'leaf']=False
            tree_df.loc[tree_df['node']==node, 'split_feature']=split_feature

            for g in split_groups:
                new_node=node+1
                while new_node in tree_df['node'].values:
                    new_node+=1
                new_row=tree_df[tree_df['node']==node].iloc[0]
                new_row[split_feature]=g
                new_row['node']=new_node
                new_row['parent_node']=node
                new_row['depth']=tree_df.loc[tree_df['node']==node, 'depth'].values[0]+1
                new_row['leaf']=True
                new_row['split_feature']=np.nan

                child_df=self.df_filter(current_df, new_row[used_features])
                if train_data.weights is not None:
                    amount=child_df[train_data.weights].sum()
                    target=(child_df[train_data.weights]*child_df[train_data.target]).sum()
                else:
                    amount=child_df.shape[0]
                    target=child_df[train_data.target].sum()

                new_row['amount']=amount
                new_row['target']=target
                new_row['nontarget']=amount-target
                new_row['er']=target/amount

                good_bad=[tree_df.loc[tree_df['node']==0, 'nontarget'].values[0],
                          tree_df.loc[tree_df['node']==0, 'target'].values[0]]
                new_row['woe']=np.log(((good_bad[1]/good_bad[0])*(alpha + new_row['amount'])/ \
                                           (new_row['amount']*((new_row['target'] + woe_adjust)/(new_row['nontarget'] + woe_adjust)) + alpha)).astype(float))
                tree_df=tree_df.append(new_row, ignore_index=True)

            #display(tree_df)
            #print(tree_df.loc[tree_df['parent_node']==node, 'amount'].sum(), tree_df.loc[tree_df['node']==node, 'amount'].values[0])

            if tree_df.loc[tree_df['parent_node']==node, 'amount'].sum()!=tree_df.loc[tree_df['node']==node, 'amount'].values[0]:
                if verbose:
                    print('Not all observations from the parent node are covered by split groups. Return input candidate')
                return self.tree

            if to_fit:
                self.tree=tree_df.copy()
                if except_nodes is None:
                    except_nodes=[]
                if node not in except_nodes:
                    except_nodes.append(node)
                self.fit(criterion=criterion, max_depth=max_depth, min_samples_leaf=min_samples_leaf, min_targets_leaf=min_targets_leaf,
                         max_leaf_nodes=max_leaf_nodes, missing_process=missing_process,
                         use_features_once=use_features_once, alpha=alpha, woe_adjust=woe_adjust, verbose=verbose, to_continue=True, to_self=True,
                         except_nodes=except_nodes, train_nodes=train_nodes)
                if criterion is not None:
                    self.parameters.update({'criterion':criterion})
                if max_depth is not None:
                    self.parameters.update({'max_depth':max_depth})
                if min_samples_leaf is not None:
                    self.parameters.update({'min_samples_leaf':min_samples_leaf})
                if min_targets_leaf is not None:
                    self.parameters.update({'min_targets_leaf':min_targets_leaf})
                if max_leaf_nodes is not None:
                    self.parameters.update({'max_leaf_nodes':max_leaf_nodes})

            else:
                tree_df=tree_df.drop(['group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'], axis=1)
                tree_df.loc[tree_df[tree_df['leaf']].sort_values('woe').index, 'group']=[i for i in range(tree_df[tree_df['leaf']].shape[0])]
                tree_df.loc[tree_df['leaf'], 'group_target']=tree_df.loc[tree_df['leaf'], 'target']
                tree_df.loc[tree_df['leaf'], 'group_nontarget']=tree_df.loc[tree_df['leaf'], 'nontarget']
                tree_df.loc[tree_df['leaf'], 'group_amount']=tree_df.loc[tree_df['leaf'], 'amount']
                tree_df.loc[tree_df['leaf'], 'group_woe']=tree_df.loc[tree_df['leaf'], 'woe']
                self.tree=tree_df.copy()

            self.parameters.update({'manual':True})
            _, self.ginis=GiniChecker().work_tree(self, out=True)

            return tree_df



    def transform(self, df, input_df=None, ret_values=None):
        '''
        Build a tree by the specified parameters

        Parameters
        -----------
        df: an input DataFrame, containing all used features from input_df (tree DataFrame)
        input_df: an input DataFrame, containing tree description
        ret_values: a list of fields from input_df to return. Available fields:
            'node', 'parent_node', 'depth', 'amount', 'nontarget', 'target',
            'er', 'woe', 'group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'

        Returns
        -----------
        a DataFrame with values from ret_values corresponding to appropriate nodes for rows from df
        '''
        if ret_values is None:
            ret_values=['woe']

        if input_df is None:
            tree_df=self.tree.copy()
        else:
            tree_df=input_df.copy()

        #features=[x for x in self.features if x in tree_df]
        features=list(tree_df.columns[:tree_df.columns.get_loc('node')])
        #tree_df.columns[:tree_df.columns.get_loc('node')]
        result=pd.DataFrame(columns=ret_values, index=df.index).reset_index(drop=True)
        real_index=df.index
        temp_index=result.index
        #result[ret_values]=np.nan
        for i in tree_df[tree_df['leaf']].index:
            df_filter=pd.Series([True]*df.shape[0], index=df.index)
            for f in features:
                condition=tree_df.loc[i][f]
                if condition is not None:
                    if isinstance(condition, list):
                        if sum([pd.isnull(x) for x in condition])>0:
                            df_filter=df_filter & (df[f].isin(condition)|(pd.isnull(df[f])))
                        else:
                            df_filter=df_filter & (df[f].isin(condition))
                    elif not isinstance(condition, list) and not isinstance(condition, tuple) and pd.isnull(condition):
                        df_filter=df_filter & (pd.isnull(df[f]))
                    elif isinstance(condition, tuple) and pd.isnull(condition[1]):
                        df_filter=df_filter & (((df[f]>=condition[0][0]) & (df[f]<condition[0][1]))|(pd.isnull(df[f])))
                    elif isinstance(condition, tuple):
                        df_filter=df_filter & (df[f]>=condition[0]) & (df[f]<condition[1])
            df_filter.index=temp_index
            result.loc[df_filter, ret_values]=pd.DataFrame([tree_df.loc[i][ret_values].tolist()]*df.shape[0],
                                                           columns=ret_values, index=temp_index)
        result.index=real_index
        return result



    def draw_tree(self, input_df=None, figsize=(12,9), draw_node_labels=True, draw_node_labels_list=None,
                  draw_leaves_by_group=True, shift_node_labels=(0,-0.04), margins=(0.05,0.15)):
        '''
        Draw a tree with nodes' numbers, edges with conditions and specified information from
        input tree description DataFrame. Size of nodes depends on amount of obseervations,
        color depends on WoE values.

        Parameters
        -----------
        input_df: an input DataFrame, containing tree description
        figsize: a tuple with sizes for a tree visualization (width, height)
        draw_node_labels: a flag for printing any additional information for nodes (except nodes' numbers)
        draw_node_labels_list: a list of fields from input_df to print on graph (in addition to features'
            conditions). Available fields:
                'node', 'parent_node', 'depth', 'amount', 'nontarget', 'target',
                'er', 'woe', 'group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'
        draw_leaves_by_group: a flag for leaves visualization according to groups (if any nodes were united
            as a group)
        shift_node_labels: a tuple with shift values for node information (0,0 corresponds to the centers of nodes)
        margins: a tuple with margins' sizes for the graph (if 0,0, then nodes' information and leaves figures may be cut)
        '''
        #-----------------------------------------------Subsidiary functions--------------------------------------------------
        def hierarchy_pos(G, root, levels=None, width=1., height=1.):
            '''
            TECH
            Generate positional information appropriate for trees visualization
            If there is a cycle that is reachable from root, then this will see infinite recursion.

            Parameters
            -----------
            G: the graph
            root: the root node
            levels: a dictionary
                        key: level number (starting from 0)
                        value: number of nodes in this level
            width: horizontal space allocated for drawing
            height: vertical space allocated for drawing

            Returns
            -----------
            a dictionary containing positional information for graph nodes
            '''
            TOTAL = "total"
            CURRENT = "current"
            def make_levels(levels, node=root, currentLevel=0, parent=None):
                """Compute the number of nodes for each level
                """
                if not currentLevel in levels:
                    levels[currentLevel] = {TOTAL : 0, CURRENT : 0}
                levels[currentLevel][TOTAL] += 1
                neighbors = G.neighbors(node)
                for neighbor in neighbors:
                    if not neighbor == parent:
                        levels =  make_levels(levels, neighbor, currentLevel + 1, node)
                return levels

            def make_pos(pos, node=root, currentLevel=0, parent=None, vert_loc=0):
                dx = 1/levels[currentLevel][TOTAL]
                left = dx/2
                pos[node] = ((left + dx*levels[currentLevel][CURRENT])*width, vert_loc)
                levels[currentLevel][CURRENT] += 1
                neighbors = G.neighbors(node)
                for neighbor in neighbors:
                    if not neighbor == parent:
                        pos = make_pos(pos, neighbor, currentLevel + 1, node, vert_loc-vert_gap)
                return pos
            if levels is None:
                levels = make_levels({})
            else:
                levels = {l:{TOTAL: levels[l], CURRENT:0} for l in levels}
            vert_gap = height / (max([l for l in levels])+1)
            return make_pos({})
        #---------------------------------------------------------------------------------------------------------------------

        if draw_node_labels_list is None:
            draw_node_labels_list=[]

        if input_df is None:
            to_draw=self.tree.copy()
        else:
            to_draw=input_df.copy()

        G=nx.DiGraph()
        features=[x for x in self.features if x in to_draw]
        node_labels={}
        for i in to_draw.index:
            G.add_node(to_draw.loc[i].node)
            node_labels[to_draw.loc[i].node]='group '+str(int(to_draw.loc[i]['group'])) if pd.isnull(to_draw.loc[i]['group'])==False else ''
            for f in features:
                node_labels[to_draw.loc[i].node]+='\n'+f+':'+(str(to_draw.loc[i][f]) if len(str(to_draw.loc[i][f]))<=20 else str(to_draw.loc[i][f])[:20]+'..')
            for v in draw_node_labels_list:
                node_labels[to_draw.loc[i].node]+='\n'+v+' = '+str(round(to_draw.loc[i][v],3) if pd.isnull(to_draw.loc[i][v])==False else to_draw.loc[i][v])
            if not pd.isnull(to_draw.loc[i].parent_node):
                f=to_draw.loc[to_draw.node==to_draw.loc[i].parent_node, 'split_feature'].values[0]
                G.add_edge(to_draw.loc[i].parent_node, to_draw.loc[i].node,
                           label=f+ ':\n'+(str(to_draw.loc[i][f]) if len(str(to_draw.loc[i][f]))<=25 else str(to_draw.loc[i][f])[:25]+'..'))


        plt.figure(figsize=(figsize[0]*(1+margins[0]), figsize[1]*(1+margins[1])))

        if draw_leaves_by_group:
            to_draw['woe']=to_draw.apply(lambda row: row['group_woe'] if pd.isnull(row['group_woe'])==False else row['woe'], axis=1)
            to_draw['amount']=to_draw.apply(lambda row: row['group_amount'] if pd.isnull(row['group_amount'])==False else row['amount'], axis=1)

        norm_color = colors.Normalize(to_draw['woe'].min(), to_draw['woe'].max())
        node_color = [colors.rgb2hex(plt.cm.get_cmap(matplotlib.cm.RdYlGn)(norm_color(to_draw[to_draw['node']==node]['woe'].values[0]))) for node in G.nodes]
        norm_size = colors.Normalize(to_draw['amount'].min(), to_draw['amount'].max())
        node_size = [(norm_size(to_draw[to_draw['node']==node]['amount'].values[0])+0.1)*3000 for node in G.nodes]
        node_lines = [120*(norm_size(to_draw[to_draw['node']==node]['amount'].values[0])+0.45) if to_draw[to_draw['node']==node]['leaf'].values[0] else 0 for node in G.nodes]

        edge_labels = nx.get_edge_attributes(G,'label')
        pos=hierarchy_pos(G, 0)
        nx.draw(G, pos, with_labels=True,
                node_size=node_size,
                node_color=node_color, node_shape="o", alpha=0.7, linewidths=node_lines, font_weight='bold', font_size=15)
        nx.draw_networkx_edge_labels(G, pos, edge_labels = edge_labels, font_weight='normal')

        #display(pos)
        if draw_node_labels:
            nx.draw_networkx_labels(G, {x:(pos[x][0]+shift_node_labels[0], pos[x][1]+shift_node_labels[1]) for x in pos},
                                    node_labels, font_size=10)
        plt.margins(x=margins[0], y=margins[1])
        plt.show()



    def unite_nodes(self, input_df=None, nodes=None, alpha=None, woe_adjust=None):
        '''
        Unite nodes in the input DataFrame, containing tree description

        Parameters
        -----------
        input_df: an input DataFrame, containing tree description
        nodes: an array-like, containing nodes to unite
        alpha: regularization parameter, used in WoE calculation
        woe_adjust: adjustment parameter for group's ER: EventRate_i = (N_target_i+woe_adjust)/(N_nontarget_i+woe_adjust)

        Returns
        -----------
        a DataFrame containing tree description with united nodes
        '''

        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        if nodes is None:
            nodes=[]

        if input_df is None:
            to_self=True
            input_df=self.tree.copy()
        else:
            to_self=False

        if not self.check_unitability(input_df, nodes):
            return input_df

        tree_df=input_df.copy()
        features=list(tree_df.columns[:tree_df.columns.get_loc('node')])
        parent_node=tree_df[tree_df['node']==nodes[0]]['parent_node'].values[0]

        if tree_df[tree_df['parent_node']==parent_node].shape[0]==2:
            tree_df=self.prune(tree_df, parent_node)
        else:
            to_regroup=not(tree_df[tree_df['node'].isin(nodes)]['leaf'].all() and len(tree_df[tree_df['node'].isin(nodes)]['group'].unique())==1)

            nodes_to_delete=self.node_and_children(tree_df, nodes[0])+self.node_and_children(tree_df, nodes[1])
            nodes_to_delete.remove(nodes[0])

            tree_df.loc[tree_df['node']==nodes[0], 'leaf'] = True

            for f in features:
                condition1=tree_df.loc[tree_df['node']==nodes[0], f].values[0]
                condition2=tree_df.loc[tree_df['node']==nodes[1], f].values[0]
                if condition1!=condition2:
                    if isinstance(condition1, list) and isinstance(condition2, list):
                        tree_df.loc[tree_df['node']==nodes[0], f] = pd.Series([[x for x in condition1+condition2 if pd.isnull(x)] + \
                                                                              [x for x in condition1+condition2 if pd.isnull(x)==False]]*tree_df.shape[0])
                    elif isinstance(condition1, tuple)==False and pd.isnull(condition1) and isinstance(condition2, tuple):
                        tree_df.loc[tree_df['node']==nodes[0], f] = pd.Series([(condition2, np.nan)] *tree_df.shape[0])
                    elif isinstance(condition2, tuple)==False and pd.isnull(condition2) and isinstance(condition1, tuple):
                        tree_df.loc[tree_df['node']==nodes[0], f] = pd.Series([(condition1, np.nan)] *tree_df.shape[0])
                    elif isinstance(condition1, tuple) and isinstance(condition2, tuple) and pd.isnull(condition1[1]):
                        tree_df.loc[tree_df['node']==nodes[0], f] = pd.Series([((min(condition1[0][0], condition2[0]), \
                                                                     max(condition1[0][1], condition2[1])), np.nan)] *tree_df.shape[0])
                    elif isinstance(condition1, tuple) and isinstance(condition2, tuple) and pd.isnull(condition2[1]):
                        tree_df.loc[tree_df['node']==nodes[0], f] = pd.Series([((min(condition1[0], condition2[0][0]), \
                                                                     max(condition1[1], condition2[0][1])), np.nan)] *tree_df.shape[0])
                    elif isinstance(condition1, tuple) and isinstance(condition2, tuple):
                        tree_df.loc[tree_df['node']==nodes[0], f] = pd.Series([(min(condition1[0], condition2[0]), \
                                                                    max(condition1[1], condition2[1]))] *tree_df.shape[0])

            if to_regroup:
                for value in ['target', 'nontarget', 'amount']:
                    tree_df.loc[tree_df['node']==nodes[0], value] = tree_df.loc[tree_df['node']==nodes[0], value].values[0]+tree_df.loc[tree_df['node']==nodes[1], value].values[0]
                tree_df=tree_df[tree_df['node'].isin(nodes_to_delete)==False].reset_index(drop=True)
                good_bad=[tree_df.loc[tree_df['node']==0, 'nontarget'].values[0],
                          tree_df.loc[tree_df['node']==0, 'target'].values[0]]

                tree_df['er']=tree_df['target']/tree_df['amount']
                tree_df['woe']=np.log(((good_bad[1]/good_bad[0])*(alpha + tree_df['amount'])/ \
                                       (tree_df['amount']*((tree_df['target'] + woe_adjust)/(tree_df['nontarget'] + woe_adjust)) + alpha)).astype(float))
                tree_df=tree_df.drop(['group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'], axis=1)
                tree_df.loc[tree_df[tree_df['leaf']].sort_values('woe').index, 'group']=[i for i in range(tree_df[tree_df['leaf']].shape[0])]
                tree_df.loc[tree_df['leaf'], 'group_target']=tree_df.loc[tree_df['leaf'], 'target']
                tree_df.loc[tree_df['leaf'], 'group_nontarget']=tree_df.loc[tree_df['leaf'], 'nontarget']
                tree_df.loc[tree_df['leaf'], 'group_amount']=tree_df.loc[tree_df['leaf'], 'amount']
                tree_df.loc[tree_df['leaf'], 'group_woe']=tree_df.loc[tree_df['leaf'], 'woe']
            else:
                tree_df=tree_df[tree_df['node'].isin(nodes_to_delete)==False].reset_index(drop=True)
                tree_df.loc[tree_df['node']==nodes[0], 'target']=tree_df.loc[tree_df['node']==nodes[0], 'group_target']
                tree_df.loc[tree_df['node']==nodes[0], 'nontarget']=tree_df.loc[tree_df['node']==nodes[0], 'group_nontarget']
                tree_df.loc[tree_df['node']==nodes[0], 'amount']=tree_df.loc[tree_df['node']==nodes[0], 'group_amount']
                tree_df.loc[tree_df['node']==nodes[0], 'woe']=tree_df.loc[tree_df['node']==nodes[0], 'group_woe']

        if to_self:
            self.tree=tree_df.copy()
            self.parameters.update({'manual':True})
            _, self.ginis=GiniChecker().work_tree(self, out=True)
        else:
            return tree_df



    def unite_groups(self, input_df=None, groups=None, alpha=None, woe_adjust=None):
        '''
        Unite groups in the input DataFrame, containing tree description and recalculate 'group_' stats

        Parameters
        -----------
        input_df: an input DataFrame, containing tree description
        groups: an array-like, containing groups to unite
        alpha: regularization parameter, used in WoE calculation
        woe_adjust: adjustment parameter for group's ER: EventRate_i = (N_target_i+woe_adjust)/(N_nontarget_i+woe_adjust)

        Returns
        -----------
        a DataFrame containing tree description with united groups
        '''

        if woe_adjust is None and self.woe_adjust is not None:
            woe_adjust=self.woe_adjust
        if alpha is None and self.alpha is not None:
            alpha=self.alpha

        if groups is None:
            groups=[]

        if input_df is None:
            to_self=True
            input_df=self.tree.copy()
        else:
            to_self=False

        if (not isinstance(groups, list) and not isinstance(groups, tuple)):
            print('Not a list/tuple. Return None.')
            return None
        elif sum([x not in input_df['group'] for x in groups])>0:
            print('Not all groups are present in input candidate. Return None.')
            return None
        else:
            tree_df=input_df.copy()
            for g in sorted(groups)[1:]:
                tree_df.loc[tree_df['group']==g, 'group']=sorted(groups)[0]
            good_bad=[tree_df.loc[tree_df['node']==0, 'nontarget'].values[0],
                      tree_df.loc[tree_df['node']==0, 'target'].values[0]]
            group_stats=tree_df[tree_df['leaf']][['group', 'target', 'nontarget', 'amount']].groupby('group', as_index=False).sum().rename({'target':'group_target', 'nontarget':'group_nontarget', 'amount':'group_amount'}, axis=1)
            group_stats['group_woe']=np.log(((good_bad[1]/good_bad[0])*(alpha + group_stats['group_amount'])/ \
                                           (group_stats['group_amount']*((group_stats['group_target'] + woe_adjust)/(group_stats['group_nontarget'] + woe_adjust)) + alpha)).astype(float))
            tree_df=tree_df.drop(['group_target', 'group_nontarget', 'group_amount', 'group_woe'], axis=1).merge(group_stats, on=['group'], how='left')
            for i in range(len(groups)-1):
                tree_df.loc[tree_df['group']>sorted(groups)[i+1]-i, 'group']=tree_df[tree_df['group']>sorted(groups)[i+1]-i]['group']-1

            if to_self:
                self.tree=tree_df.copy()
                self.parameters.update({'manual':True})
                _, self.ginis=GiniChecker().work_tree(self, out=True)
            else:
                return tree_df



    def prune(self, input_df=None, nodes=None):
        '''
        Prune node/nodes in input DataFrame containing tree description

        Parameters
        -----------
        input_df: an input DataFrame, containing tree description
        nodes: a node or an array-like, containing nodes to be pruned

        Returns
        -----------
        a DataFrame containing tree description with pruned nodes
        '''
        if nodes is None:
            nodes=[]

        if input_df is None:
            to_self=True
            input_df=self.tree.copy()
        else:
            to_self=False

        tree_df=input_df.copy()
        if isinstance(nodes, list)==False and isinstance(nodes, tuple)==False:
            nodes=[nodes]
        to_regroup=False
        for node in nodes:
            if node in tree_df['node'].tolist():
                to_prune=self.node_and_children(tree_df, node)
                to_prune.remove(node)
                if len(tree_df[tree_df['parent_node']==node]['group'].unique())==1 and \
                    tree_df[tree_df['parent_node']==node]['leaf'].all():
                    tree_df.loc[tree_df['node']==node, 'group']=tree_df[tree_df['parent_node']==node]['group'].unique()
                    tree_df.loc[tree_df['node']==node, 'group_target']=tree_df[tree_df['parent_node']==node]['group_target'].unique()
                    tree_df.loc[tree_df['node']==node, 'group_nontarget']=tree_df[tree_df['parent_node']==node]['group_nontarget'].unique()
                    tree_df.loc[tree_df['node']==node, 'group_amount']=tree_df[tree_df['parent_node']==node]['group_amount'].unique()
                    tree_df.loc[tree_df['node']==node, 'group_woe']=tree_df[tree_df['parent_node']==node]['group_woe'].unique()
                else:
                    to_regroup=True
                tree_df.loc[tree_df['node']==node,'leaf']=True
                tree_df.loc[tree_df['node']==node,'split_feature']=np.nan
                tree_df=tree_df[tree_df['node'].isin(to_prune)==False].reset_index(drop=True)

        if to_regroup:
            tree_df=tree_df.drop(['group', 'group_target', 'group_nontarget', 'group_amount', 'group_woe'], axis=1)
            tree_df.loc[tree_df[tree_df['leaf']].sort_values('woe').index, 'group']=[i for i in range(tree_df[tree_df['leaf']].shape[0])]
            tree_df.loc[tree_df['leaf'], 'group_target']=tree_df.loc[tree_df['leaf'], 'target']
            tree_df.loc[tree_df['leaf'], 'group_nontarget']=tree_df.loc[tree_df['leaf'], 'nontarget']
            tree_df.loc[tree_df['leaf'], 'group_amount']=tree_df.loc[tree_df['leaf'], 'amount']
            tree_df.loc[tree_df['leaf'], 'group_woe']=tree_df.loc[tree_df['leaf'], 'woe']

        if to_self:
            self.tree=tree_df.copy()
            self.parameters.update({'manual':True})
            _, self.ginis=GiniChecker().work_tree(self, out=True)
        else:
            return tree_df



    def check_unitability(self, input_df, nodes, verbose=False):
        '''
        Check possibility of two nodes to be united into one. Considers parent node and conditions (for interval parent
        split feature conditions should have adjacent intervals or one of condition should correspond to missing values)

        Parameters
        -----------
        input_df: an input DataFrame, containing tree description
        nodes: an array-like containing nodes to unite
        verbose: should the reasons for False result be printed or not

        Returns
        -----------
        a DataFrame containing tree description with united groups
        '''
        if (not isinstance(nodes, list) and not isinstance(nodes, tuple)) or len(nodes)>2:
            if verbose:
                print('Not a list/tuple or length is greater then 2. Return input candidate.')
            return False
        elif sum([x not in input_df['node'].tolist() for x in nodes])>0:
            if verbose:
                print('Not all nodes are present in input candidate. Return input candidate.')
            return False
        elif input_df.loc[input_df['node']==nodes[0], 'parent_node'].values[0]!=input_df.loc[input_df['node']==nodes[1], 'parent_node'].values[0]:
            if verbose:
                print('Specified nodes are not sharing the same parent node. Return input candidate.')
            return False
        else:
            parent_node=input_df.loc[input_df['node']==nodes[0], 'parent_node'].values[0]
            parent_split_feature=input_df.loc[input_df['node']==parent_node, 'split_feature'].values[0]
            cond1=input_df.loc[input_df['node']==nodes[0], parent_split_feature].values[0]
            cond2=input_df.loc[input_df['node']==nodes[1], parent_split_feature].values[0]
            if isinstance(cond1, tuple) and isinstance(cond2, tuple):
                cond1=cond1[0] if pd.isnull(cond1[1]) else cond1
                cond2=cond2[0] if pd.isnull(cond2[1]) else cond2
                if not (cond1[0]==cond2[1] or cond1[1]==cond2[0]):
                    if verbose:
                        print('Specified nodes are not having adjacent intervals. Return input candidate.')
                    return False
        return True



    def node_and_children(self, tree_df, node):
        '''
        TECH
        Return input node and all its children from an input DataFrame, containing tree description

        Parameters
        -----------
        input_df: an input DataFrame, containing tree description
        node: a node to check for children

        Returns
        -----------
        a list, containing the initial node and all its children
        '''
        if node in tree_df['node'].tolist():
            children=[node]
            if tree_df.loc[tree_df['node']==node,'leaf'].values[0]==False:
                for child_node in tree_df[tree_df['parent_node']==node]['node']:
                    children=children+self.node_and_children(tree_df, child_node)
            return children
        else:
            return []



    def export_trees(self, out=None):
        '''
        Export tree candidates and their gini values in DataFrame and Excel file

        Parameters
        -----------
        out: a string with the path to export file (.xlsx or .xls extension required)

        Returns
        -----------
        a DataFrame with tree candidates and
        a DataFrame with gini values of candidates on different samples
        '''
        out_trees=pd.DataFrame(columns=['candidate', 'selected']+self.features+list(self.tree.columns[self.tree.columns.get_loc('node'):]))
        out_ginis=[]
        #pd.DataFrame(columns=['candidate', 'selected']+list(self.ginis))
        for i in self.trees_correct:
            to_append_gini=self.trees_ginis[i].copy()
            to_append_gini['candidate']=i
            to_append_gini['selected']=(self.trees_ginis[i]==self.ginis)
            if to_append_gini not in out_ginis:
                out_ginis.append(to_append_gini)

                to_append_tree=self.trees_correct[i].copy()
                for f in self.features:
                    if f not in to_append_tree:
                        to_append_tree[f]=None
                to_append_tree['candidate']=i
                to_append_tree['selected']=(self.trees_ginis[i]==self.ginis)
                out_trees=out_trees.append(to_append_tree, ignore_index=True)
        if self.ginis not in [self.trees_ginis[x] for x in self.trees_ginis]:
            to_append_gini=self.ginis.copy()
            to_append_gini['candidate']=max(x['candidate'] for x in out_ginis)+1
            to_append_gini['selected']=True
            out_ginis.append(to_append_gini)

            to_append_tree=self.tree.copy()
            for f in self.features:
                if f not in to_append_tree:
                    to_append_tree[f]=None
            to_append_tree['candidate']=out_trees['candidate'].max()+1
            to_append_tree['selected']=True
            out_trees=out_trees.append(to_append_tree, ignore_index=True)

        out_trees=out_trees[['candidate', 'selected']+self.features+list(self.tree.columns[self.tree.columns.get_loc('node'):])].dropna(how='all', axis=1)
        used_features=[x for x in self.features if x in out_trees]
        out_trees[used_features]=out_trees[used_features].astype(str)
        out_ginis=pd.DataFrame(out_ginis, columns=['candidate', 'selected']+list(self.ginis))
        if out is not None and isinstance(out, str):
            if out[-5:]=='.xlsx' or out[-4:]=='.xls':
                writer = pd.ExcelWriter(out, engine='openpyxl')
                #woes_values=woes_df[df_columns].values.reshape(-1,).tolist()
                #woes_df.style.apply(color_background,
                #                    mn=np.mean(woes_values)-2*np.std(woes_values),
                #                    mx=np.mean(woes_values)+2*np.std(woes_values),
                #                    cmap='RdYlGn', subset=df_columns,
                #                    high=out_woe_high, low=out_woe_low)
                out_trees.to_excel(writer, sheet_name='Trees', index=False)
                out_ginis.style.apply(color_background,
                                  mn=0,
                                  mx=out_ginis.max().max(),
                                  cmap='RdYlGn', subset=list(self.ginis),
                                  high=0, low=0).to_excel(writer, sheet_name='Gini by Samples', index=False)
                # Get the openpyxl objects from the dataframe writer object.
                for worksheet in writer.sheets:
                    for cn in range(1, worksheet.max_column+1):
                        cl = openpyxl.utils.get_column_letter(cn)
                        worksheet.column_dimensions[cl].width = 12
                    #for x in writer.sheets[worksheet].columns:
                    #    writer.sheets[worksheet].column_dimensions[x[0].column].width = 12#40 if x[0].column=='A' else 12
                writer.save()
            else:
                print('Unknown format for export file. Use .xlsx. Skipping export.')

        return out_trees, out_ginis



    def import_trees(self, data, candidates=None, replace=False, exclude_rest=False):
        '''
        Import a tree with its candidates from data, which can be a pandas.DataFrame or a string with
        the path to Excel file,  containing trees descriptions

        Parameters
        -----------
        data: a string with the path to import file or a pandas.DataFrame with trees
        replace: if there is a candidate with the importing candidate number in self should new candidate
            replace existing or be skipped
        exclude_rest: should the candidates, which were not imported, be deleted or not
        '''
        if isinstance(data, str):
            if data[-5:]=='.xlsx' or data[-4:]=='.xls':
                data=pd.read_excel(data)
            else:
                print('Unknown format of import file (use .xlsx or .xls). Abort.')
                return None

        if candidates is not None:
            for i in candidates:
                if i not in data['candidate']:
                    print('No', i, 'tree candidate in input data. Abort.')
                    return None
            data=data[data['candidate'].isin(candidates)].reset_index(drop=True)
        else:
            candidates=sorted(data['candidate'].unique().tolist())

        for candidate in candidates:
            if (candidate in self.trees_correct and replace) or candidate not in self.trees_correct:
                if candidate in self.trees_correct:
                    print('Replacing', candidate,'candidate ..')
                else:
                    print('Adding', candidate,'candidate ..')
                current_candidate=data[data['candidate']==candidate].dropna(how='all', axis=1).reset_index(drop=True)
                tech_vars=[]
                for tech_var in ['tree', 'candidate', 'selected']:
                    if tech_var in current_candidate:
                        tech_vars.append(tech_var)
                current_candidate=current_candidate.drop(tech_vars, axis=1)
                used_features=list(current_candidate.columns[:current_candidate.columns.get_loc('node')])
                current_candidate[used_features]=current_candidate[used_features]\
                                                    .applymap(lambda x: x if isinstance(x, str)==False \
                                                              else None if x=='None' else np.nan if x=='nan' else \
                                                                eval(x.replace('), nan)','), np.nan)').replace(', nan,',', np.nan,')\
                                                                      .replace('[nan,','[np.nan,').replace(', nan]',', np.nan]').replace('[nan]','[np.nan]')\
                                                                      .replace('(inf,','(np.inf,').replace(', inf)',', np.inf)')\
                                                                      .replace('(-inf,','(-np.inf,').replace(', -inf)',', -np.inf)')))
                current_candidate['parent_node']=current_candidate['parent_node'].astype(object)
                current_candidate.loc[pd.isnull(current_candidate['parent_node']), 'parent_node']=None
                self.trees_correct[candidate]=current_candidate.dropna(how='all', axis=1)
                self.trees_parameters[candidate]={'manual':True}
                _, self.trees_ginis[candidate]=GiniChecker().work_tree(self, self.trees_correct[candidate], out=True)
            else:
                print('Candidate', candidate, 'was not replaced (to import candidate set replace=True). Skipping..')

        if replace and data[data['selected']].shape[0]>0:
            self.tree_select(sorted(data[data['selected']]['candidate'].unique().tolist())[0])

        if exclude_rest:
            print('Excluding candidates not specified in input data:')
            candidates_to_check=list(self.trees_correct)
            for candidate in candidates_to_check:
                if candidate not in candidates:
                    print('.. dropping', candidate)
                    del self.trees_correct[candidate]
                    del self.trees_parameters[candidate]
                    del self.trees_ginis[candidate]




